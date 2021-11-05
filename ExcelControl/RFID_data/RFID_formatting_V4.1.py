import openpyxl
import pandas as pd
import xlrd
import os
import sys
import json
import requests
import wget

from datetime import date, datetime
from xml.etree.ElementTree import parse
from openpyxl_image_loader import SheetImageLoader
from openpyxl.styles import Font
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment
from openpyxl.styles import numbers
from shutil import copyfile
from urllib.parse import urlencode, quote_plus

START_DATE = date.today()
START_LINE = 0
END_LINE = 0

STYLE_BORDER = Border(left=Side(style='thin'),
                      right=Side(style='thin'),
                      top=Side(style='thin'),
                      bottom=Side(style='thin'))
STYLE_FONT = Font(size=12)
STYLE_ALIGN = Alignment(horizontal='center', vertical='top', wrap_text=True)
STYLE_ALIGN_LEFT = Alignment(horizontal='left', vertical='top', wrap_text=True)

SAMPLE_SHEET = "Data/sampledata211102.xlsx"
CLASSFICATION_DATA = "Data/classData210909.xml"

CUR_PATH = os.getcwd()
SRC_FILE = CUR_PATH + '\\raw.xlsx'
DST_PATH = CUR_PATH + '\\'

TOTAL_OUTPUT_FILE = DST_PATH + "totalResult.xlsx"
EACH_OUTPUT_FILE = DST_PATH + "eachResult.xlsx"

SORTING_KEY = "운용부서"


def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_path, relative_path)


def g2b_prd_call(prdNo):
    url = 'http://apis.data.go.kr/1230000/ThngListInfoService/getThngPrdnmLocplcAccotListInfoInfoPrdlstSearch'
    queryParams = '?' + urlencode({
        quote_plus('serviceKey'):
            'BE+Tsy32/R43W0b5Ab8rkDXpPkQyplKcrVghwj3OkFjNwiVp5J5FHXnhwm4eGBN583zA0/gRPpPpwsFa29n1Rg==',
        quote_plus('numOfRows'): '10',
        quote_plus('pageNo'): '1',
        quote_plus('prdctIdntNo'): prdNo,
        quote_plus('inqryBgnDt'): '199501010000',
        quote_plus('inqryEndDt'): '202212310000',
        quote_plus('type'): 'json'
    })
    try:
        req = url + queryParams
        reply = requests.get(req).text

        jData = json.loads(reply)
        if jData['response']['body']['totalCount'] > 0:
            data = jData['response']['body']['items'][0]['prdctImgLrge']
            res = True
        else:
            data = "Not in List"
            res = False
    except Exception as e:
        data = "Error"
        res = False
        write_log(f"Error;404;{data};{prdNo};{e}\n")
    finally:
        return data, res


def printProgress(iteration, total, prefix='', suffix='', decimals=1, barLength=100):
    formatStr = "{0:." + str(decimals) + "f}"
    percent = formatStr.format(100 * (iteration / float(total)))
    filledLength = int(round(barLength * iteration / float(total)))
    bar = '#' * filledLength + '-' * (barLength - filledLength)
    sys.stdout.write('\r%s |%s| %s%s %s' % (prefix, bar, percent, '%', suffix)),
    if iteration == total:
        sys.stdout.write('\n')
    sys.stdout.flush()


def download(url, out_path):
    try:
        wget.download(url, out=out_path)
    except Exception as e:
        write_log(f"Error;404;{url};{e}\n")
        # print(e)


def write_log(string):
    dat = str(START_DATE).replace("-", "")
    f = open(CUR_PATH + f'\\log_{dat}.txt', "a+")
    f.write(f'{string}\n')
    f.close()


def convert_xls_to_xlsx(xls_file_path):
    xlsBook = xlrd.open_workbook(xls_file_path)
    workbook = openpyxl.Workbook()

    for i in range(0, xlsBook.nsheets):
        xlsSheet = xlsBook.sheet_by_index(i)
        sheet = workbook.active if i == 0 else workbook.create_sheet()
        sheet.title = xlsSheet.name

        for row in range(0, xlsSheet.nrows):
            for col in range(0, xlsSheet.ncols):
                colvalue = xlsSheet.cell_value(row, col)
                if isinstance(colvalue, str):
                    colvalue = colvalue.replace('L', ' ', 3)

                sheet.cell(row=row + 1, column=col + 1).value = colvalue
    return workbook


def parse_xml(src_file, xmlroot, arg):
    tree = parse(src_file)
    root = tree.getroot()
    xml_root = root.findall(xmlroot)

    xml_val = list()
    for i in range(0, len(arg)):
        xml_val.append([x.findtext(arg[i]) for x in xml_root])

    return xml_val


if __name__ == "__main__":
    sort_key = 0
    while sort_key not in [1, 2]:
        sort_key = int(input("분류기준 (1 or 2)\n"
                             "1. 운용부서\n"
                             "2. 설치장소\n"
                             "Input Number : "))
    if sort_key == 1:
        SORTING_KEY = "운용부서"
    elif sort_key == 2:
        SORTING_KEY = "설치장소"

    create_key = 0
    while create_key not in [1, 2]:
        create_key = int(input("출력파일 (1 or 2)\n"
                               "1. 총괄 시트 파일\n"
                               "2. 부서별 시트 파일\n"
                               "Input Number : "))

    create_sol = create_key
    errChk = True
    filename = "None"
    if os.path.isfile(SRC_FILE):
        filename = SRC_FILE
    else:
        errChk = False
    if errChk:
        sample_file = resource_path(SAMPLE_SHEET)
        class_file = resource_path(CLASSFICATION_DATA)
        totalResult_file = TOTAL_OUTPUT_FILE
        eachResult_file = EACH_OUTPUT_FILE

        col_range = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']

        raw_wb = openpyxl.load_workbook(filename, data_only=True)
        raw_ws = raw_wb.active

        if START_LINE == 0:
            startRow = 2
        else:
            startRow = START_LINE

        if END_LINE == 0:
            endRows = raw_ws.max_row + 1
        else:
            endRows = END_LINE

        # print(endRows)

        for i in range(1, endRows):
            row = str(i)
            if raw_ws[f'C{row}'].value == "물품목록번호":
                # print(row)
                startRow = i + 1
                break

        columns = []
        for col in col_range:
            if col == 'D':
                columns.append('대분류')
                columns.append('중분류')
            columns.append(raw_ws[col+row].value)

        if startRow == 0:
            print("Error 1")

        # Classification
        xml_root = "classification"
        xml_src = ["classname", "midclass", "highclass"]
        parseData = parse_xml(class_file, xml_root, xml_src)

        classname = parseData[0]
        midclass = parseData[1]
        highclass = parseData[2]

        # Pre-Processing
        tbl = []
        for j in range(startRow, endRows):
            row = str(j)
            tup = []
            for k in col_range:
                val = k + row
                if k == 'D':
                    try:
                        idx = classname.index(raw_ws[val].value)
                        tup.append(highclass[idx])
                        tup.append(midclass[idx])
                    except Exception as e:
                        write_log(e)
                        tup.append("None")
                        tup.append("None")
                tup.append(raw_ws[val].value)

            tbl.append(tup)

        # Sorting Data
        sortData = pd.DataFrame(tbl, columns=columns)
        sortData = sortData.sort_values(by=[SORTING_KEY, '대분류', '중분류', '물품분류명'], ascending=[True, True, True, True])

        res = sortData.drop_duplicates([SORTING_KEY])[SORTING_KEY].tolist()

        # Main Processing
        if create_sol == 1:
            # ### TotalData Creating
            totalRes_wb = openpyxl.load_workbook(sample_file)
            totalRes_ws = totalRes_wb.active

            proc_ws = totalRes_wb.copy_worksheet(totalRes_ws)
            proc_ws.title = "총괄표"
            cnt1 = 0

            proc_ws['B1'].value = raw_ws['C1'].value
            totalData = sortData.values.tolist()

            # ### Empty Data Proc
            for data in totalData:
                if data[1] == None:
                    pass

                else:
                    data.insert(10, '')
                    data.insert(11, '')
                    data.insert(12, '')
                    data.insert(13, '')
                    data.insert(18, '')
                    data.insert(19, '')
                    data.insert(20, '')
                    data.insert(25, '')
                    proc_ws.append(data)

            # ### Product Expire Date
            for j in range(3, proc_ws.max_row + 1):
                row = str(j)
                startUsageDate = proc_ws[f'G{row}'].value
                usableDate = proc_ws[f'Q{row}'].value
                prodDate = startUsageDate.split(".")

                endUsableDateYear = int(prodDate[0]) + int(usableDate)
                endUsableDateMonth = int(prodDate[1])
                endUsableDateDay = int(prodDate[2])

                dt = datetime.today()

                if dt.year > endUsableDateYear:
                    resDate = True
                else:
                    resDate = False
                    if (dt.year == endUsableDateYear) and (dt.month > endUsableDateMonth):
                        resDate = True
                    else:
                        resDate = False
                        if (dt.year == endUsableDateYear) and (dt.month == endUsableDateMonth and dt.day > endUsableDateDay):
                            resDate = True
                        else:
                            resDate = False

                if resDate:
                    resDateData = "연수초과"
                else:
                    resDateData = "연수남음"

                proc_ws[f'X{row}'].value = f'{endUsableDateYear}.{endUsableDateMonth}.{endUsableDateDay}'
                proc_ws[f'Y{row}'].value = resDateData

            # ### Call G2B API for Image
            for i in range(3, proc_ws.max_row + 1):
                row = str(i)

                prdCls = proc_ws['B' + row].value
                if type(str()) == type(prdCls):
                    val = prdCls.split('-')[1]

                    if not os.path.isdir(resource_path("/img")):
                        os.mkdir(resource_path("/img"))

                    if not os.path.isfile(resource_path(f"/img/{val}.jpg")):
                        imgUrl, resVal = g2b_prd_call(val)
                        if resVal:
                            download(imgUrl, resource_path(f"/img/{val}.jpg"))

                    if os.path.isfile(resource_path(f"/img/{val}.jpg")):
                        img = openpyxl.drawing.image.Image(resource_path(f"/img/{val}.jpg"))
                        img.anchor = f'AA{row}'
                        img.width = 65
                        img.height = 65
                        proc_ws.add_image(img)

                    printProgress(i, proc_ws.max_row, 'Progress:', 'Complete ', 1, 50)

            ck1 = 1
            ck2 = 1
            # ### xlsx Template Apply
            # print(len(totalData))
            # for data in totalData:
            #     print(len(data))
            for row in range(3, len(totalData) + 15):
                proc_ws.row_dimensions[row].height = 50
                ck1 += 1
                # print(f"ck1 : {str(ck1)} / {str(row)}")
                for col in range(1, 28):
                    ck2 += 1
                    # print(f"ck2 : {str(ck2)}")
                    proc_ws.cell(row=row, column=col).font = STYLE_FONT
                    proc_ws.cell(row=row, column=col).border = STYLE_BORDER
                    proc_ws.cell(row=row, column=col).alignment = STYLE_ALIGN
                    if col in [15, 16]:
                        proc_ws.cell(row=row, column=col).number_format = numbers.BUILTIN_FORMATS[3]

                    if col in [5, 6, 19, 26]:
                        proc_ws.cell(row=row, column=col).alignment = STYLE_ALIGN_LEFT

            totalRes_wb.remove(totalRes_wb['Sample'])
            totalRes_wb.save(totalResult_file)
            totalRes_wb.close()
            # print("1. OK")

        elif create_sol == 2:
            # ### Each Sheet Processing
            eachRes_wb = openpyxl.load_workbook(sample_file)
            eachRes_ws = eachRes_wb.active
            prog = 0

            for ws in res:
                if ws == '':
                    title = "NaN"
                    sortList = sortData[sortData[SORTING_KEY].isnull()].values.tolist()
                elif str(type(ws)).find("str") != -1:
                    title = ws
                    sortList = sortData[sortData[SORTING_KEY] == ws].values.tolist()
                else:
                    title = "NaN"
                    sortList = sortData[sortData[SORTING_KEY].isnull()].values.tolist()

                proc_ws = eachRes_wb.copy_worksheet(eachRes_ws)
                proc_ws.title = title
                cnt = 0

                proc_ws['B1'].value = raw_ws['C1'].value
                proc_ws['E1'].value = title

                # ### Empty Data Proc
                for data in sortList:
                    if data[1] == None:
                        pass
                    else:
                        data.insert(10, '')
                        data.insert(11, '')
                        data.insert(12, '')
                        data.insert(13, '')
                        data.insert(18, '')
                        data.insert(19, '')
                        data.insert(20, '')
                        data.insert(25, '')
                        proc_ws.append(data)

                # ### Product Expire Date
                for j in range(3, proc_ws.max_row + 1):
                    row = str(j)
                    startUsageDate = proc_ws[f'G{row}'].value
                    usableDate = proc_ws[f'Q{row}'].value
                    prodDate = startUsageDate.split(".")

                    endUsableDateYear = int(prodDate[0]) + int(usableDate)
                    endUsableDateMonth = int(prodDate[1])
                    endUsableDateDay = int(prodDate[2])

                    dt = datetime.today()

                    if dt.year > endUsableDateYear:
                        resDate = True
                    else:
                        resDate = False
                        if (dt.year == endUsableDateYear) and (dt.month > endUsableDateMonth):
                            resDate = True
                        else:
                            resDate = False
                            if (dt.year == endUsableDateYear) and (
                                    dt.month == endUsableDateMonth and dt.day > endUsableDateDay):
                                resDate = True
                            else:
                                resDate = False

                    if resDate:
                        resDateData = "연수초과"
                    else:
                        resDateData = "연수남음"

                    proc_ws[f'X{row}'].value = f'{endUsableDateYear}.{endUsableDateMonth}.{endUsableDateDay}'
                    proc_ws[f'Y{row}'].value = resDateData

                # ### Call G2B API for Image
                for i in range(3, proc_ws.max_row + 1):
                    row = str(i)

                    prdCls = proc_ws['B' + row].value
                    if type(str()) == type(prdCls):
                        val = prdCls.split('-')[1]

                        if not os.path.isdir(resource_path("/img")):
                            os.mkdir(resource_path("/img"))

                        if not os.path.isfile(resource_path(f"/img/{val}.jpg")):
                            imgUrl, resVal = g2b_prd_call(val)
                            if resVal:
                                download(imgUrl, resource_path(f"/img/{val}.jpg"))

                        if os.path.isfile(resource_path(f"/img/{val}.jpg")):
                            img = openpyxl.drawing.image.Image(resource_path(f"/img/{val}.jpg"))
                            img.anchor = f'AA{row}'
                            img.width = 65
                            img.height = 65
                            proc_ws.add_image(img)

                # ### xlsx Template Apply
                for data in sortList:
                    for row in range(3, len(data) + 15):
                        proc_ws.row_dimensions[row].height = 50
                        for col in range(1, 28):
                            proc_ws.cell(row=row, column=col).font = STYLE_FONT
                            proc_ws.cell(row=row, column=col).border = STYLE_BORDER
                            proc_ws.cell(row=row, column=col).alignment = STYLE_ALIGN
                            if col in [15, 16]:
                                proc_ws.cell(row=row, column=col).number_format = numbers.BUILTIN_FORMATS[3]

                            if col in [5, 6, 19, 26]:
                                proc_ws.cell(row=row, column=col).alignment = STYLE_ALIGN_LEFT

                prog += 1
                printProgress(prog, len(res), 'Progress:', 'Complete ', 1, 50)

            eachRes_wb.remove(eachRes_wb['Sample'])
            eachRes_wb.save(eachResult_file)
            eachRes_wb.close()

        # print("2. OK")
    else:
        print("Error:invalid Filename")
        write_log("Error:invalid Filename")

    os.system("pause")