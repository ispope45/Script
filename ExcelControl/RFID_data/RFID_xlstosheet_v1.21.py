import openpyxl
import pandas as pd
import xlrd
import os
import sys

from datetime import date
from xml.etree.ElementTree import parse
from openpyxl_image_loader import SheetImageLoader
from openpyxl.styles import Font
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment
from shutil import copyfile

START_DATE = date.today()
START_LINE = 0
END_LINE = 0

STYLE_BORDER = Border(left=Side(style='thin'),
                      right=Side(style='thin'),
                      top=Side(style='thin'),
                      bottom=Side(style='thin'))
STYLE_FONT = Font(size=20)
STYLE_ALIGN = Alignment(horizontal='left', vertical='top', wrap_text=True)

CUR_PATH = os.getcwd()
SRC_FILE = CUR_PATH + '\\raw.xls'
SRC_FILE2 = CUR_PATH + '\\raw.xlsx'
DST_PATH = CUR_PATH + '\\'


def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_path, relative_path)


def printProgress(iteration, total, prefix='', suffix='', decimals=1, barLength=100):
    formatStr = "{0:." + str(decimals) + "f}"
    percent = formatStr.format(100 * (iteration / float(total)))
    filledLength = int(round(barLength * iteration / float(total)))
    bar = '#' * filledLength + '-' * (barLength - filledLength)
    sys.stdout.write('\r%s |%s| %s%s %s' % (prefix, bar, percent, '%', suffix)),
    if iteration == total:
        sys.stdout.write('\n')
    sys.stdout.flush()


def write_log(string):
    dat = str(START_DATE).replace("-", "")
    f = open(CUR_PATH + f'log_{dat}.txt', "a+")
    f.write(f'{string}\n')
    f.close()


def convert_xls_to_xlsx(xls_file_path):
    xlsBook = xlrd.open_workbook(xls_file_path)
    workbook = openpyxl.Workbook()

    for i in range(0, xlsBook.nsheets):
        xlsSheet = xlsBook.sheet_by_index(i)
        sheet = workbook.active if i == 0 else workbook.create_sheet()
        sheet.title = xlsSheet.name

        for row in range(0, xlsSheet.nrows):
            for col in range(0, xlsSheet.ncols):
                colvalue = xlsSheet.cell_value(row, col)
                if isinstance(colvalue, str):
                    colvalue = colvalue.replace('L', ' ', 3)

                sheet.cell(row=row + 1, column=col + 1).value = colvalue
    return workbook


def parse_xml(src_file, xmlroot, arg):
    tree = parse(src_file)
    root = tree.getroot()
    xml_root = root.findall(xmlroot)

    xml_val = list()
    for i in range(0, len(arg)):
        xml_val.append([x.findtext(arg[i]) for x in xml_root])

    return xml_val


if __name__ == "__main__":

    sch_name = input("INPUT [학교명] : ")

    errChk = True

    if os.path.isfile(SRC_FILE):
        filename = SRC_FILE
    elif os.path.isfile(SRC_FILE2):
        filename = SRC_FILE2
    else:
        errChk = False

    if errChk:
        sample_file = resource_path("Data/sampledata210909.xlsx")
        class_file = resource_path("Data/classData210909.xml")
        result_file = DST_PATH + "result.xlsx"

        if filename.find(".xlsx") == -1:
            try:
                load_wb = convert_xls_to_xlsx(filename)
            except Exception as e:
                write_log(e)
                b_filename = filename
                a_filename = filename.replace(".xls", ".xlsx")
                copyfile(b_filename, a_filename)
                filename = a_filename
                load_wb = openpyxl.load_workbook(filename, data_only=True)
        else:
            load_wb = openpyxl.load_workbook(filename, data_only=True)

        load_ws = load_wb.active

        if START_LINE == 0:
            startRow = 2
        else:
            startRow = START_LINE

        if END_LINE == 0:
            totalRows = load_ws.max_row + 1
        else:
            totalRows = END_LINE

        # Image Load(Extract)
        img_dict = {}
        image_load = SheetImageLoader(load_ws)
        for i in range(startRow, totalRows):
            row = str(i)
            if image_load.image_in(f'O{row}'):
                key = load_ws[f'A{row}'].value
                val = image_load.get(f'O{row}')
                img_dict[key] = val

        # Classification
        xml_root = "classification"
        xml_src = ["classname", "midclass", "highclass"]
        parseData = parse_xml(class_file, xml_root, xml_src)

        classname = parseData[0]
        midclass = parseData[1]
        highclass = parseData[2]

        load_ws.insert_cols(2)
        load_ws.insert_cols(2)
        load_ws['B1'].value = "대분류"
        load_ws['C1'].value = "중분류"

        for i in range(startRow, totalRows):
            row = str(i)
            try:
                idx = classname.index(load_ws['D' + row].value)

                load_ws['B' + row].value = midclass[idx]
                load_ws['C' + row].value = midclass[idx]
            except Exception as e:
                write_log(e)
                load_ws['B' + row].value = "None"
                load_ws['C' + row].value = "None"

        load_wb.save(resource_path("res.xlsx"))
        res_file = resource_path("res.xlsx")

        sortData = pd.read_excel(res_file)
        sortData = sortData.sort_values(by=['운용부서', '대분류', '중분류', '물품분류명'], ascending=[True, True, True, True])

        res = sortData.drop_duplicates(['운용부서'])['운용부서'].tolist()

        res_wb = openpyxl.load_workbook(sample_file)
        res_ws = res_wb.active
        prog = 0
        for ws in res:

            if str(type(ws)).find("str") != -1:
                title = ws
                sortList = sortData[sortData['운용부서'] == ws].values.tolist()
            else:
                title = "NaN"
                sortList = sortData[sortData['운용부서'].isnull()].values.tolist()

            proc_ws = res_wb.copy_worksheet(res_ws)
            proc_ws.title = title
            cnt = 0

            for data in sortList:
                cnt += 1
                data.insert(0, cnt)
                data.insert(10, '')
                data.insert(15, '')
                data.insert(16, '')
                data.insert(17, '')
                proc_ws.append(data)
                proc_ws['C1'].value = sch_name
                proc_ws['D1'].value = title

            for data in sortList:
                # print(data)
                # print(len(sortList))
                for row in range(3, len(sortList) + 15):
                    proc_ws.row_dimensions[row].height = 100
                    for col in range(1, 22):
                        proc_ws.cell(row=row, column=col).font = STYLE_FONT
                        proc_ws.cell(row=row, column=col).border = STYLE_BORDER
                        proc_ws.cell(row=row, column=col).alignment = STYLE_ALIGN
                        # print(f'{data[8]} / {str(col)} : {str(row)} ')

            for i in range(3, proc_ws.max_row + 1):
                row = str(i)
                if proc_ws[f'B{row}'].value in img_dict:
                    img = openpyxl.drawing.image.Image(img_dict[proc_ws[f'B{row}'].value])
                    img.anchor = f'V{row}'
                    img.width = 130
                    img.height = 130
                    proc_ws.add_image(img)

            prog += 1
            printProgress(prog, len(res), 'Progress:', 'Complete', 1, 50)

        os.system(f'del {resource_path("res.xlsx")}')

        res_wb.remove(res_wb['Sample'])
        res_wb.save(result_file)
    else:
        print("Error:invalid Filename")
        write_log("Error:invalid Filename")

    os.system("pause")