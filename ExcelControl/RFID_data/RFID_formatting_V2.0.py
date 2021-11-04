import openpyxl
import pandas as pd
import xlrd
import os
import sys
import json
import requests
import wget

from datetime import date
from xml.etree.ElementTree import parse
from openpyxl_image_loader import SheetImageLoader
from openpyxl.styles import Font
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment
from shutil import copyfile
from urllib.parse import urlencode, quote_plus

START_DATE = date.today()
START_LINE = 0
END_LINE = 0

STYLE_BORDER = Border(left=Side(style='thin'),
                      right=Side(style='thin'),
                      top=Side(style='thin'),
                      bottom=Side(style='thin'))
STYLE_FONT = Font(size=20)
STYLE_ALIGN = Alignment(horizontal='left', vertical='top', wrap_text=True)

CUR_PATH = os.getcwd()
SRC_FILE = CUR_PATH + '\\raw44.xlsx'
DST_PATH = CUR_PATH + '\\'


def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_path, relative_path)


def g2b_prd_call(prdNo):
    url = 'http://apis.data.go.kr/1230000/ThngListInfoService/getThngPrdnmLocplcAccotListInfoInfoPrdlstSearch'
    queryParams = '?' + urlencode({
        quote_plus('serviceKey'):
            'BE+Tsy32/R43W0b5Ab8rkDXpPkQyplKcrVghwj3OkFjNwiVp5J5FHXnhwm4eGBN583zA0/gRPpPpwsFa29n1Rg==',
        quote_plus('numOfRows'): '10',
        quote_plus('pageNo'): '1',
        quote_plus('prdctIdntNo'): prdNo,
        quote_plus('inqryBgnDt'): '199501010000',
        quote_plus('inqryEndDt'): '202110010000',
        quote_plus('type'): 'json'
    })
    try:
        req = url + queryParams
        reply = requests.get(req).text

        jData = json.loads(reply)
        if jData['response']['body']['totalCount'] > 0:
            data = jData['response']['body']['items'][0]['prdctImgLrge']
            res = True
        else:
            data = "Not in List"
            res = False
    except Exception as e:
        data = "Error"
        res = False
        write_log(f"Error;404;{data};{e}\n")
    finally:
        return data, res


def printProgress(iteration, total, prefix='', suffix='', decimals=1, barLength=100):
    formatStr = "{0:." + str(decimals) + "f}"
    percent = formatStr.format(100 * (iteration / float(total)))
    filledLength = int(round(barLength * iteration / float(total)))
    bar = '#' * filledLength + '-' * (barLength - filledLength)
    sys.stdout.write('\r%s |%s| %s%s %s' % (prefix, bar, percent, '%', suffix)),
    if iteration == total:
        sys.stdout.write('\n')
    sys.stdout.flush()


def download(url, out_path):
    try:
        wget.download(url, out=out_path)
    except Exception as e:
        write_log(f"Error;404;{url};{e}\n")
        # print(e)


def write_log(string):
    dat = str(START_DATE).replace("-", "")
    f = open(CUR_PATH + f'\\log_{dat}.txt', "a+")
    f.write(f'{string}\n')
    f.close()


def convert_xls_to_xlsx(xls_file_path):
    xlsBook = xlrd.open_workbook(xls_file_path)
    workbook = openpyxl.Workbook()

    for i in range(0, xlsBook.nsheets):
        xlsSheet = xlsBook.sheet_by_index(i)
        sheet = workbook.active if i == 0 else workbook.create_sheet()
        sheet.title = xlsSheet.name

        for row in range(0, xlsSheet.nrows):
            for col in range(0, xlsSheet.ncols):
                colvalue = xlsSheet.cell_value(row, col)
                if isinstance(colvalue, str):
                    colvalue = colvalue.replace('L', ' ', 3)

                sheet.cell(row=row + 1, column=col + 1).value = colvalue
    return workbook


def parse_xml(src_file, xmlroot, arg):
    tree = parse(src_file)
    root = tree.getroot()
    xml_root = root.findall(xmlroot)

    xml_val = list()
    for i in range(0, len(arg)):
        xml_val.append([x.findtext(arg[i]) for x in xml_root])

    return xml_val


if __name__ == "__main__":
    errChk = True
    filename = "None"
    if os.path.isfile(SRC_FILE):
        filename = SRC_FILE
    else:
        errChk = False
    if errChk:
        sample_file = resource_path("Data/sampledata210909.xlsx")
        class_file = resource_path("Data/classData210909.xml")
        result_file = DST_PATH + "result.xlsx"

        col_range = ['A', 'B', 'D', 'F', 'G', 'I', 'L', 'M', 'O', 'Q', 'R', 'S']
        # print(filename)?
        # if filename.find(".xlsx") == -1:
        #     try:
        #         raw_wb = convert_xls_to_xlsx(filename)
        #         # raw_wb.save(CUR_PATH + "\\test.xlsx")
        #     except Exception as e:
        #         write_log(e)
        #         b_filename = filename
        #         a_filename = filename.replace(".xls", ".xlsx")
        #         copyfile(b_filename, a_filename)
        #         filename = a_filename
        #         raw_wb = openpyxl.load_workbook(filename, data_only=True)
        # else:
        #     raw_wb = openpyxl.load_workbook(filename, data_only=True)

        raw_wb = openpyxl.load_workbook(filename, data_only=True)
        raw_ws = raw_wb.active

        if START_LINE == 0:
            startRow = 2
        else:
            startRow = START_LINE

        if END_LINE == 0:
            endRows = raw_ws.max_row - 5
        else:
            endRows = END_LINE

        for i in range(1, endRows):
            row = str(i)
            if raw_ws[f'A{row}'].value == "물품목록번호":
                # print(row)
                startRow = i + 1
                break

        columns = []
        for col in col_range:
            if col == 'B':
                columns.append('대분류')
                columns.append('중분류')
            columns.append(raw_ws[col+row].value)

        if startRow == 0:
            print("Error 1")

        # Classification
        xml_root = "classification"
        xml_src = ["classname", "midclass", "highclass"]
        parseData = parse_xml(class_file, xml_root, xml_src)

        classname = parseData[0]
        midclass = parseData[1]
        highclass = parseData[2]

        # Pre-Processing
        tbl = []
        for j in range(startRow, endRows):
            row = str(j)
            tup = []
            for k in col_range:
                val = k + row
                if k == 'B':
                    try:
                        idx = classname.index(raw_ws[val].value)
                        tup.append(highclass[idx])
                        tup.append(midclass[idx])
                    except Exception as e:
                        write_log(e)
                        tup.append("None")
                        tup.append("None")
                tup.append(raw_ws[val].value)

            tbl.append(tup)

        sortData = pd.DataFrame(tbl, columns=columns)
        sortData = sortData.sort_values(by=['운용부서', '대분류', '중분류', '물품분류명'], ascending=[True, True, True, True])

        res = sortData.drop_duplicates(['운용부서'])['운용부서'].tolist()
        # print(res)
        res_wb = openpyxl.load_workbook(sample_file)
        res_ws = res_wb.active

        # Main Processing
        prog = 0
        for ws in res:
            # print(type(ws))
            # print(ws)
            if ws == '':
                title = "NaN"
                sortList = sortData[sortData['운용부서'].isnull()].values.tolist()
            elif str(type(ws)).find("str") != -1:
                title = ws
                sortList = sortData[sortData['운용부서'] == ws].values.tolist()
            else:
                title = "NaN"
                sortList = sortData[sortData['운용부서'].isnull()].values.tolist()

            proc_ws = res_wb.copy_worksheet(res_ws)
            proc_ws.title = title
            cnt = 0

            for data in sortList:
                cnt += 1
                data.insert(0, cnt)
                data.insert(10, '')
                data.insert(15, '')
                data.insert(16, '')
                data.insert(17, '')
                proc_ws.append(data)
                # print(raw_ws.max_row-2)
                # print(raw_ws['P39'].value.split(" ")[1])
                # print(raw_ws[f'P{str(raw_ws.max_row-2)}'].value)
                proc_ws['A1'].value = raw_ws[f'P{str(raw_ws.max_row-2)}'].value
                proc_ws['D1'].value = title

            for data in sortList:
                for row in range(3, len(sortList) + 15):
                    proc_ws.row_dimensions[row].height = 100
                    for col in range(1, 23):
                        proc_ws.cell(row=row, column=col).font = STYLE_FONT
                        proc_ws.cell(row=row, column=col).border = STYLE_BORDER
                        proc_ws.cell(row=row, column=col).alignment = STYLE_ALIGN

            for i in range(3, proc_ws.max_row):
                row = str(i)

                prdCls = proc_ws['B' + row].value
                # print(type(prdCls))
                if type(str()) == type(prdCls):
                    # print(prdCls.split('-'))
                    val = prdCls.split('-')[1]
                    # print(val)

                    if not os.path.isdir(resource_path("/img")):
                        os.mkdir(resource_path("/img"))

                    imgUrl, resVal = g2b_prd_call(val)
                    # print(imgUrl)
                    if resVal:
                        if not os.path.isfile(resource_path(f"/img/{val}.jpg")):
                            # print(imgUrl)
                            download(imgUrl, resource_path(f"/img/{val}.jpg"))

                        if os.path.isfile(resource_path(f"/img/{val}.jpg")):
                            # print(resource_path(f"/img/{val}.jpg"))
                            img = openpyxl.drawing.image.Image(resource_path(f"/img/{val}.jpg"))
                            # print(img)
                            img.anchor = f'V{row}'
                            img.width = 130
                            img.height = 130
                            proc_ws.add_image(img)

            prog += 1
            printProgress(prog, len(res), 'Progress:', 'Complete', 1, 50)

        res_wb.remove(res_wb['Sample'])
        res_wb.save(result_file)
    else:
        print("Error:invalid Filename")
        write_log("Error:invalid Filename")

    os.system("pause")