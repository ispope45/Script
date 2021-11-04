import openpyxl
import pandas as pd
import xlrd
import os
import sys

from datetime import date
from xml.etree.ElementTree import parse
from openpyxl_image_loader import SheetImageLoader
from openpyxl.styles import Font
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment
from shutil import copyfile

START_DATE = date.today()
START_LINE = 0
END_LINE = 0

STYLE_BORDER = Border(left=Side(style='thin'),
                      right=Side(style='thin'),
                      top=Side(style='thin'),
                      bottom=Side(style='thin'))
STYLE_FONT = Font(size=10)
STYLE_ALIGN = Alignment(horizontal='left', vertical='top', wrap_text=True)

CUR_PATH = os.getcwd()
SRC_FILE = CUR_PATH + '\\tag.xlsx'
DST_PATH = CUR_PATH + '\\'


def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_path, relative_path)


def printProgress(iteration, total, prefix='', suffix='', decimals=1, barLength=100):
    formatStr = "{0:." + str(decimals) + "f}"
    percent = formatStr.format(100 * (iteration / float(total)))
    filledLength = int(round(barLength * iteration / float(total)))
    bar = '#' * filledLength + '-' * (barLength - filledLength)
    sys.stdout.write('\r%s |%s| %s%s %s' % (prefix, bar, percent, '%', suffix)),
    if iteration == total:
        sys.stdout.write('\n')
    sys.stdout.flush()


def write_log(string):
    dat = str(START_DATE).replace("-", "")
    f = open(CUR_PATH + f'log_{dat}.txt', "a+")
    f.write(f'{string}\n')
    f.close()


if __name__ == "__main__":

    errChk = True

    if os.path.isfile(SRC_FILE):
        filename = SRC_FILE
    else:
        filename = ""
        errChk = False

    if errChk:
        sample_file = resource_path("Data/sampleTagData.xlsx")
        result_file = DST_PATH + "resultTagData.xlsx"

        sortData = pd.read_excel(filename)
        sortData = sortData.sort_values(by=['실제설치장소'], ascending=[True])

        res = sortData.drop_duplicates(['실제설치장소'])['실제설치장소'].tolist()

        res_wb = openpyxl.load_workbook(sample_file)
        res_ws = res_wb.active

        prog = 0
        print(res)
        for ws in res:

            if str(type(ws)).find("str") != -1:
                title = ws
                sortList = sortData[sortData['실제설치장소'] == ws].values.tolist()
            else:
                title = "NaN"
                sortList = sortData[sortData['실제설치장소'].isnull()].values.tolist()

            proc_ws = res_wb.copy_worksheet(res_ws)
            proc_ws.title = title

            proc_ws['A1'].value = title

            result = {}
            for data in sortList:
                val = []
                if str(type(data[21])).find("str") != -1:
                    data[4] = data[21]

                if data[4] in result:
                    result[data[4]][0] += data[13]
                else:
                    val.append(data[13])
                    val.append(data[3])
                    result[data[4]] = val

            print(result)
            r = 4
            for attr in result:
                for row in range(4, len(result) + 4):
                    for col in range(1, 7):
                        proc_ws.cell(row=row, column=col).font = STYLE_FONT
                        proc_ws.cell(row=row, column=col).border = STYLE_BORDER
                        proc_ws.cell(row=row, column=col).alignment = STYLE_ALIGN

                proc_ws[f'A{str(r)}'].value = result[attr][1]
                proc_ws[f'B{str(r)}'].value = attr
                proc_ws[f'C{str(r)}'].value = result[attr][0]
                r += 1

        prog += 1
        printProgress(prog, len(res), 'Progress:', 'Complete', 1, 50)

        res_wb.remove(res_wb['Sample'])
        res_wb.save(result_file)
    else:
        print("Error:invalid Filename")
        write_log("Error:invalid Filename")

    os.system("pause")
