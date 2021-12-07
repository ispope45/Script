import openpyxl
import os

from datetime import date, datetime

from openpyxl.styles import Font
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment
from openpyxl.styles import numbers


if __name__ == "__main__":
    STYLE_BORDER = Border(left=Side(style='thin'),
                          right=Side(style='thin'),
                          top=Side(style='thin'),
                          bottom=Side(style='thin'))
    STYLE_FONT = Font(size=12)
    STYLE_ALIGN = Alignment(horizontal='center', vertical='top', wrap_text=True)
    STYLE_ALIGN_LEFT = Alignment(horizontal='left', vertical='top', wrap_text=True)

    SAMPLE_SHEET = "Data/mergesampledata211126.xlsx"

    CUR_PATH = os.getcwd() + '\\'
    SRC_FILE = CUR_PATH + 'merge_src.xlsx'
    OUTPUT_FILE = CUR_PATH + "mergeResult.xlsx"

    COL_LIST = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L',
                'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']



    resWb = openpyxl.load_workbook(SAMPLE_SHEET)
    resWs = resWb.active
    resWs.title = "총괄표"

    procWb = openpyxl.load_workbook(SRC_FILE)
    procWsNum = len(procWb.sheetnames)

    resRowNum = 2

    for i in range(0, procWsNum):
        procWs = procWb[procWb.sheetnames[i]]
        endNum = procWs.max_row + 1
        for j in range(3, endNum):
            row = str(j)
            if procWs[f'A{row}'].value is None:
                print("Hit")
                break

            resRow = str(resRowNum)
            for k in range(0, len(COL_LIST)):
                resWs[COL_LIST[k] + resRow].value = procWs[COL_LIST[k] + row].value
                print(COL_LIST[k] + row + " > " + COL_LIST[k] + resRow)

            resRowNum += 1

    for row in range(2, resRowNum + 1):
        resWs.row_dimensions[row].height = 50
        for col in range(1, 27):
            resWs.cell(row=row, column=col).font = STYLE_FONT
            resWs.cell(row=row, column=col).border = STYLE_BORDER
            resWs.cell(row=row, column=col).alignment = STYLE_ALIGN
            if col in [15, 16]:
                resWs.cell(row=row, column=col).number_format = numbers.BUILTIN_FORMATS[3]

            if col in [5, 6, 19, 26]:
                resWs.cell(row=row, column=col).alignment = STYLE_ALIGN_LEFT

    procWb.close()
    resWb.save(OUTPUT_FILE)
    resWb.close()

