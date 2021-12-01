import openpyxl
import os
import sys

# CUR_PATH = os.getcwd()
CUR_PATH = 'C:\\Users\\work.Jungly\\Desktop\\'
SRC_FILE = CUR_PATH + '\\raw.xlsx'
DST_PATH = CUR_PATH + '\\'

OUTPUT_FILE = DST_PATH + "Result.xlsx"

START_LINE = 0
END_LINE = 0

EXTRA_COMMENT = (
    '---------------- 추가설정(직접입력) ---------------- \n\n'
    'conf terminal\n'
    'logging 172.28.228.78\n'
    'snmp-server enable traps\n'
    'snmp-server host 172.28.228.78 version 3 priv sen_class\n'
    'snmp-server group sen_class v3 priv\n'
    'snmp-server user sen_class sen_class v3\n'
    'sha\n'
    'sen_key02\n'
    'sen_key02\n'
    'aes\n'
    'sen_enc02\n'
    'sen_enc02\n\n'

    'write memory\n'
    'y\n\n'
    '---------------- 계정정보(직접입력) ---------------- \n\n'
    '기본값 : root // [패스워드없음]\n'
    '** root 입력시 계정재생성 메세지출력됨\n'
    '변경값 : admin // frontier1!\n\n')


def printProgress(iteration, total, prefix='', suffix='', decimals=1, barLength=100):
    formatStr = "{0:." + str(decimals) + "f}"
    percent = formatStr.format(100 * (iteration / float(total)))
    filledLength = int(round(barLength * iteration / float(total)))
    bar = '#' * filledLength + '-' * (barLength - filledLength)
    sys.stdout.write('\r%s |%s| %s%s %s' % (prefix, bar, percent, '%', suffix)),
    if iteration == total:
        sys.stdout.write('\n')
    sys.stdout.flush()


if __name__ == "__main__":
    raw_wb = openpyxl.load_workbook(SRC_FILE, data_only=True)
    raw_ws = raw_wb.active

    if START_LINE == 0:
        startRow = 2
    else:
        startRow = START_LINE

    if END_LINE == 0:
        endRows = raw_ws.max_row + 1
    else:
        endRows = END_LINE

    proc_wb = openpyxl.Workbook()
    proc_ws = proc_wb.active

    procRow = 1
    # ip_d = 0

    proc_ws[f'A1'].value = "No"
    proc_ws[f'B1'].value = "schoolName"
    proc_ws[f'C1'].value = "executionName"
    proc_ws[f'D1'].value = "hostName"
    proc_ws[f'E1'].value = "deviceIp"
    # proc_ws[f'F1'].value = "ipAddr"
    # proc_ws[f'G1'].value = "netMask"
    # proc_ws[f'H1'].value = "gatewayIp"

    for row in range(startRow, endRows):
        no = str(raw_ws[f'A{row}'].value)
        execuName = raw_ws[f'C{row}'].value
        name = raw_ws[f'D{row}'].value
        deviceHostname = raw_ws[f'E{row}'].value
        l_poeCnt = raw_ws[f'H{row}'].value
        deviceIp_a = raw_ws[f'P{row}'].value
        deviceIp_b = raw_ws[f'Q{row}'].value
        deviceIp_c = raw_ws[f'R{row}'].value
        deviceIp_d = raw_ws[f'S{row}'].value
        deviceNetmask = str(raw_ws[f'M{row}'].value)
        deviceGateway = raw_ws[f'N{row}'].value

        for i in range(0, 5):
            deviceIp = f'{str(deviceIp_a)}.{str(deviceIp_b)}.{str(deviceIp_c)}.{str(deviceIp_d + i)}'
            if l_poeCnt + (i + 1) < 10:
                poeNo = '0' + str(l_poeCnt + (i + 1))
            else:
                poeNo = str(l_poeCnt + (i + 1))

            script = (
                f'en\n'
                f'config terminal\n'
                f'hostname {deviceHostname}_PoE{poeNo}\n'
                f'service ssh\n'
                f'ip ssh port 9992\n'
                f'interface vlan 1\n'
                f'ip address {deviceIp}/{deviceNetmask}\n'
                f'exit\n\n'
                
                f'ip route 0.0.0.0/0 {deviceGateway}\n'
                f'end\n\n'
                f'write memory\n'
                f'y\n\n')

            script += EXTRA_COMMENT

            savePath = DST_PATH + execuName
            if not(os.path.isdir(savePath)):
                os.makedirs(savePath)

            f = open(savePath + f'\\{no}_{name}_PoE{poeNo}.txt', "w+")
            f.write(script)
            f.close()
            procRow += 1
            p_row = str(procRow)

            proc_ws[f'A{p_row}'].value = no
            proc_ws[f'B{p_row}'].value = name
            proc_ws[f'C{p_row}'].value = execuName
            proc_ws[f'D{p_row}'].value = f'{deviceHostname}_PoE{poeNo}'
            proc_ws[f'E{p_row}'].value = f"{deviceIp}/{deviceNetmask}"
            # proc_ws[f'F{p_row}'].value = "ipAddr"
            # proc_ws[f'G{p_row}'].value = "netMask"
            # proc_ws[f'H{p_row}'].value = "gatewayIp"
            # proc_ws[f'I{p_row}'].value = "gatewayIp"


        printProgress(row, endRows, 'Progress:', 'Complete', 1, 50)
    proc_wb.save(OUTPUT_FILE)
