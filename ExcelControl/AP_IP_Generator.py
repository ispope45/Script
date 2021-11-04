import openpyxl
import os
import sys

CUR_PATH = os.getcwd()
SRC_FILE = CUR_PATH + '\\raw.xlsx'
DST_PATH = CUR_PATH + '\\'

OUTPUT_FILE = DST_PATH + "Result.xlsx"

START_LINE = 0
END_LINE = 0


def printProgress(iteration, total, prefix='', suffix='', decimals=1, barLength=100):
    formatStr = "{0:." + str(decimals) + "f}"
    percent = formatStr.format(100 * (iteration / float(total)))
    filledLength = int(round(barLength * iteration / float(total)))
    bar = '#' * filledLength + '-' * (barLength - filledLength)
    sys.stdout.write('\r%s |%s| %s%s %s' % (prefix, bar, percent, '%', suffix)),
    if iteration == total:
        sys.stdout.write('\n')
    sys.stdout.flush()


if __name__ == "__main__":
    raw_wb = openpyxl.load_workbook(SRC_FILE, data_only=True)
    raw_ws = raw_wb.active

    if START_LINE == 0:
        startRow = 2
    else:
        startRow = START_LINE

    if END_LINE == 0:
        endRows = raw_ws.max_row + 1
    else:
        endRows = END_LINE

    proc_wb = openpyxl.Workbook()
    proc_ws = proc_wb.active

    rowNo = 0
    rowCnt = 0
    # ip_d = 0

    proc_ws[f'A1'].value = "No"
    proc_ws[f'B1'].value = "schoolName"
    proc_ws[f'C1'].value = "labelName"
    proc_ws[f'D1'].value = "hostName"
    proc_ws[f'E1'].value = "networkId"
    proc_ws[f'F1'].value = "ipAddr"
    proc_ws[f'G1'].value = "netMask"
    proc_ws[f'H1'].value = "gatewayIp"

    for row in range(startRow, endRows):
        networkId = str(raw_ws[f'G{row}'].value) + "/" + str(raw_ws[f'H{row}'].value)
        gatewayIp = raw_ws[f'L{row}'].value
        netMask = raw_ws[f'H{row}'].value
        schoolName = raw_ws[f'A{row}'].value

        ip_a = raw_ws[f'I{row}'].value
        ip_b = raw_ws[f'J{row}'].value
        ip_c = raw_ws[f'K{row}'].value

        cnt = 0
        cnt2 = 0

        ap2Cnt = raw_ws[f'D{row}'].value
        ap1Cnt = raw_ws[f'C{row}'].value

        start1Ip = raw_ws[f'M{row}'].value
        end1Ip = raw_ws[f'N{row}'].value
        start2Ip = raw_ws[f'O{row}'].value
        end2Ip = raw_ws[f'P{row}'].value

        ip_d = start2Ip

        for apNo in range(ap1Cnt + 1, ap2Cnt + ap1Cnt + 1):
            rowCnt += 1
            rowNo = rowCnt + 1
            if apNo < 10:
                apNumber = "0" + str(apNo)
            else:
                apNumber = str(apNo)
            hostName = raw_ws[f'F{row}'].value + "_AP" + apNumber
            labelName = raw_ws[f'A{row}'].value + "_AP" + apNumber

            if start2Ip != 0 and (ip_d in range(start2Ip, end2Ip)):
                ip_d = start2Ip + cnt2
                cnt2 += 1
            else:
                ip_d = start1Ip + cnt
                cnt += 1

            ipAddr = f'{ip_a}.{ip_b}.{ip_c}.{ip_d}'
            proc_ws[f'A{rowNo}'].value = rowCnt
            proc_ws[f'B{rowNo}'].value = schoolName
            proc_ws[f'C{rowNo}'].value = labelName
            proc_ws[f'D{rowNo}'].value = hostName
            proc_ws[f'E{rowNo}'].value = networkId
            proc_ws[f'F{rowNo}'].value = ipAddr
            proc_ws[f'G{rowNo}'].value = netMask
            proc_ws[f'H{rowNo}'].value = gatewayIp

        printProgress(row, endRows, 'Progress:', 'Complete', 1, 50)
    proc_wb.save(OUTPUT_FILE)
