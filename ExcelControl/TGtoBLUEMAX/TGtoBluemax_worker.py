import openpyxl
import os
import sys
import pandas as pd
from datetime import date

START_DATE = date.today()

PG_NAME = "TGtoBluemax"

SRC_DIR = os.getcwd() + "//"
DST_FILE = SRC_DIR + "result.xlsx"
COL_LIST = ['PRIORITY', 'ENABLED', 'SRC', 'SRC ADDR', 'DST', 'DST ADDR', 'SVC', 'SVC SPEC',
            'ACTION', 'REVERSIBLE']

SAMPLE_FILE = "SRC/BLUEMAX_2.5.3_pol.xlsx"
SERVICE_FILE = "SRC/obj.csv"

# exceptionAddrName = ['172.28.228.101_NiaAdm', '172.28.228.102_NiaAdm', '172.28.228.103_NiaAdm',
#                      '172.28.228.104_NiaAdm', '172.28.228.105_NiaAdm', '172.28.228.106_NiaAdm',
#                      '172.28.228.107_NiaAdm', '172.28.228.108_NiaAdm', '172.28.228.109_NiaAdm',
#                      '172.28.228.110_NiaAdm', '172.28.228.11_NiaAuthSvr1', '172.28.228.21_NiaAuthSvr2',
#                      '172.28.228.77_wNMS1', '172.28.228.78_wNMS2', '172.28.228.79_wNMS3',
#                      '172.28.228.80_wNMS4', '172.28.228.71_SODE', '172.28.228.70_WNMS',
#                      '172.28.228.70_SODE', '172.28.228.78_wNMS1', '172.28.228.79_wNMS1',
#                      '172.28.228.80_wNMS1', '172.28.228.21_NiaAuthSvr1', '72.28.228.70_WNMS',
#                      '192.168.72.82_NiaEMS1', '192.168.72.83_NiaEMS2', '192.168.72.82_NiaEMS2']
#
# excpAddrName = ['test248249', 'Cloud_Server_195-196', 'Cloud_Server_13-40', '클라우드', '207.189.104.86',
#                 'Cloud_Server_197-200', '클라우드VM_1', '클라우드VM_2', '클라우드VM_3', '클라우드VM_4', 'Cloud_IP_192',
#                 'Cloud_IP_194', 'Cloud_IP_196', 'Cloud_IP_198']
#
# excpAddr = ['3.3.3.0/24', '192.168.72.248-192.168.72.249', '10.197.1.195-10.197.1.196/0',
#             '10.197.1.13-10.197.1.40/32', '10.197.1.195-10.197.1.196/32', '207.189.104.86', '10.192.0.0/16',
#             '10.194.0.0/16', '10.196.0.0/16', '10.198.0.0/16', '10.197.1.197-10.197.1.200/32',
#             '10.197.1.197-10.197.1.200/0']
#
# excpPort = ['VIRUS_TCP_PORT_1', 'VIRUS_UDP_PORT_4', 'Virus_TCP_Port_1', 'Virus_UDP_Port_1']
#

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_path, relative_path)


def printProgress(iteration, total, prefix='', suffix='', decimals=1, barLength=100):
    formatStr = "{0:." + str(decimals) + "f}"
    percent = formatStr.format(100 * (iteration / float(total)))
    filledLength = int(round(barLength * iteration / float(total)))
    bar = '#' * filledLength + '-' * (barLength - filledLength)
    sys.stdout.write('\r%s |%s| %s%s %s' % (prefix, bar, percent, '%', suffix)),
    if iteration == total:
        sys.stdout.write('\n')
    sys.stdout.flush()


def write_log(string):
    dat = str(START_DATE).replace("-", "")
    f = open(SRC_DIR + f'{PG_NAME}_log_{dat}.txt', "a+")
    f.write(f'{string}\n')
    f.close()


if __name__ == "__main__":
    print("Preparing TG to Bluemax Migration Tool....")

    fileList = os.listdir(SRC_DIR)
    BLMSvcList = []
    try:
        svcFile = pd.read_csv(resource_path(SERVICE_FILE), engine='python')
        svcData = svcFile['Object']
        BLMSvcList = svcData.values.tolist()

    except Exception as e:
        write_log(f'svcFile;{e}')
    #
    # print(svcList)
    # print(type(svcList))

    prog = 0
    for f in fileList:
        if f.find(".csv") == -1:
            continue

        res_wb = openpyxl.load_workbook(resource_path(SAMPLE_FILE))
        res_ws = res_wb.active

        proc_wb = openpyxl.Workbook()
        proc_ws = proc_wb.active
        proc_ws.append(COL_LIST)
        try:
            fileNo = f.split("_")[0]
            fileOrg = f.split("_")[1]
            fileSch = f.split("_")[2].split(".")[0]

            data = pd.read_csv(SRC_DIR + f, engine='python', encoding='cp949')
            p_data = data[COL_LIST]
            totalValue = p_data.values.tolist()
        except Exception as e:
            write_log(f'{f};{e}')
            continue

        polItem = []
        ruleList = []
        srcList = []
        dstList = []
        svcList = []
        reversibleCnt = 0
        ruleCount = 1

        # print(totalValue)
        for val in totalValue:
            proc_ws.append(val)
            # print(val)

        for row in range(2, proc_ws.max_row + 1):

            if not(pd.isna(proc_ws[f'A{row}'].value)):
                ruleList.append(ruleCount)
                ruleCount += 1
                ruleList.append(proc_ws[f'B{row}'].value)
                ruleList.append(proc_ws[f'I{row}'].value)
                ruleList.append(proc_ws[f'J{row}'].value)

            if not(pd.isna(proc_ws[f'C{row}'].value)):
                srcList.append([proc_ws[f'C{row}'].value, proc_ws[f'D{row}'].value])

            if not(pd.isna(proc_ws[f'E{row}'].value)):
                dstList.append([proc_ws[f'E{row}'].value, proc_ws[f'F{row}'].value])

            if not(pd.isna(proc_ws[f'G{row}'].value)):
                svcList.append([proc_ws[f'G{row}'].value, proc_ws[f'H{row}'].value])

            if not(pd.isna(proc_ws[f'A{row + 1}'].value)):
                ruleList.append(srcList)
                ruleList.append(dstList)
                ruleList.append(svcList)
                polItem.append(ruleList)

                if ruleList[3] == 'Yes':
                    ruleList = []
                    ruleList.append(ruleCount)
                    ruleCount += 1
                    ruleList.append(polItem[-1][1])
                    ruleList.append(polItem[-1][2])
                    ruleList.append(polItem[-1][3])

                    ruleList.append(dstList)
                    ruleList.append(srcList)
                    ruleList.append(svcList)
                    polItem.append(ruleList)
                ruleList = []
                srcList = []
                dstList = []
                svcList = []

        ruleList.append(srcList)
        ruleList.append(dstList)
        ruleList.append(svcList)
        polItem.append(ruleList)
        ruleList = []
        srcList = []
        dstList = []
        svcList = []

        # print(polItem)
        # print(len(polItem))

        res_ws_row = 4
        # for a in polItem:
        #     print(a)

        svcChkList = []
        wireless21 = []
        deleteList = []
        ruleNum = 0
        #
        # for valid in polItem:
        #     # print(valid)
        #     isExcept = False
        #
        #     for val1 in valid[4]:
        #         # print(val1[0])
        #         if val1[0] in exceptionAddrName:
        #             isExcept = True
        #             wireless21 = valid[5][0]
        #
        #         if val1[0] in excpAddrName:
        #             isExcept = True
        #
        #         if val1[1] in excpAddr:
        #             isExcept = True
        #
        #     for val2 in valid[5]:
        #         if val2[0] in exceptionAddrName:
        #             isExcept = True
        #             wireless21 = valid[5][0]
        #
        #         if val2[0] in excpAddrName:
        #             isExcept = True
        #
        #         if val2[1] in excpAddr:
        #             isExcept = True
        #
        #     for val3 in valid[6]:
        #         # print(val3)
        #         if val3[0] in excpPort:
        #             # print(val3[0])
        #             isExcept = True
        #
        #     if isExcept:
        #         deleteList.append(valid)
        #
        # for dList in deleteList:
        #     polItem.remove(dList)
        #
        # if wireless21:
        nms_cfg = [
            [1, 'Yes', 'Allow', 'No', [],
             [['UTM_RIP1', '1.1.1.1/32']],
             [['PING', 'icmp type=8 code=0']]],\
            [2, 'Yes', 'Allow', 'No', [
                ['SEN_NMS', '192.168.75.219-192.168.75.230']],
             [['UTM_RIP1', '1.1.1.1/32']],
             [['SNMP', 'udp 1-65535 161-161']]]]
        # print(type(nia_cfg))
        polItem.insert(0, nms_cfg[1])
        polItem.insert(0, nms_cfg[0])

        # print(polItem)

        for pol in polItem:
            # print(pol)

            '''
            [3.0, 'Yes', 'Deny', 'No', 
                [
                    ['34.78.211.173', '34.78.211.173/32'], 
                    ['94.99.216.4', '94.99.216.4/32'],
                    ['192.168.0.0/24', '192.168.0.0/24']
                ],
                [
                    ['192.168.0.0/24', '192.168.0.0/24'], 
                    ['all', '0.0.0.0/0']
                ],
                [
                    ['Sharing_TCP_file', 'tcp 1-65535 137-139']
                ]
            ]
            '''
            ruleNum += 1

            ruleCnt = []
            ruleCnt.append(len(pol[4]))
            ruleCnt.append(len(pol[5]))
            ruleCnt.append(len(pol[6]))

            res_ws[f'B{res_ws_row}'].value = 'default'
            if pol[1] == 'Yes':
                res_ws[f'C{res_ws_row}'].value = 1
            else:
                res_ws[f'C{res_ws_row}'].value = 0

            res_ws[f'E{res_ws_row}'].value = pol[2]

            # print(max(ruleCnt))
            for i in range(res_ws_row, res_ws_row + max(ruleCnt)):
                # res_ws[f'A{i}'].value = pol[0]
                res_ws[f'A{i}'].value = ruleNum

            row = res_ws_row
            for src in pol[4]:
                if src[1] == '0.0.0.0/0':
                    continue

                res_ws[f'J{row}'].value = src[0]
                srcCk = src[1].split('/')

                ipOctet = srcCk[0].split('.')
                if int(ipOctet[0]) == 10:
                    res_ws[f'F{row}'].value = 1
                    if int(ipOctet[1]) % 10 in [8, 0]:
                        res_ws[f'F{row}'].value = 2
                else:
                    if int(ipOctet[0]) in [218] and (int(ipOctet[1]) in [48]):
                        res_ws[f'F{row}'].value = 1
                    else:
                        res_ws[f'F{row}'].value = 2

                if srcCk[0].find("-") != -1:
                    res_ws[f'I{row}'].value = 'N'
                    res_ws[f'K{row}'].value = srcCk[0]
                elif int(srcCk[1]) < 32:
                    res_ws[f'I{row}'].value = 'N'
                    res_ws[f'K{row}'].value = f'{srcCk[0]}/{srcCk[1]}'
                else:
                    res_ws[f'I{row}'].value = 'H'
                    res_ws[f'K{row}'].value = srcCk[0]

                row += 1
                # print(src)

            row = res_ws_row
            for dst in pol[5]:
                if dst[1] == '0.0.0.0/0':
                    continue

                res_ws[f'P{row}'].value = dst[0]
                dstCk = dst[1].split('/')

                ipOctet = dstCk[0].split('.')
                if int(ipOctet[0]) == 10:
                    res_ws[f'L{row}'].value = 1
                    if int(ipOctet[1]) % 10 in [8]:
                        res_ws[f'L{row}'].value = 2
                else:
                    if int(ipOctet[0]) in [218] and (int(ipOctet[1]) in [48]):
                        res_ws[f'L{row}'].value = 1
                    else:
                        res_ws[f'L{row}'].value = 2

                if dstCk[0].find("-") != -1:
                    res_ws[f'O{row}'].value = 'N'
                    res_ws[f'Q{row}'].value = dstCk[0]
                elif int(dstCk[1]) < 32:
                    res_ws[f'O{row}'].value = 'N'
                    res_ws[f'Q{row}'].value = f'{dstCk[0]}/{dstCk[1]}'
                else:
                    res_ws[f'O{row}'].value = 'H'
                    res_ws[f'Q{row}'].value = dstCk[0]

                row += 1
                # print(dst)

            row = res_ws_row
            for svc in pol[6]:
                if svc[0] == 'all':
                    continue

                svcChkVal = True
                for val in svcChkList:
                    if svc[1] in val:
                        svcChkVal = False
                        svc[0] = val[0]

                if svcChkVal:
                    svcChkList.append(svc)

                # print(svcChkList)
                #
                # print(svc)
                svcVal = svc[1].split(' ')
                if svcVal[0] == "icmp":
                    res_ws[f'S{row}'].value = "PING"
                    res_ws[f'T{row}'].value = "ICMP"
                    res_ws[f'U{row}'].value = "NONE"
                    res_ws[f'V{row}'].value = "*"
                    res_ws[f'W{row}'].value = "*"
                elif svc[1] == "ip proto=2":
                    res_ws[f'S{row}'].value = "IGMP"
                    res_ws[f'T{row}'].value = "IGMP"
                    res_ws[f'U{row}'].value = "NONE"
                    res_ws[f'V{row}'].value = "*"
                    res_ws[f'W{row}'].value = "*"
                elif svc[1] == "ip proto=47":
                    res_ws[f'S{row}'].value = "GRE"
                    res_ws[f'T{row}'].value = "GRE"
                    res_ws[f'U{row}'].value = "NONE"
                    res_ws[f'V{row}'].value = "*"
                    res_ws[f'W{row}'].value = "*"
                elif svc[1] == "ip proto=89":
                    res_ws[f'S{row}'].value = "OSPF"
                    res_ws[f'T{row}'].value = "OSPF"
                    res_ws[f'U{row}'].value = "NONE"
                    res_ws[f'V{row}'].value = "*"
                    res_ws[f'W{row}'].value = "*"
                else:
                    # print(type(svc[0].upper()))
                    # print(svcList)
                    if svc[0].upper() in BLMSvcList:
                        res_ws[f'S{row}'].value = svc[0].upper() + "_1"
                    else:
                        res_ws[f'S{row}'].value = svc[0].upper()
                    res_ws[f'T{row}'].value = svcVal[0].upper()
                    res_ws[f'U{row}'].value = "NONE"

                    srcPortCk = svcVal[1].split('-')
                    if srcPortCk[0] == srcPortCk[1]:
                        res_ws[f'V{row}'].value = srcPortCk[0]
                    elif svcVal[1] == '1-65535':
                        res_ws[f'V{row}'].value = "*"
                    else:
                        res_ws[f'V{row}'].value = svcVal[1]

                    dstPortCk = svcVal[2].split('-')
                    if dstPortCk[0] == dstPortCk[1]:
                        res_ws[f'W{row}'].value = dstPortCk[0]
                    elif svcVal[2] == '1-65535':
                        res_ws[f'W{row}'].value = "*"
                    else:
                        res_ws[f'W{row}'].value = svcVal[2]
                row += 1

            res_ws_row += max(ruleCnt)

        prog += 1
        printProgress(prog, len(fileList), 'Progress:', 'Complete ', 1, 50)
        res_wb.save(SRC_DIR + f"{fileNo}_{fileOrg}_{fileSch}_BLM.xlsx")
        proc_wb.close()
        res_wb.close()




