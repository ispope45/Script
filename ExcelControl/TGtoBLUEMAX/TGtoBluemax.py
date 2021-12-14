import openpyxl
import os
import sys
import pandas as pd
import numpy as np

SRC_DIR = os.getcwd() + "//"
DST_FILE = SRC_DIR + "result.xlsx"
COL_LIST = ['PRIORITY', 'ENABLED', 'SRC', 'SRC ADDR', 'DST', 'DST ADDR', 'SVC', 'SVC SPEC',
            'ACTION', 'REVERSIBLE']

SAMPLE_FILE = "SRC/BLUEMAX_2.5.3_pol.xlsx"


def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_path, relative_path)


if __name__ == "__main__":

    fileList = os.listdir(SRC_DIR)
    res_wb = openpyxl.load_workbook(resource_path(SAMPLE_FILE))
    res_ws = res_wb.active

    for f in fileList:
        if f.find(".csv") == -1:
            continue

        proc_wb = openpyxl.Workbook()
        proc_ws = proc_wb.active
        proc_ws.append(COL_LIST)

        fileNo = f.split("_")[0]
        fileOrg = f.split("_")[1]
        fileSch = f.split("_")[2].split(".")[0]

        data = pd.read_csv(SRC_DIR + f)
        p_data = data[COL_LIST]
        totalValue = p_data.values.tolist()
        polItem = []
        ruleList = []
        srcList = []
        dstList = []
        svcList = []

        # print(totalValue)
        for val in totalValue:
            proc_ws.append(val)
            print(val)

        for row in range(2, proc_ws.max_row + 1):
            # print(row)
            # print(proc_ws[f'A{row}'].value)
            # print(type(proc_ws[f'A{row}'].value))
            if not(pd.isna(proc_ws[f'A{row}'].value)):
                ruleList.append(proc_ws[f'A{row}'].value)
                # print("OK")

            if not(pd.isna(proc_ws[f'B{row}'].value)):
                ruleList.append(proc_ws[f'B{row}'].value)

            if not(pd.isna(proc_ws[f'I{row}'].value)):
                ruleList.append(proc_ws[f'I{row}'].value)

            if not(pd.isna(proc_ws[f'J{row}'].value)):
                ruleList.append(proc_ws[f'J{row}'].value)

            if not(pd.isna(proc_ws[f'C{row}'].value)):
                srcList.append([proc_ws[f'C{row}'].value, proc_ws[f'D{row}'].value])

            if not(pd.isna(proc_ws[f'E{row}'].value)):
                dstList.append([proc_ws[f'E{row}'].value, proc_ws[f'F{row}'].value])

            if not(pd.isna(proc_ws[f'G{row}'].value)):
                svcList.append([proc_ws[f'G{row}'].value, proc_ws[f'H{row}'].value])

            if not(pd.isna(proc_ws[f'A{row}'].value)):
                ruleList.append(srcList)
                ruleList.append(dstList)
                ruleList.append(svcList)
                polItem.append(ruleList)

                ruleList = []
                srcList = []
                dstList = []
                svcList = []

        # print(polItem)
        # print(len(polItem))

        res_ws_row = 4

        for pol in polItem:
            print(pol)

            '''
            [3.0, 'Yes', 'Deny', 'No', 
                [
                    ['34.78.211.173', '34.78.211.173/32'], 
                    ['94.99.216.4', '94.99.216.4/32'],
                    ['192.168.0.0/24', '192.168.0.0/24']
                ],
                [
                    ['192.168.0.0/24', '192.168.0.0/24'], 
                    ['all', '0.0.0.0/0']
                ],
                [
                    ['Sharing_TCP_file', 'tcp 1-65535 137-139']
                ]
            ]
            '''

            ruleCnt = []
            ruleCnt.append(len(pol[4]))
            ruleCnt.append(len(pol[5]))
            ruleCnt.append(len(pol[6]))

            res_ws[f'B{res_ws_row}'].value = 'default'
            if pol[1] == 'Yes':
                res_ws[f'C{res_ws_row}'].value = 1
            else:
                res_ws[f'C{res_ws_row}'].value = 0

            res_ws[f'E{res_ws_row}'].value = pol[2]

            # print(max(ruleCnt))
            for i in range(res_ws_row, res_ws_row + max(ruleCnt)):
                res_ws[f'A{i}'].value = pol[0]

            row = res_ws_row
            for src in pol[4]:
                if src[1] == '0.0.0.0/0':
                    continue

                res_ws[f'J{row}'].value = src[0]
                srcCk = src[1].split('/')

                ipOctet = srcCk[0].split('.')
                if int(ipOctet[0]) == 10:
                    res_ws[f'F{row}'].value = 1
                    if int(ipOctet[1]) % 10 == 7:
                        res_ws[f'F{row}'].value = 2
                else:
                    res_ws[f'F{row}'].value = 2

                if srcCk[0].find("-") != -1:
                    res_ws[f'I{row}'].value = 'N'
                    res_ws[f'K{row}'].value = srcCk[0]
                elif int(srcCk[1]) > 32:
                    res_ws[f'I{row}'].value = 'N'
                    res_ws[f'K{row}'].value = f'{srcCk[0]}/{srcCk[1]}'
                else:
                    res_ws[f'I{row}'].value = 'H'
                    res_ws[f'K{row}'].value = srcCk[0]

                row += 1
                # print(src)

            row = res_ws_row
            for dst in pol[5]:
                if dst[1] == '0.0.0.0/0':
                    continue

                res_ws[f'P{row}'].value = dst[0]
                dstCk = dst[1].split('/')

                ipOctet = dstCk[0].split('.')
                if int(ipOctet[0]) == 10:
                    res_ws[f'L{row}'].value = 1
                    if int(ipOctet[1]) % 10 == 7:
                        res_ws[f'L{row}'].value = 2
                else:
                    res_ws[f'L{row}'].value = 2

                if dstCk[0].find("-") != -1:
                    res_ws[f'O{row}'].value = 'N'
                    res_ws[f'Q{row}'].value = dstCk[0]
                elif int(dstCk[1]) > 32:
                    res_ws[f'O{row}'].value = 'N'
                    res_ws[f'Q{row}'].value = f'{dstCk[0]}/{dstCk[1]}'
                else:
                    res_ws[f'O{row}'].value = 'H'
                    res_ws[f'Q{row}'].value = dstCk[0]

                row += 1
                # print(dst)

            row = res_ws_row
            for svc in pol[6]:
                if svc[0] == 'all':
                    continue

                # print(svc)
                res_ws[f'S{row}'].value = svc[0]
                svcVal = svc[1].split(' ')
                res_ws[f'T{row}'].value = svcVal[0]
                res_ws[f'U{row}'].value = "NONE"
                res_ws[f'V{row}'].value = "*"
                portCk = svcVal[2].split('-')
                if portCk[0] == portCk[1]:
                    res_ws[f'W{row}'].value = portCk[0]
                else:
                    res_ws[f'W{row}'].value = svcVal[2]
                row += 1

            res_ws_row += max(ruleCnt)

        res_wb.save(SRC_DIR + "result2.xlsx")




