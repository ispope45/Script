import openpyxl
import wget
import math
import os
import time
import random

OS_HOME_DRIVE = os.environ['HOMEDRIVE']
OS_HOME_PATH = os.environ['HOMEPATH']
HOME_PATH = OS_HOME_DRIVE + OS_HOME_PATH

SRC_PATH = HOME_PATH + '\\Desktop\\src\\'
DST_PATH = HOME_PATH + '\\Desktop\\dst\\'

SRC_FILE = SRC_PATH + 'sch_list.xlsx'

START_LINE = 0
END_LINE = 0
ERROR_CNT = 0


def bar_custom(current, total, width=80):
    width=30
    avail_dots = width-2
    shaded_dots = int(math.floor(float(current) / total * avail_dots))
    percent_bar = '[' + '■'*shaded_dots + ' '*(avail_dots-shaded_dots) + ']'
    progress = "%d%% %s [%d / %d]" % (current / total * 100, percent_bar, current, total)
    return progress


def download(url, out_path, error_cnt):
    try:
        wget.download(url, out=out_path, bar=bar_custom)
        print(" Ok")
        return int(error_cnt)
    except Exception as e:
        print(url + " Error : 404 Not Exist")
        # f = open(DST_PATH + "log.txt", "a+")
        # f.write(out_path + f" Error : 404 Not Exist : {e}\n")
        # f.close()
        return int(error_cnt + 1)


if __name__ == "__main__":
    # url = BASEURL + "7/report.pdf"
    wb = openpyxl.load_workbook(SRC_FILE)
    ws = wb.active

    if START_LINE == 0:
        startRow = 2
    else:
        startRow = START_LINE

    if END_LINE == 0:
        totalRows = ws.max_row + 1
    else:
        totalRows = END_LINE

    for i in range(startRow, totalRows):
        val_id = ws['A' + str(i)].value
        val_name = ws['B' + str(i)].value
        dir_name = DST_PATH + str(val_id) + "_" + val_name + "\\"
        if not os.path.isdir(dir_name):
            os.makedirs(dir_name, exist_ok=True)
            print("Create Dir " + dir_name)

        for Lid in range(1, 7):
            reqUrl = f'http://edu.im.neostep.kr/upload/school/{str(val_id)}/drawing/drawing_{Lid}.jpeg'
            # print(BASEURL + str(CONTENTS_ID[i]) + "/" + str(Lid) + ".mp4")
            fileName = dir_name + str(val_id) + "_" + str(val_name) + "_" + str(Lid) + ".jpeg"
            print(fileName)
            # time.sleep(random.randrange(1, 60))

            # print(TARGET_DIR + fileName)
            ERROR_CNT = download(reqUrl, fileName, ERROR_CNT)
            if ERROR_CNT > 5:
                ERROR_CNT = 0
                break

        for Lid in range(1, 7):
            reqUrl = f'http://edu.im.neostep.kr/upload/school/{str(val_id)}/drawing/drawing_{Lid}.png'
            # print(BASEURL + str(CONTENTS_ID[i]) + "/" + str(Lid) + ".mp4")
            fileName = dir_name + str(val_id) + "_" + str(val_name) + "_" + str(Lid) + ".png"
            print(fileName)
            # time.sleep(random.randrange(1, 60))

            # print(TARGET_DIR + fileName)
            ERROR_CNT = download(reqUrl, fileName, ERROR_CNT)
            if ERROR_CNT > 5:
                ERROR_CNT = 0
                break


