import requests
import openpyxl
import os
import sys
import paramiko
import time
import urllib3
from datetime import date

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

START_DATE = date.today()
MAIN_URL = "https://192.168.0.254:50005/"
LOGIN_API = "login/loginAction"
BACKUP_API = "firewall/firewall/ipv4Policy/list?json=true&exportAction=true&queryString=&"
BACKUP_API2 = "firewall/firewall/ipv4Policy/exportAction?profileId=&"

CUR_PATH = os.getcwd()
SRC_DIR = CUR_PATH + "\\"
SRC_FILE = "Form1.xlsx"


def printProgress(iteration, total, prefix='', suffix='', decimals=1, barLength=100):
    formatStr = "{0:." + str(decimals) + "f}"
    percent = formatStr.format(100 * (iteration / float(total)))
    filledLength = int(round(barLength * iteration / float(total)))
    bar = '#' * filledLength + '-' * (barLength - filledLength)
    sys.stdout.write('\r%s |%s| %s%s %s' % (prefix, bar, percent, '%', suffix)),
    if iteration == total:
        sys.stdout.write('\n')
    sys.stdout.flush()


def write_log(string):
    dat = str(START_DATE).replace("-", "")
    f = open(SRC_DIR + f'log_{dat}.txt', "a+")
    f.write(f'{string}\n')
    f.close()


if __name__ == "__main__":
    print("Preparing TG Backup Tool....")
    wb = openpyxl.load_workbook(SRC_DIR + SRC_FILE)
    ws = wb.active

    for row in range(2, ws.max_row + 1):
        no = str(ws[f'A{row}'].value)
        org = ws[f'B{row}'].value
        name = ws[f'C{row}'].value
        utmIp = ws[f'D{row}'].value

        USERID = ws[f'E{row}'].value
        USERPASS = ws[f'F{row}'].value

        login_data = {'disconnectType': '',
                      'webDeviceForceLogin': '',
                      'force': 'false',
                      'userId': USERID,
                      'password': USERPASS,
                      'language': 0}

        MAIN_URL = f"https://{utmIp}:50005/"

        with requests.Session() as s:
            with s.post(MAIN_URL + LOGIN_API, data=login_data, verify=False) as res:
                write_log(f"{no}_{name}; {MAIN_URL + LOGIN_API}; {res.status_code}; ")

            with s.get(MAIN_URL + BACKUP_API2, verify=False) as res:
                write_log(f"{no}_{name}; {MAIN_URL + BACKUP_API2}; {res.status_code}; 2.7.1")
                if len(res.text) < 1:
                    with s.get(MAIN_URL + BACKUP_API, verify=False) as res:
                        write_log(f"{no}_{name}; {MAIN_URL + BACKUP_API}; {res.status_code}; 2.7.1 later")

                result = res.text.replace("\n\n", "\n")

                f2 = open(SRC_DIR + f"{no}_{org}_{name}.csv", 'w', newline='')
                f2.write(result)
                f2.close()

        s.close()
        printProgress(row, ws.max_row, 'Progress:', 'Complete ', 1, 50)
    wb.close()
    os.system("pause")
