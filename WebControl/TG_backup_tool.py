import requests
import openpyxl
import os
import sys
import paramiko
import time

MAIN_URL = "https://192.168.0.254:50005/"
LOGIN_API = "login/loginAction"
BACKUP_API = "firewall/firewall/ipv4Policy/list?json=true&exportAction=true&queryString=&"

USERID = "manager"
USERPASS = "qwe123!@#"

login_data = {'disconnectType': '',
              'webDeviceForceLogin': '',
              'force': 'false',
              'userId': USERID,
              'password': USERPASS,
              'language': 0}
CUR_PATH = os.getcwd()
SRC_DIR = CUR_PATH + "\\"
SRC_FILE = "form.xlsx"


def printProgress(iteration, total, prefix='', suffix='', decimals=1, barLength=100):
    formatStr = "{0:." + str(decimals) + "f}"
    percent = formatStr.format(100 * (iteration / float(total)))
    filledLength = int(round(barLength * iteration / float(total)))
    bar = '#' * filledLength + '-' * (barLength - filledLength)
    sys.stdout.write('\r%s |%s| %s%s %s' % (prefix, bar, percent, '%', suffix)),
    if iteration == total:
        sys.stdout.write('\n')
    sys.stdout.flush()


cmd1 = ['show_all\n']


def main(sw_ip, sw_user, sw_pass, sw_port, command):
    host = sw_ip
    username = sw_user
    password = sw_pass

    output = list()
    try:
        conn = paramiko.SSHClient()
        conn.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        conn.connect(host, username=username, password=password, port=sw_port, timeout=100)
        channel = conn.invoke_shell()
        time.sleep(3)

        for line in command:
            res = ""
            res += send(channel, line)
            output.append(res)

        return output

    except Exception as e:
        if "port" in str(e):
            return "Port Error"
        if "WinError 10060" in str(e):
            return "Connection Error"
        print(e)
        return "Error"

    finally:
        if conn is not None:
            conn.close()


def send(channel, cmd):
    channel.send(cmd)
    out_data, err_data = wait_streams(channel)
    a = out_data.replace("\\r\\n", "\n").replace("\\r", "")
    b = a.replace('b"', "").replace("b'", "").replace('"', "")
    return b


def wait_streams(channel):
    time.sleep(1)
    out_data = ""
    err_data = ""
    while True:
        time.sleep(1)
        if channel.recv_ready():
            out_data += str(channel.recv(100000))
            if channel.recv_stderr_ready():
                err_data += str(channel.recv(100000))
            break

    return out_data, err_data


if __name__ == "__main__":
    wb = openpyxl.load_workbook(SRC_DIR + SRC_FILE)
    ws = wb.active

    for row in range(2, ws.max_row + 1):
        no = str(ws[f'A{row}'].value)
        org = ws[f'B{row}'].value
        name = ws[f'C{row}'].value
        utmIp = ws[f'D{row}'].value

        MAIN_URL = f"https://{utmIp}:50005/"

        with requests.Session() as s:
            with s.post(MAIN_URL + LOGIN_API, data=login_data, verify=False) as res:
                print(res.status_code)

            with s.get(MAIN_URL + BACKUP_API, verify=False) as res:
                print(res.status_code)
                result = res.text.replace("\n\n", "\n")

                res = main(utmIp, USERID, USERPASS, 22, cmd1)

                f1 = open(SRC_DIR + f"{no}_{org}_{name}.txt", "w")
                for val in res:
                    f1.write(val)
                f1.close()

                f2 = open(SRC_DIR + f"{no}_{org}_{name}.csv", 'w', newline='')
                f2.write(result)
                f2.close()
        s.close()

        printProgress(row, ws.max_row, 'Progress:', 'Complete ', 1, 50)
    wb.close()
    os.system("pause")
