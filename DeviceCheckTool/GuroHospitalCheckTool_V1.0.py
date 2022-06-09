import paramiko
import time
import openpyxl
import os

CUR_PATH = os.getcwd()


def main(sw_ip, sw_user, sw_pass, sw_port, sw_proto):
    host = sw_ip
    username = sw_user
    password = sw_pass
    # port = sw_port
    # proto = sw_proto

    output = list()
    command = list()
    command.append('enable\n')
    command.append('show syslog non-volatile tail\n')
    command.append('show temperature\n')

    try:
        conn = paramiko.SSHClient()
        conn.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        conn.connect(host, username=username, password=password)
        channel = conn.invoke_shell()
        time.sleep(3)

        for line in command:
            channel.send(line)
            out_data, err_data = wait_streams(channel)

            a = out_data.replace("\\r\\n", "\n").replace("\\r", "")
            b = a.replace('b"', "").replace("b'", "").replace(" '", "").replace('"', "")

            output.append(b)

        return output

    except Exception as e:
        print(e)
        return "Error"

    finally:
        if conn is not None:
            conn.close()


def wait_streams(channel):
    time.sleep(1)
    out_data = ""
    err_data = ""

    while channel.recv_ready():
        out_data += str(channel.recv(1000))
    while channel.recv_stderr_ready():
        err_data += str(channel.recv_stderr(1000))

    return out_data, err_data


if __name__ == "__main__":

    SRC_FILE = CUR_PATH + "\\Python_diag.xlsx"
    wb = openpyxl.load_workbook(SRC_FILE)
    ws = wb.active

    for row in range(2, ws.UsedRange.Rows.Count+1):
        f = open(CUR_PATH + f'{str(int(ws.Cells(row, 1).Value))}_{ws.Cells(row, 2).Value}.txt', "w+")
        sw_idx = ws[f'A{row}'].value
        sw_hostname = ws[f'B{row}'].value
        sw_ip = ws[f'C{row}'].value
        sw_user = ws[f'D{row}'].value
        sw_pass = ws[f'E{row}'].value
        sw_port = ws[f'F{row}'].value
        sw_proto = ws[f'G{row}'].value
        e = main(sw_ip, sw_user, sw_pass, sw_port, sw_proto)
        for line in e:
            print(line)
            f.write(line)
        f.close()

    wb.close()
