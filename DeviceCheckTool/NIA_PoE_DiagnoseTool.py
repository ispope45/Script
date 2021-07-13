import paramiko
import time
import os
import openpyxl

# GLOBAL
OS_HOME_DRIVE = os.environ['HOMEDRIVE']
OS_HOME_PATH = os.environ['HOMEPATH']
HOME_PATH = OS_HOME_DRIVE + OS_HOME_PATH

SRC_PATH = HOME_PATH + '\\Desktop\\src\\'
DST_PATH = HOME_PATH + '\\Desktop\\dst\\'
SRC_FILE = HOME_PATH + '\\Desktop\\Dev_List.xlsx'
DST_FILE = HOME_PATH + '\\Desktop\\dst.xlsx'

form_desc = '■ PoE Switch Diagnose Tool \n\n'

#  0 = 처음(2)부터 끝까지
START_LINE = 0  # 부터
END_LINE = 0  # 이전까지

cmd1 = ['enable\n', 'show interface status\n', ' ', ' ', 'show power inline port-status\n', ' ', ' ']
cmd2 = ['enable\n', 'show power inline port-status\n']


def main(sw_ip, sw_user, sw_pass, sw_port, command):
    host = sw_ip
    username = sw_user
    password = sw_pass

    output = list()

    try:
        conn = paramiko.SSHClient()
        conn.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        conn.connect(host, username=username, password=password, port=sw_port)
        channel = conn.invoke_shell()
        time.sleep(2)

        for line in command:
            channel.send(line)
            out_data, err_data = wait_streams(channel)
            tmp = out_data.split("\\r\\n")

            for v in tmp:
                output.append(v)

        return output

    except Exception as e:
        if "port" in str(e):
            return "Port Error"
        if "WinError 10060" in str(e):
            return "Connection Error"
        print(e)
        return "Error"

    finally:
        if conn is not None:
            conn.close()


def wait_streams(channel):
    time.sleep(0.1)
    out_data = ""
    err_data = ""

    while channel.recv_ready():
        out_data += str(channel.recv(1000))
    while channel.recv_stderr_ready():
        err_data += str(channel.recv_stderr(1000))

    return out_data, err_data


if __name__ == "__main__":

    print(form_desc)

    wb = openpyxl.load_workbook(SRC_FILE)
    ws = wb.active

    target_wb = openpyxl.load_workbook(DST_FILE)
    target_ws = target_wb.active

    cnt = 1

    if START_LINE == 0:
        startRow = 2
    else:
        startRow = START_LINE

    if END_LINE == 0:
        totalRows = ws.max_row + 1
    else:
        totalRows = END_LINE

    for i in range(startRow, totalRows):

        row = str(i)
        val_name = ws['B' + row].value
        val_devName = ws['C' + row].value
        val_ip = ws['D' + row].value
        val_port = ws['E' + row].value
        val_id = "admin"
        val_password = "frontier1!"

        print(f'\n\n{val_name} : {val_devName}({val_ip})')

        try:
            s_time = time.time()
            res = main(val_ip, val_id, val_password, val_port, cmd1)
            e_time = time.time()
            print("SSH 소요시간 : " + str(e_time - s_time))
            if "Error" in res:
                f = open(DST_PATH + "log.txt", "a+")
                f.write(f'{val_name} : {val_devName}({val_ip}) : {res}\n')
                f.close()
                print(res)
                continue
            # res2 = main(val_ip, val_id, val_password, val_port, cmd2)
            port_val = list()
            power_val = list()
            port_stat = list()
            power_stat = list()

            for r in res:
                if "Gi0/" in r:
                    val = (r.replace("b'", "").replace("\\r", "").replace("'", ""))
                    port_val = [val[:6].strip(), val[28:38].strip()]
                    port_stat.append(port_val)
                elif "Giga0/" in r:
                    val = (r.replace("b'", "").replace("\\r", "").replace("'", ""))
                    power_val = [val[:8].strip(), val[12:25].strip(), val[27:47].strip()]
                    power_stat.append(power_val)

            # for r in res2:
            #     if "Gi" in r:
            #         val = (r.replace("b'", "").replace("\\r", "").replace("'", ""))
            #         power_val = [val[:8].strip(), val[12:25].strip(), val[27:47].strip()]
            #         power_stat.append(power_val)

            for j in range(0, len(power_stat)):
                cnt += 1
                target_ws['A' + str(cnt)].value = val_name
                target_ws['B' + str(cnt)].value = val_devName
                target_ws['C' + str(cnt)].value = val_ip
                target_ws['D' + str(cnt)].value = port_stat[j][0]
                target_ws['E' + str(cnt)].value = port_stat[j][1]
                target_ws['F' + str(cnt)].value = power_stat[j][0]
                target_ws['G' + str(cnt)].value = power_stat[j][1]
                target_ws['H' + str(cnt)].value = power_stat[j][2]

            target_wb.save(filename=DST_FILE)

        except Exception as e:
            print(e)
    target_wb.close()
