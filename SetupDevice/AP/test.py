import openpyxl
import os
import sys
import time
import paramiko
import socket

from datetime import date

CUR_PATH = os.getcwd()
SRC_FILE = CUR_PATH + "\\ap_config_src.xlsx"
START_DATE = date.today()


def davo4038_setting(ip, netmask, gw):
    script = [
        f'config\n',
        f'basic wan\n',
        f'static\n',
        f'{ip}\n',
        f'{netmask}\n',
        f'{gw}\n',
        f'168.126.63.1\n',
        f'168.126.63.2\n',
        f'y\n',
        f'apply\n']
    return script


def davo504_setting(ip, netmask, gw):
    script = [
        f'config\n',
        f'configure\n',
        f'config wan\n',
        f'static\n',
        f'{ip}\n',
        f'{netmask}\n',
        f'{gw}\n',
        f'168.126.63.1\n',
        f'168.126.63.2\n',
        f'y\n',
        f'apply\n']
    return script


def samsung_setting(ip, netmask, gw):
    script = [f'config interface address {ip} {netmask} {gw}\n']
    return script


def main(sw_ip, sw_user, sw_pass, sw_port, command):
    host = sw_ip
    username = sw_user
    password = sw_pass

    output = list()

    try:
        conn = paramiko.SSHClient()
        conn.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        conn.connect(host, username=username, password=password, port=sw_port, timeout=10)
        channel = conn.invoke_shell()
        time.sleep(2)

        for line in command:
            channel.send(line)
            time.sleep(0.1)
            # out_data, err_data = wait_streams(channel)
            # output.append(out_data)
        return "OK"

    except Exception as e:
        if "port" in str(e):
            return "Port Error"
        if "WinError 10060" in str(e):
            return "Connection Error"
        # print(e)
        return e

    finally:
        if conn is not None:
            conn.close()


def wait_streams(channel):
    time.sleep(0.2)
    out_data = ""
    err_data = ""

    while channel.recv_ready():
        out_data += str(channel.recv(1000))
    while channel.recv_stderr_ready():
        err_data += str(channel.recv_stderr(1000))

    return out_data, err_data


def write_log(string):
    dat = str(START_DATE).replace("-", "")
    f = open(CUR_PATH + f'\\log_{dat}.txt', "a+")
    now = time.localtime()
    cur_time = "%02d/%02d,%02d:%02d:%02d" % (now.tm_mon, now.tm_mday, now.tm_hour, now.tm_min, now.tm_sec)
    f.write(f'{cur_time};{string}\n')
    f.close()


def port_check(ip, port):
    try:
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(1.5)
        result = sock.connect_ex((ip, port))
        sock.close()
    except socket.error as e:
        result = 1
        print("Error : " + e)

    if result == 0:
        return True
    else:
        return False


def print_progress(iteration, total, prefix='', suffix='', decimals=1, barLength=100):
    formatStr = "{0:." + str(decimals) + "f}"
    percent = formatStr.format(100 * (iteration / float(total)))
    filledLength = int(round(barLength * iteration / float(total)))
    bar = '#' * filledLength + '-' * (barLength - filledLength)
    sys.stdout.write('\r%s |%s| %s%s %s' % (prefix, bar, percent, '%', suffix)),
    if iteration == total:
        sys.stdout.write('\n')
    sys.stdout.flush()


if __name__ == "__main__":
    wb = openpyxl.load_workbook(SRC_FILE)
    ws = wb.active

    for row in range(2, ws.max_row + 1):

        schName = ws[f'A{row}'].value
        schModel = ws[f'B{row}'].value
        preIP = ws[f'C{row}'].value
        postIP = ws[f'D{row}'].value
        netmask = ws[f'E{row}'].value
        gw = ws[f'F{row}'].value

        accessPort = ws[f'G{row}'].value
        devId = ws[f'H{row}'].value
        devPass = ws[f'I{row}'].value

        if str(schName) == "None":
            continue

        if "504" in str(schModel):
            script = davo504_setting(postIP, netmask, gw)
        elif "4038" in str(schModel):
            script = davo4038_setting(postIP, netmask, gw)
        elif "WEA" in str(schModel):
            script = samsung_setting(postIP, netmask, gw)
        else:
            write_log(f'{schName};{preIP};Model Invalid;False')
            ws[f'L{row}'].value = "Model Invalid"
            continue

        if port_check(preIP, accessPort):
            write_log(f'{schName};{preIP}:{accessPort};port check;OK')
            ws[f'J{row}'].value = "O"

            res = main(preIP, devId, devPass, accessPort, script)
            if res:
                ws[f'L{row}'].value = str(res)
                write_log(f'{schName};{preIP}:{accessPort};{str(res)};')

        else:
            write_log(f'{schName};{preIP}:{accessPort};port check;False')
            ws[f'J{row}'].value = "X"

        print_progress(row - 1, ws.max_row - 1, 'Progress:', 'Complete ', 1, 50)
        wb.save(SRC_FILE)

    print("Post Port Check Phase....")
    time.sleep(10)

    for row in range(2, ws.max_row + 1):
        schName = ws[f'A{row}'].value
        postIP = ws[f'D{row}'].value

        accessPort = ws[f'G{row}'].value

        if port_check(postIP, accessPort):
            write_log(f'{schName};{postIP}:{accessPort};port check;OK')
            ws[f'K{row}'].value = "O"
        else:
            write_log(f'{schName};{preIP}:{accessPort};port check;False')
            ws[f'K{row}'].value = "X"

        print_progress(row - 1, ws.max_row - 1, 'Progress:', 'Complete ', 1, 50)

    wb.save(SRC_FILE)
