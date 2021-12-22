import openpyxl
import os
from datetime import date
import sys

CUR_PATH = os.getcwd()
SRC_FILE = CUR_PATH + "\\wifi_src.xlsx"
START_DATE = date.today()


def ip_calculator(ip, mask):
    ipOctet_A = int(ip.split('.')[0])
    ipOctet_B = int(ip.split('.')[1])
    ipOctet_C = int(ip.split('.')[2])
    # ipOctet_D = int(ip.split('/')[0].split('.')[3])

    ipOctet_Cp = ipOctet_C + (2 ** (24 - mask)) - 1

    subnetmask = ''
    if mask // 8 > 0:
        subnetmask += '255.'

    if mask // 8 > 1:
        subnetmask += '255.'

    if mask // 8 > 2:
        subnetmask += '255.'

    subnetmask += str(256 - 2 ** (8 - (mask % 8))) + '.'
    subnetmask += '0'

    gateway = 254

    l2_ip = f'{str(ipOctet_A)}.{str(ipOctet_B)}.{str(ipOctet_Cp)}.220'
    gw_ip = f'{str(ipOctet_A)}.{str(ipOctet_B)}.{str(ipOctet_Cp)}.{str(gateway)}'
    return l2_ip, gw_ip, subnetmask


if __name__ == "__main__":
    wb = openpyxl.load_workbook(SRC_FILE)
    ws = wb.active

    for i in range(2, ws.max_row + 1):
        row = str(i)

        execution = ws[f'C{row}'].value
        Org = ws[f'B{row}'].value
        schName = ws[f'A{row}'].value
        hostname = ws[f'F{row}'].value

        networkId = ws[f'G{row}'].value
        netmask = ws[f'H{row}'].value

        l2_ip, gw_ip, subnetmask = ip_calculator(networkId, netmask)

        script = (
            f'enable\n'
            f'config\n'
            f'hostname {hostname}\n'
            f'ip sshd port 9992\n\n'
            f'interface vlan 1\n'
            f'ip address {l2_ip} {subnetmask}\n'
            f'exit\n\n'
            f'ip route 0.0.0.0 0.0.0.0 {gw_ip}\n\n'
            'snmp-server group sen_class v3 auth\n'
            'snmp-server user sen_class sen_class v3 auth md5 sen_enc02\n'
            'snmp-server host 172.28.228.78 version v3 priv sen_class\n'
            'snmp-server trap-logs\n\n'
            'write\n\n'
            )
        print(script)
        if not(os.path.isdir(CUR_PATH + f'\\{execution}')):
            os.makedirs(CUR_PATH + f'\\{execution}')

        f = open(CUR_PATH + f"\\{execution}\\{Org}_{schName}.txt", "w+")
        f.write(script)
        f.close()

