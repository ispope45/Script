import openpyxl
import os
from datetime import date

CUR_PATH = os.getcwd()
SRC_FILE = CUR_PATH + "\\backup.xlsx"
START_DATE = date.today()


def ip_calculator(ip, d_octet):
    subnet = int(ip.split('/')[1])
    ipOctet_A = int(ip.split('/')[0].split('.')[0])
    ipOctet_B = int(ip.split('/')[0].split('.')[1])
    ipOctet_C = int(ip.split('/')[0].split('.')[2])
    # ipOctet_D = int(ip.split('/')[0].split('.')[3])

    ipOctet_Cp = ipOctet_C + (2 ** (24 - subnet)) - 1
    ipOctet_Dp = d_octet
    ipOctet_Dpp = 254

    l2_ip = f'{str(ipOctet_A)}.{str(ipOctet_B)}.{str(ipOctet_Cp)}.{str(ipOctet_Dp)}/{str(subnet)}'
    gw_ip = f'{str(ipOctet_A)}.{str(ipOctet_B)}.{str(ipOctet_Cp)}.{str(ipOctet_Dpp)}'
    return l2_ip, gw_ip


if __name__ == "__main__":
    wb = openpyxl.load_workbook(SRC_FILE)
    ws = wb.active

    for i in range(2, ws.max_row + 1):
        row = str(i)

        item_set = []

        No = ws[f'A{row}'].value
        Org = ws[f'B{row}'].value
        schName = ws[f'D{row}'].value
        hostname = ws[f'E{row}'].value
        execution = ws[f'T{row}'].value

        # hostname_t = hostname + "_TL2-1"
        # hostname_s = hostname + "_SL2-1"
        hostname_w = hostname + "_WL2-1"
        # hostname_e = hostname + "_EL2-1"

        item = []
        if ws[f'H{row}'].value != "없음":
            ip_w = ws[f'H{row}'].value.split(',')
            ip_w_l2, ip_w_gw = ip_calculator(ip_w[0], 242)
            item.append(No)
            item.append(Org)
            item.append(schName)
            item.append("무선망PoE")
            item.append(hostname_w)
            item.append(ip_w_l2)
            item.append(ip_w_gw)

            item_set.append(item)

        for val in item_set:
            script = (
                f'en\n'
                f'config terminal\n'
                f'interface vlan 1\n'
                f'ip address {val[5]}\n'
                f'exit\n\n'
                f'hostname {val[4]}\n'
                f'ip route 0.0.0.0/0 {val[6]}\n'
                'spanning-tree shutdown\n'
                'service ssh\n'
                'ip ssh port 2004\n'
                f'end\n\n'
                f'write memory\n'
                f'y\n\n\n\n'
                '---------------- 계정정보(직접입력) ---------------- \n\n'
                '기본값 : root // [패스워드없음]\n'
                '** root 입력시 계정재생성 메세지출력됨\n'
                '변경값 : mainsw // Sen16701396!\n\n')

            if not(os.path.isdir(CUR_PATH + f'\\{execution}')):
                os.makedirs(CUR_PATH + f'\\{execution}')

            f = open(CUR_PATH + f"\\{val[0]}_{val[1]}_{val[2]}_{val[3]}.txt", "w+")
            f.write(script)
            f.close()

