import paramiko
import time
import os, sys
import openpyxl
import socket
from datetime import date

# GLOBAL
OS_HOME_DRIVE = os.environ['HOMEDRIVE']
OS_HOME_PATH = os.environ['HOMEPATH']
HOME_PATH = OS_HOME_DRIVE + OS_HOME_PATH

CUR_PATH = os.getcwd()
START_DATE = date.today()

SRC_PATH = HOME_PATH + '\\Desktop\\src\\'
DST_PATH = HOME_PATH + '\\Desktop\\dst\\'

SRC_FILE = HOME_PATH + '\\Desktop\\src\\Dev_List.xlsx'
DST_FILE = HOME_PATH + '\\Desktop\\dst\\dst.xlsx'

form_desc = '■ PoE Switch Diagnose Tool \n\n'

cmd1 = ['enable\n',
        'configure terminal\n',
        'snmp-server group schoolnet4 v3 priv write snmpview\n',
        'snmp-server view snmpview 1.3.6.1 included\n',
        'snmp-server user schoolnet4 schoolnet4 v3\n',
        'sha\n',
        'Sen16701396!\n',
        'Sen16701396!\n',
        'aes\n',
        'Sen16701396!\n',
        'Sen16701396!\n',
        '\n',
        'snmp-server host 192.168.75.231 trap version 3 priv schoolnet4\n',
        'snmp-server enable traps interface\n',
        'snmp-server enable traps interface backup\n',
        'snmp-server enable traps envmon fan supply temperature ext-supply\n',
        'snmp-server enable traps port-monitor\n',
        'snmp-server enable traps traffic-control\n',
        'snmp-server enable traps resource\n',
        'snmp-server enable traps sld state-change\n',
        'snmp-server enable traps snmp authFail coldStart warmStart\n',
        'logging 192.168.75.231\n',
        'logging trap errors\n',
        'end\n',
        'wr m\n',
        'y\n']


def main(sw_ip, sw_user, sw_pass, sw_port, command):
    host = sw_ip
    username = sw_user
    password = sw_pass

    output = list()

    try:
        conn = paramiko.SSHClient()
        conn.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        conn.connect(host, username=username, password=password, port=sw_port, timeout=10)
        channel = conn.invoke_shell()
        time.sleep(2)

        for line in command:
            channel.send(line)
            out_data, err_data = wait_streams(channel)
            output.append(out_data)

        return "OK"

    except Exception as e:
        if "port" in str(e):
            return "Port Error"
        if "WinError 10060" in str(e):
            return "Connection Error"
        # print(e)
        return e

    finally:
        if conn is not None:
            conn.close()


def wait_streams(channel):
    time.sleep(0.2)
    out_data = ""
    err_data = ""

    while channel.recv_ready():
        out_data += channel.recv(1000).decode('ascii')
    while channel.recv_stderr_ready():
        err_data += channel.recv_stderr(1000).decode('ascii')

    return out_data, err_data


def write_log(string):
    dat = str(START_DATE).replace("-", "")
    f = open(CUR_PATH + f'\\log_{dat}.txt', "a+")
    now = time.localtime()
    cur_time = "%02d/%02d,%02d:%02d:%02d" % (now.tm_mon, now.tm_mday, now.tm_hour, now.tm_min, now.tm_sec)
    f.write(f'{cur_time};{string}\n')
    f.close()


def port_check(ip, port):
    try:
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(1.5)
        result = sock.connect_ex((ip, port))
        sock.close()
    except socket.error as e:
        result = 1
        print("Error : " + e)

    if result == 0:
        return True
    else:
        return False


def print_progress(iteration, total, prefix='', suffix='', decimals=1, barLength=100):
    formatStr = "{0:." + str(decimals) + "f}"
    percent = formatStr.format(100 * (iteration / float(total)))
    filledLength = int(round(barLength * iteration / float(total)))
    bar = '#' * filledLength + '-' * (barLength - filledLength)
    sys.stdout.write('\r%s |%s| %s%s %s' % (prefix, bar, percent, '%', suffix)),
    if iteration == total:
        sys.stdout.write('\n')
    sys.stdout.flush()


if __name__ == "__main__":
    SRC_FILE = CUR_PATH + "\\L2_IP_LIST.xlsx"
    wb = openpyxl.load_workbook(SRC_FILE)
    ws = wb.active

    # row = 6
    cnt = 0
    for row in range(2, ws.max_row + 1):

        schNo = ws[f'A{row}'].value
        schOrg = ws[f'C{row}'].value
        schName = ws[f'D{row}'].value
        schCls = ws[f'E{row}'].value
        schL2Ip = ws[f'H{row}'].value
        schL2AccessPort = ws[f'I{row}'].value
        schL2Id = ws[f'J{row}'].value
        schL2Password = ws[f'K{row}'].value

        if port_check(schL2Ip, schL2AccessPort):
            ws[f'L{row}'].value = "O"
            res = main(schL2Ip, schL2Id, schL2Password, schL2AccessPort, cmd1)
            if res:
                ws[f'M{row}'].value = str(res)
        else:
            ws[f'L{row}'].value = "X"

        print_progress(row - 1, ws.max_row - 1, 'Progress:', 'Complete ', 1, 50)

        if row % 10 == 5:
            wb.save(SRC_FILE)

    wb.save(SRC_FILE)

