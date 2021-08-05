import requests
import hashlib
import openpyxl
import os
import socket
import datetime
import time
import urllib3

# id / 집선청 / 구분 / 설치업체 / 학교명 / vendor / name / ip / port / ssid / 패스워드

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# GLOBAL
OS_HOME_DRIVE = os.environ['HOMEDRIVE']
OS_HOME_PATH = os.environ['HOMEPATH']
HOME_PATH = OS_HOME_DRIVE + OS_HOME_PATH

SRC_PATH = HOME_PATH + '\\Desktop\\src\\'
DST_PATH = HOME_PATH + '\\Desktop\\dst\\'

# SRC_FILE = SRC_PATH + 'Dev_List_20210706_PoE.xlsx'
SRC_FILE = SRC_PATH + 'TEST1.xlsx'

START_LINE = 0
END_LINE = 0

form_desc = '■ Auto Ap Config Tool\n' \
            'Read IP, Port number and AP configuraion. \n\n' \
            'id / 집선청 / 구분 / 설치업체 / 학교명 / vendor / name / ip / port / ssid / 패스워드 \n\n'


def port_check(ip, port):
    try:
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(0.5)
        result = sock.connect_ex((ip, port))
        sock.close()
    except socket.error as e:
        result = 1
        print("Error : " + e)

    if result == 0:
        return True
    else:
        return False


def kaon_web(ip, ssid, pwd):
    MAIN_URL = f'http://{ip}:8080/api/auth/token'
    SETTING_URL = f'http://{ip}:8080/api/wireless'
    RESTART_URL = f'http://{ip}:8080/api/operation/reboot'

    LOGIN_DATA = {'auth': {"username": "YWRtaW4=", "authHash": ""}}
    SETTING_DATA = {"in-body":{"band-steering":True,"5g":{"radio":{"id":0,"bandwidth":7,"channel":0,"wps":True,"sideband":-1,"dfs":False,"dfs-status":"IDLE","airtime-frame":False,"power":"100","nphy":"-1","rate":"54000000","frag-threshold":"2346","rts-threshold":"2347","dtim-interval":"1","beacon-interval":"100","beacon-rotation":False,"preamble-type":"1","max-assoc":"64"},"wlans":[{"id":0,"ssid_id":0,"status":True,"ssid":ssid,"enc-method":"WPA2 Mixed","enc-algorithm":"aes","enc-key":pwd,"r-server-ip":"0.0.0.0","r-server-port":"1812","r-server-key":"1234","hidden":True,"isolate":False,"wmm":True,"wmf":True},{"id":0,"ssid_id":1,"status":False,"ssid":"KaonAP-Guest01-5G","enc-method":"WPA2 Mixed","enc-algorithm":"tkip+aes","enc-key":"sen2021!wi","r-server-ip":"0.0.0.0","r-server-port":"1812","r-server-key":"1234","hidden":False,"isolate":False,"wmm":True,"wmf":True},{"id":0,"ssid_id":2,"status":False,"ssid":"KaonAP-Guest02-5G","enc-method":"WPA2 Mixed","enc-algorithm":"tkip+aes","enc-key":"sen2021!wi","r-server-ip":"0.0.0.0","r-server-port":"1812","r-server-key":"1234","hidden":False,"isolate":False,"wmm":True,"wmf":True},{"id":0,"ssid_id":3,"status":False,"ssid":"KaonAP-Guest03-5G","enc-method":"WPA2 Mixed","enc-algorithm":"tkip+aes","enc-key":"sen2021!wi","r-server-ip":"0.0.0.0","r-server-port":"1812","r-server-key":"1234","hidden":False,"isolate":False,"wmm":True,"wmf":True},{"id":0,"ssid_id":4,"status":False,"ssid":"KaonAP-Guest04-5G","enc-method":"WPA2 Mixed","enc-algorithm":"tkip+aes","enc-key":"sen2021!wi","r-server-ip":"0.0.0.0","r-server-port":"1812","r-server-key":"1234","hidden":False,"isolate":False,"wmm":True,"wmf":True}]},"2g":{"radio":{"id":1,"bandwidth":1,"channel":0,"wps":True,"sideband":"-1","airtime-frame":False,"power":"100","nphy":"-1","rate":"54000000","frag-threshold":"2346","rts-threshold":"2347","dtim-interval":"1","beacon-interval":"100","beacon-rotation":False,"preamble-type":"1","max-assoc":"64"},"wlans":[{"id":1,"ssid_id":0,"status":True,"ssid":ssid,"enc-method":"WPA2 Mixed","enc-algorithm":"aes","enc-key":pwd,"r-server-ip":"0.0.0.0","r-server-port":"1812","r-server-key":"1234","hidden":True,"isolate":False,"wmm":True,"wmf":True},{"id":1,"ssid_id":1,"status":False,"ssid":"KaonAP-Guest01-2.4G","enc-method":"WPA2 Mixed","enc-algorithm":"tkip+aes","enc-key":"sen2021!wi","r-server-ip":"0.0.0.0","r-server-port":"1812","r-server-key":"1234","hidden":False,"isolate":False,"wmm":True,"wmf":True},{"id":1,"ssid_id":2,"status":False,"ssid":"KaonAP-Guest02-2.4G","enc-method":"WPA2 Mixed","enc-algorithm":"tkip+aes","enc-key":"sen2021!wi","r-server-ip":"0.0.0.0","r-server-port":"1812","r-server-key":"1234","hidden":False,"isolate":False,"wmm":True,"wmf":True},{"id":1,"ssid_id":3,"status":False,"ssid":"KaonAP-Guest03-2.4G","enc-method":"WPA2 Mixed","enc-algorithm":"tkip+aes","enc-key":"sen2021!wi","r-server-ip":"0.0.0.0","r-server-port":"1812","r-server-key":"1234","hidden":False,"isolate":False,"wmm":True,"wmf":True},{"id":1,"ssid_id":4,"status":False,"ssid":"KaonAP-Guest04-2.4G","enc-method":"WPA2 Mixed","enc-algorithm":"tkip+aes","enc-key":"sen2021!wi","r-server-ip":"0.0.0.0","r-server-port":"1812","r-server-key":"1234","hidden":False,"isolate":False,"wmm":True,"wmf":True}]}}}

    try:
        with requests.Session() as s:
            with s.get(MAIN_URL) as res:
                val = res.json()
                challenge = val['auth']['challenge']
                key = "admin_password_" + challenge
                enc = hashlib.md5()
                enc.update(key.encode('utf-8'))
                encKey = enc.hexdigest()
                LOGIN_DATA['auth']['authHash'] = encKey
                print(res.url + " : " + str(res.status_code))

            with s.post(MAIN_URL, json=LOGIN_DATA) as res:
                cookies = res.cookies.get_dict()
                val = res.json()
                e_id = val['token']['id']
                headers = {'X-Auth-Token': e_id, 'X-Requested-With': 'XMLHttpRequest', 'Accept-Encoding': 'gzip, deflate'}
                print(res.url + " : " + str(res.status_code))

            with s.put(SETTING_URL, json=SETTING_DATA, headers=headers, cookies=cookies) as res:
                print(res.url + " : " + str(res.status_code))

            with s.post(RESTART_URL) as res:
                print(res.url + " : " + str(res.status_code))

        return "OK"

    except requests.exceptions.Timeout as errd:
        print("Timeout Error : ", errd)
        return "Timeout Error : ", errd

    except requests.exceptions.ConnectionError as errc:
        print("Error Connecting : ", errc)
        return "Error Connecting : ", errc

    except requests.exceptions.HTTPError as errb:
        print("Http Error : ", errb)
        return "Http Error : ", errb

    # Any Error except upper exception
    except requests.exceptions.RequestException as erra:
        print("AnyException : ", erra)
        return "AnyException : ", erra


def allradio_web(ip, ssid, pwd):
    MAIN_URL = f'https://{ip}:8080/admin/login/login'
    SETTING_URL1 = f'https://{ip}:8080/admin/wlan/wlan5gssid1'
    SETTING_URL2 = f'https://{ip}:8080/admin/wlan/wlan2gssid1'

    LOGIN_DATA = {
        'submit': '로그인',
        'loginid': 'admin',
        'loginpwd': 'allradio'
    }
    SETTING_DATA = {
        'wepkeyid': '',
        'enable': '1',
        'ssid': ssid,  # SSID
        'hiddenssid': '1',
        'passphrase': pwd,  # Password
        'viewpasswd': 'on',
        'wpakeyhex': '0',
        'authtype': 'wpapsk',
        'wepauthtype': 'auto',
        'wepkeylength': '5',
        'wepkeyhex': '1',
        'wepkey0': '',
        'dynamicwep': '0',
        'macauth': '0',
        'wpaversion': 'wpa1_2',
        'wpacipher': 'tkip%2Baes',
        'authsvraddr': '10.0.200.200',
        'authsvrport': '1812',
        'authsvrsecret': 'secret',
        'authsvraddr1': '',
        'authsvrport1': '',
        'authsvrsecret1': '',
        'vlanid': '',
        'assoccount': ''
    }

    try:
        with requests.Session() as s:
            with s.post(MAIN_URL, data=LOGIN_DATA, verify=False) as res:
                print(res.url + " : " + str(res.status_code))

            with s.post(SETTING_URL1, data=SETTING_DATA, verify=False) as res:
                print(res.url + " : " + str(res.status_code))

            with s.post(SETTING_URL2, data=SETTING_DATA, verify=False) as res:
                print(res.url + " : " + str(res.status_code))

        return "OK"

    except requests.exceptions.Timeout as errd:
        print("Timeout Error : ", errd)
        return "Timeout Error : ", errd

    except requests.exceptions.ConnectionError as errc:
        print("Error Connecting : ", errc)
        return "Error Connecting : ", errc

    except requests.exceptions.HTTPError as errb:
        print("Http Error : ", errb)
        return "Http Error : ", errb

    # Any Error except upper exception
    except requests.exceptions.RequestException as erra:
        print("AnyException : ", erra)
        return "AnyException : ", erra


def davo_web(ip, ssid, pwd):
    MAIN_URL = f'http://{ip}:88/login'
    SETTING_URL1 = f'http://{ip}:88/setting/wl5_wlan'
    SETTING_URL2 = f'http://{ip}:88/setting/wl24_wlan'
    RESTART_URL = f'http://{ip}:88/system/restart'

    dt = datetime.datetime.today()
    timeval = str(dt.year) + str(dt.month) + str(dt.day) + str(dt.hour) + str(dt.minute) + str(dt.second) + str(dt.microsecond)[0:3]

    LOGIN_DATA = {
        'user_id': 'admin',
        'user_pwd': 'tmakxmdpdj~!@'
    }

    SETTING_DATA1 = {
        'disabled': '0',
        'ssid': ssid,
        'hidden': '1',
        'key': pwd,
        'key1': '',
        'key_type': 'ascii',
        'wep_key_type': 'hex',
        'wep_key_len': '64',
        'wpa_type': 'psk-mixed',
        'rsn_pairwise': 'aes',
        'vap_[]': 'vap00',
        'mode': 'none',
        'sta_auth': '0',
        'max_sta': '512',
        'encryption': 'psk-mixed+aes',
        'wmm': '1',
        'dummyVal': timeval,
        'act': 'set_info',
    }

    SETTING_DATA2 = {
        'disabled': '0',
        'ssid': ssid,
        'hidden': '1',
        'key': pwd,
        'key1': '',
        'key_type': 'ascii',
        'wep_key_type': 'hex',
        'wep_key_len': '64',
        'wpa_type': 'psk-mixed',
        'rsn_pairwise': 'aes',
        'vap_[]': 'vap10',
        'mode': 'none',
        'sta_auth': '0',
        'max_sta': '512',
        'encryption': 'psk-mixed+aes',
        'wmm': '1',
        'dummyVal': timeval,
        'act': 'set_info',
    }

    RESTART_DATA = {
        'dummyVal': timeval,
        'act': 'wifi_restart',
    }

    try:
        with requests.Session() as s:
            with s.post(MAIN_URL, data=LOGIN_DATA) as res:
                print(res.url + " : " + str(res.status_code))

            with s.post(SETTING_URL1, data=SETTING_DATA1) as res:
                print(res.url + " : " + str(res.status_code))

            with s.post(SETTING_URL2, data=SETTING_DATA2) as res:
                print(res.url + " : " + str(res.status_code))

            with s.post(RESTART_URL, data=RESTART_DATA) as res:
                print(res.url + " : " + str(res.status_code))

        return "OK"

    except requests.exceptions.Timeout as errd:
        print("Timeout Error : ", errd)
        return "Timeout Error : ", errd

    except requests.exceptions.ConnectionError as errc:
        print("Error Connecting : ", errc)
        return "Error Connecting : ", errc

    except requests.exceptions.HTTPError as errb:
        print("Http Error : ", errb)
        return "Http Error : ", errb

    # Any Error except upper exception
    except requests.exceptions.RequestException as erra:
        print("AnyException : ", erra)
        return "AnyException : ", erra


if __name__ == "__main__":
    print(form_desc)

    wb = openpyxl.load_workbook(SRC_FILE)
    ws = wb.active
    cnt = 0

    if START_LINE == 0:
        startRow = 2
    else:
        startRow = START_LINE

    if END_LINE == 0:
        totalRows = ws.max_row + 1
    else:
        totalRows = END_LINE

    res = ""

    for i in range(startRow, totalRows):
        val_devName = ws['G' + str(i)].value
        val_vendor = ws['F' + str(i)].value
        val_ip = ws['H' + str(i)].value
        val_port = ws['I' + str(i)].value
        val_ssid = ws['J' + str(i)].value
        val_pass = ws['K' + str(i)].value

        ck = port_check(val_ip, val_port)
        startTime = time.time()

        if ck:
            print(val_devName + " TCP Port Test OK")
            if val_vendor == "가온미디어":
                print("Device Vendor is 가온미디어")
                res = kaon_web(val_ip, val_ssid, val_pass)
            elif val_vendor == "다보링크":
                print("Device Vendor is 다보링크")
                res = davo_web(val_ip, val_ssid, val_pass)
            elif val_vendor == "올레디오":
                print("Device Vendor is 올레디오")
                res = allradio_web(val_ip, val_ssid, val_pass)

            ws['L' + str(i)].value = str(res)

        else:
            ws['L' + str(i)].value = "TCP Port Closed"
            print("TCP Port Closed")

        endTime = time.time()
        print("Remain time : ", endTime - startTime)
        print("=========================================================\n\n")

        cnt += 1
        if cnt > 50:
            wb.save(filename=SRC_FILE)
            cnt = 0

    wb.save(filename=SRC_FILE)
