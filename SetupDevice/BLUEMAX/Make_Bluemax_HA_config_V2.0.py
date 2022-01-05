# from SSHConnector import SSHConnector
import requests
import base64
import openpyxl
import json
import os
import paramiko

from multiprocessing import Process

from Cryptodome.Cipher import AES
from Cryptodome import Random

import random
from datetime import date
import time
import urllib3

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

MAIN_URL = 'https://192.168.10.10'
LOGIN_API = '/api/au/login'
HA_API = '/ha:443:ha_grp_2'
BACKUP_API = '/api/sm/backup/manual'
ROUTING_API = '/api/sm/static-routings/1'
VIP_API = '/api/sm/ha/virtual-ips'
DHCP_API = '/api/sm/dhcp/server/dynamic-areas'
DHCP_API2 = '/api/sm/dhcp/server/config'
DHCP_API3 = '/api/sm/dhcp/server/apply'
EQUIP_API = "/api/sm/info/equipment"

BACKUP_PARAM = {"ha_backup": 1, "target": "POVSL"}

proxy = {'https': 'http://127.0.0.1:8080'}

CUR_PATH = os.getcwd()
SRC_FILE = CUR_PATH + "\\src.xlsx"
START_DATE = date.today()


# ### gw1: KT, gw2: SK
def make_routing_config(gw1, gw2):
    cfg1 = {"type": 0, "dest_addr": "0.0.0.0", "dest_mask": 0, "metric": 0,
            "gw": [
                {"id": f"{str(random.randint(2000, 9999))}_{gw1}", "addr": gw1, "ifc_id": 12, "ifc_name": "eth11",
                 "weight": 1, "itemIndex": 0}
            ],
            "gwTostring": "", "route_type": 0, "status_check": 0, "desc": ""}
    cfg2 = {"type": 0, "dest_addr": "0.0.0.0", "dest_mask": 0, "metric": 0,
            "gw": [
                {"id": f"{str(random.randint(2000, 9999))}_{gw2}", "addr": gw2, "ifc_id": 13, "ifc_name": "eth12",
                 "weight": 1, "itemIndex": 0}
            ],
            "gwTostring": "", "route_type": 0, "status_check": 0, "desc": ""}
    return cfg1, cfg2


def make_vip_config(zone, zone_id, mmbr_id, mmbr_uuid, mmbr_hostname, dev_name, vrid, vip, vip_mask):
    cfg1 = {"zone": zone, "zone_id": zone_id, "ha_mmbr_id": mmbr_id, "mmbr_uuid": mmbr_uuid,
            "mmbr_hostname": mmbr_hostname, "ifc_name": dev_name, "vrid": vrid, "ip_ver": 4,
            "ip": vip, "netmask": vip_mask, "rip": None}
    return cfg1


def make_host_config(hostname):
    hostdata = {"hostname": hostname,
                "location": "KR",
                "desc": ""}
    return hostdata


def make_dhcp_config(start, end, net, subnet, gw, idx):
    cfg1 = {"area_index": idx, "start_area": start, "end_area": end, "net_addr": net,
            "subnet_mask": subnet, "default_gw": gw}
    cfg2 = {"normal_serv_use_enable": 1, "normal_serv_rent_tm": 1440, "normal_serv_auth": 0,
            "dns_serv1": "168.126.63.1", "dns_serv2": "8.8.8.8", "wins_serv1": "", "wins_serv2": ""}
    return cfg1, cfg2


def make_login_data(login_pw, csrf_token):
    login_data = {
        "lang": "ko",
        "force": "true",
        "readonly": 0,
        "login_id": "admin",
        "login_pw": login_pw,
        "csrf_token": csrf_token,
        "type": "local"
    }
    return login_data


def _pad(s):
    bs = 16
    return s + (bs - len(s) % bs) * chr(bs - len(s) % bs)


def write_log(string):
    dat = str(START_DATE).replace("-", "")
    f = open(CUR_PATH + f'\\log_{dat}.txt', "a+")
    now = time.localtime()
    cur_time = "%02d/%02d,%02d:%02d:%02d" % (now.tm_mon, now.tm_mday, now.tm_hour, now.tm_min, now.tm_sec)
    f.write(f'{cur_time}:::{string}\n')
    f.close()


def write_log_ssh(string):
    dat = str(START_DATE).replace("-", "")
    f = open(CUR_PATH + f'\\log_ssh_{dat}.txt', "a+")
    now = time.localtime()
    cur_time = "%02d/%02d,%02d:%02d:%02d" % (now.tm_mon, now.tm_mday, now.tm_hour, now.tm_min, now.tm_sec)
    f.write(f'{cur_time}:::{string}\n')
    f.close()


def ip_calculator(ip):
    subnet = int(ip.split('/')[1])
    ipOctet_A = int(ip.split('/')[0].split('.')[0])
    ipOctet_B = int(ip.split('/')[0].split('.')[1])
    ipOctet_C = int(ip.split('/')[0].split('.')[2])
    # ipOctet_D = int(ip.split('/')[0].split('.')[3])

    ipOctet_Cp = ipOctet_C + (2 ** (24 - subnet)) - 1
    ipOctet_Dp = 254

    vip = f'{str(ipOctet_A)}.{str(ipOctet_B)}.{str(ipOctet_Cp)}.{str(ipOctet_Dp)}/{str(subnet)}'
    rip1 = f'{str(ipOctet_A)}.{str(ipOctet_B)}.{str(ipOctet_Cp)}.{str(ipOctet_Dp - 1)}/{str(subnet)}'
    rip2 = f'{str(ipOctet_A)}.{str(ipOctet_B)}.{str(ipOctet_Cp)}.{str(ipOctet_Dp - 2)}/{str(subnet)}'

    return [vip, rip1, rip2]


def dhcp_calculator(net):
    subnet = int(net.split('/')[1])
    ipOctet_A = int(net.split('/')[0].split('.')[0])
    ipOctet_B = int(net.split('/')[0].split('.')[1])
    ipOctet_C = int(net.split('/')[0].split('.')[2])
    # ipOctet_D = int(net.split('/')[0].split('.')[3])

    ipOctet_Cp = ipOctet_C + (2 ** (24 - subnet)) - 1
    ipOctet_Dp = 254

    val = 256 - 2 ** (24 - subnet)

    startIp = f'{str(ipOctet_A)}.{str(ipOctet_B)}.{str(ipOctet_C)}.1'
    endIp = f'{str(ipOctet_A)}.{str(ipOctet_B)}.{str(ipOctet_Cp)}.100'
    network = f'{str(ipOctet_A)}.{str(ipOctet_B)}.{str(ipOctet_C)}.0'
    netmask = f'255.255.{str(val)}.0'
    gateway = f'{str(ipOctet_A)}.{str(ipOctet_B)}.{str(ipOctet_Cp)}.{str(ipOctet_Dp)}'

    return startIp, endIp, network, netmask, gateway


def ssh_make_config(interface, del_cnt, ip_set):
    cmd = list()
    cmd.append(f'interface {interface}\n')
    for delCnt in range(1, del_cnt + 1):
        cmd.append('no ip add\n')

    for ip in ip_set:
        cmd.append(f'ip add {ip}\n')

    cmd.append('exit\n')
    cmd.append('y\n')

    return cmd


class SSHConnector():
    def __init__(self):
        self.preamble = ['y\n', 'config\n']
        self.hostname = ['hostname\n', 'set hostname\n', 'exit\n']
        self.interface = ['interface eth1\n', 'no ip add\n', 'ip add 100.100.100.100/24\n', 'exit\n', 'y\n']

    def main(self, sw_ip, sw_user, sw_pass, sw_port, command_set):
        host = sw_ip
        username = sw_user
        password = sw_pass

        output = list()

        try:
            conn = paramiko.SSHClient()
            conn.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            conn.connect(host, username=username, password=password, port=sw_port, timeout=100)
            channel = conn.invoke_shell()
            time.sleep(3)
            for command in command_set:
                for line in command:
                    channel.send(line)

                    out_data, err_data = self.wait_streams(channel)
                    output.append(out_data)
                    write_log_ssh(out_data)

            return output

        except Exception as e:
            if "port" in str(e):
                return "Port Error"
            if "WinError 10060" in str(e):
                return "Connection Error"
            print(e)
            return "Error"

        finally:
            if conn is not None:
                conn.close()

    @staticmethod
    def wait_streams(channel):
        time.sleep(1)
        out_data = ""
        err_data = ""
        while True:
            time.sleep(2)
            if channel.recv_ready():
                out_data += channel.recv(1000).decode('ascii')
                if channel.recv_stderr_ready():
                    err_data += channel.recv(1000).decode('ascii')
                if out_data.find("#") != -1 or out_data.find("[y|n]") != -1:
                    break

        return out_data, err_data

    def ssh_connect(self, con_info, ip_cfg):
        print(ip_cfg)
        self.main(con_info, "admin", "secui00@!", 22, ip_cfg)


if __name__ == "__main__":
    # Excel Data Formatting
    wb = openpyxl.load_workbook(SRC_FILE)
    ws = wb.active

    ip_t = []
    ip_s = []
    ip_w = []
    ip_e = []

    ip_t_vip = [1]
    ip_t_rip1 = []
    ip_t_rip2 = []
    ip_s_vip = [2]
    ip_s_rip1 = []
    ip_s_rip2 = []
    ip_w_vip = [3, 4]
    ip_w_rip1 = []
    ip_w_rip2 = []
    ip_e_vip = [5]
    ip_e_rip1 = []
    ip_e_rip2 = []

    for i in range(2, ws.max_row + 1):
        row = str(i)

        # COMMON
        no = ws[f'A{row}'].value
        orgName = ws[f'B{row}'].value
        schName = ws[f'D{row}'].value
        haCls = ws[f'C{row}'].value

        if haCls == "SA":
            continue

        execution = ws[f'T{row}'].value
        exec_date = str(ws[f'S{row}'].value).replace('-', '').split(' ')[0]

        ip_t_delCnt = len(ip_t)
        ip_s_delCnt = len(ip_s)
        ip_w_delCnt = len(ip_w)
        ip_e_delCnt = len(ip_e)

        vip_delCnt = (len(ip_t_vip) + len(ip_s_vip) + len(ip_w_vip) + len(ip_e_vip)) * 2
        if vip_delCnt != 0:
            vip_delCnt += 2

        ip_t = []
        ip_s = []
        ip_w = []
        ip_e = []

        ifCfg_M = []
        ifCfg_S = []

        ip_t_vip = []
        ip_t_rip1 = []
        ip_t_rip2 = []
        ip_s_vip = []
        ip_s_rip1 = []
        ip_s_rip2 = []
        ip_w_vip = []
        ip_w_rip1 = []
        ip_w_rip2 = []
        ip_e_vip = []
        ip_e_rip1 = []
        ip_e_rip2 = []

        # IF Standalone
        if haCls == "HA":
            con_info = ['192.168.10.10', '192.168.10.11']
            m_hostname = ws[f'E{row}'].value + "_FW1"
            s_hostname = ws[f'E{row}'].value + "_FW2"

            # [10.10.10.0/24, 20.20.20.0/24]
            if ws[f'F{row}'].value != "없음":
                ip_t = ws[f'F{row}'].value.split(',')
            if ws[f'G{row}'].value != "없음":
                ip_s = ws[f'G{row}'].value.split(',')
            if ws[f'I{row}'].value != "없음":
                ip_w = ws[f'I{row}'].value.split(',')
            if ws[f'H{row}'].value != "없음":
                ip_w += ws[f'H{row}'].value.split(',')
                dhcp_start, dhcp_end, dhcp_net, dhcp_mask, dhcp_gateway = dhcp_calculator(ws[f'H{row}'].value)
            if ws[f'J{row}'].value != "없음":
                ip_e = ws[f'J{row}'].value.split(',')

            ip_skL3 = ws[f'K{row}'].value + "/29"
            ip_skFw = ws[f'L{row}'].value + "/29"
            ip_ktL3 = ws[f'M{row}'].value + "/28"
            ip_ktFw_vip = ws[f'N{row}'].value + "/28"
            ip_ktFw_rip1 = ws[f'O{row}'].value + "/28"
            ip_ktFw_rip2 = ws[f'P{row}'].value + "/28"

            for val in ip_t:
                ip_res = ip_calculator(val)
                ip_t_vip.append(ip_res[0])
                ip_t_rip1.append(ip_res[1])
                ip_t_rip2.append(ip_res[2])

            for val in ip_s:
                ip_res = ip_calculator(val)
                ip_s_vip.append(ip_res[0])
                ip_s_rip1.append(ip_res[1])
                ip_s_rip2.append(ip_res[2])

            for val in ip_w:
                ip_res = ip_calculator(val)
                ip_w_vip.append(ip_res[0])
                ip_w_rip1.append(ip_res[1])
                ip_w_rip2.append(ip_res[2])

            for val in ip_e:
                ip_res = ip_calculator(val)
                ip_e_vip.append(ip_res[0])
                ip_e_rip1.append(ip_res[1])
                ip_e_rip2.append(ip_res[2])

            ifCfg_M.append(['y\n', 'config\n'])
            ifCfg_S.append(['y\n', 'config\n'])

            ifCfg_M.append(ssh_make_config('eth1', 1 + ip_t_delCnt, ip_t_rip1))
            ifCfg_S.append(ssh_make_config('eth1', 1 + ip_t_delCnt, ip_t_rip2))

            ifCfg_M.append(ssh_make_config('eth2', 1 + ip_s_delCnt, ip_s_rip1))
            ifCfg_S.append(ssh_make_config('eth2', 1 + ip_s_delCnt, ip_s_rip2))

            ifCfg_M.append(ssh_make_config('eth3', 1 + ip_e_delCnt, ip_e_rip1))
            ifCfg_S.append(ssh_make_config('eth3', 1 + ip_e_delCnt, ip_e_rip2))

            ifCfg_M.append(ssh_make_config('eth9', 1 + ip_w_delCnt, ip_w_rip1))
            ifCfg_S.append(ssh_make_config('eth9', 1 + ip_w_delCnt, ip_w_rip2))

            ifCfg_M.append(ssh_make_config('eth11', 1, [ip_skFw]))

            ifCfg_M.append(ssh_make_config('eth12', 1, [ip_ktFw_rip1]))
            ifCfg_S.append(ssh_make_config('eth12', 1, [ip_ktFw_rip2]))

        else:
            continue

        # for c1 in ifCfg_M:
        #     print(c1)
        # for c2 in ifCfg_S:
        #     print(c2)

        with requests.Session() as s:
            # WEB GUI Initialize
            with s.get(MAIN_URL, verify=False) as res:
                val = res.text
                find_num = val.find("csrf_token")
                csrf_token = val[find_num + 51:find_num + 115]
                pw = "secui00@!"
                iv = Random.new().read(AES.block_size)

                key = bytes(csrf_token[:32], 'utf-8')
                raw = bytes(_pad(pw), 'utf-8')
                cipher = AES.new(key, AES.MODE_CBC, iv)
                hex_data = (iv.hex() + cipher.encrypt(raw).hex()).encode('utf-8')
                login_pw = base64.b64encode(hex_data).decode('utf-8')

                login_data = make_login_data(login_pw, csrf_token)
                # write_log(f'{no}_{schName} : {res.url} / {res.text}')

            # Login Phase
            with s.post(MAIN_URL + LOGIN_API, json=login_data, verify=False) as res:
                cookies = res.cookies.get_dict()
                res_dict = json.loads(res.text)
                auth_key = res_dict['result']['api_token']
                headers = {'Authorization': auth_key}

            # VIP Delete
            for j in range(0, vip_delCnt):
                with s.delete(MAIN_URL + VIP_API + '/' + str(j + 1), headers=headers,
                              verify=False) as res:
                    write_log(f'{no}_{schName} : {res.url} / {res.text}')

            with s.put(MAIN_URL + VIP_API + '/apply', headers=headers, verify=False) as res:
                write_log(f'{no}_{schName} : {res.url} / {res.text}')

            host_data1 = make_host_config(m_hostname)
            with s.put(MAIN_URL + EQUIP_API, json=host_data1, headers=headers, verify=False) as res:
                write_log(f'{no}_{schName} : {res.url}')

            host_data2 = make_host_config(s_hostname)
            with s.put(MAIN_URL + HA_API + EQUIP_API, json=host_data2, headers=headers, verify=False) as res:
                write_log(f'{no}_{schName} : {res.url}')

            s.close()

            time.sleep(5)

            # INTERFACE & HOSTNAME Setting
            ssh = SSHConnector()
            th1 = Process(target=ssh.ssh_connect, args=(con_info[0], ifCfg_M))

            ssh2 = SSHConnector()
            th2 = Process(target=ssh2.ssh_connect, args=(con_info[1], ifCfg_S))

            th1.start()
            time.sleep(30)
            th2.start()

            th1.join()
            th2.join()

        with requests.Session() as s:
            # WEB GUI Initialize
            with s.get(MAIN_URL, verify=False) as res:
                val = res.text
                find_num = val.find("csrf_token")
                csrf_token = val[find_num + 51:find_num + 115]
                pw = "secui00@!"
                iv = Random.new().read(AES.block_size)

                key = bytes(csrf_token[:32], 'utf-8')
                raw = bytes(_pad(pw), 'utf-8')
                cipher = AES.new(key, AES.MODE_CBC, iv)
                hex_data = (iv.hex() + cipher.encrypt(raw).hex()).encode('utf-8')
                login_pw = base64.b64encode(hex_data).decode('utf-8')

                login_data = make_login_data(login_pw, csrf_token)

            # Login Phase
            with s.post(MAIN_URL + LOGIN_API, json=login_data, verify=False) as res:
                cookies = res.cookies.get_dict()
                res_dict = json.loads(res.text)
                auth_key = res_dict['result']['api_token']
                headers = {'Authorization': auth_key}

            # VIP Setting
            vrid = 0
            zoneid = 0
            vip_cfg = []
            for t in ip_t_vip:
                dev_name = 'eth1'
                vrid += 1
                zoneid += 1
                vip = t.split('/')[0]
                vip_mask = int(t.split('/')[1])
                zone = 1

                mmbr_id = 1
                mmbr_uuid = 'ha_grp_1'
                mmbr_hostname = m_hostname
                vip_cfg.append(make_vip_config(zone, zoneid, mmbr_id, mmbr_uuid, mmbr_hostname,
                                               dev_name, vrid, vip, vip_mask))

                mmbr_id = 2
                mmbr_uuid = 'ha_grp_2'
                mmbr_hostname = s_hostname
                vip_cfg.append(make_vip_config(zone, zoneid, mmbr_id, mmbr_uuid, mmbr_hostname,
                                               dev_name, vrid, vip, vip_mask))

            for st in ip_s_vip:
                dev_name = 'eth2'
                vrid += 1
                zoneid += 1
                vip = st.split('/')[0]
                vip_mask = int(st.split('/')[1])
                zone = 1

                mmbr_id = 1
                mmbr_uuid = 'ha_grp_1'
                mmbr_hostname = m_hostname
                vip_cfg.append(make_vip_config(zone, zoneid, mmbr_id, mmbr_uuid, mmbr_hostname,
                                               dev_name, vrid, vip, vip_mask))

                mmbr_id = 2
                mmbr_uuid = 'ha_grp_2'
                mmbr_hostname = s_hostname
                vip_cfg.append(make_vip_config(zone, zoneid, mmbr_id, mmbr_uuid, mmbr_hostname,
                                               dev_name, vrid, vip, vip_mask))

            for e in ip_e_vip:
                dev_name = 'eth3'
                vrid += 1
                zoneid += 1
                vip = e.split('/')[0]
                vip_mask = int(e.split('/')[1])
                zone = 1

                mmbr_id = 1
                mmbr_uuid = 'ha_grp_1'
                mmbr_hostname = m_hostname
                vip_cfg.append(make_vip_config(zone, zoneid, mmbr_id, mmbr_uuid, mmbr_hostname,
                                               dev_name, vrid, vip, vip_mask))

                mmbr_id = 2
                mmbr_uuid = 'ha_grp_2'
                mmbr_hostname = s_hostname
                vip_cfg.append(make_vip_config(zone, zoneid, mmbr_id, mmbr_uuid, mmbr_hostname,
                                               dev_name, vrid, vip, vip_mask))

            for w in ip_w_vip:
                dev_name = 'eth9'
                vrid += 1
                zoneid += 1
                vip = w.split('/')[0]
                vip_mask = int(w.split('/')[1])
                zone = 1

                mmbr_id = 1
                mmbr_uuid = 'ha_grp_1'
                mmbr_hostname = m_hostname
                vip_cfg.append(make_vip_config(zone, zoneid, mmbr_id, mmbr_uuid, mmbr_hostname,
                                               dev_name, vrid, vip, vip_mask))

                mmbr_id = 2
                mmbr_uuid = 'ha_grp_2'
                mmbr_hostname = s_hostname
                vip_cfg.append(make_vip_config(zone, zoneid, mmbr_id, mmbr_uuid, mmbr_hostname,
                                               dev_name, vrid, vip, vip_mask))

            zoneid += 1
            dev_name = 'eth12'
            vrid += 1
            vip = ip_ktFw_vip.split('/')[0]
            vip_mask = int(ip_ktFw_vip.split('/')[1])
            zone = 2

            mmbr_id = 1
            mmbr_uuid = 'ha_grp_1'
            mmbr_hostname = m_hostname
            vip_cfg.append(make_vip_config(zone, zoneid, mmbr_id, mmbr_uuid, mmbr_hostname,
                                           dev_name, vrid, vip, vip_mask))

            mmbr_id = 2
            mmbr_uuid = 'ha_grp_2'
            mmbr_hostname = s_hostname
            vip_cfg.append(make_vip_config(zone, zoneid, mmbr_id, mmbr_uuid, mmbr_hostname,
                                           dev_name, vrid, vip, vip_mask))

            for cfg in vip_cfg:
                with s.post(MAIN_URL + VIP_API, json=cfg, headers=headers, verify=False) as res:
                    write_log(f'{no}_{schName} : {res.url} / {res.text}')

            with s.put(MAIN_URL + VIP_API + '/apply', headers=headers, verify=False) as res:
                write_log(f'{no}_{schName} : {res.url} / {res.text}')

            # Routing Setting
            ip_ktGw = ip_ktL3.split('/')[0]
            ip_skGw = ip_skL3.split('/')[0]
            rt_cfg1, rt_cfg2 = make_routing_config(ip_skGw, ip_ktGw)

            with s.put(MAIN_URL + ROUTING_API, json=rt_cfg1, headers=headers, verify=False) as res:
                write_log(f'{no}_{schName} : {res.url} / {res.text}')

            with s.put(MAIN_URL + HA_API + ROUTING_API, json=rt_cfg2, headers=headers, verify=False) as res:
                write_log(f'{no}_{schName} : {res.url} / {res.text}')

            # DHCP Setting
            if ip_w:
                with s.get(MAIN_URL + DHCP_API, headers=headers, verify=False) as res:
                    write_log(f'{no}_{schName} : {res.url} / {res.text}')
                    var = res.json()

                    idx = var['result'][0]['area_index']
                    dhcp_cfg1, dhcp_cfg2 = make_dhcp_config(dhcp_start, dhcp_end, dhcp_net, dhcp_mask, dhcp_gateway, idx)

                with s.put(MAIN_URL + DHCP_API + '/' + str(idx), json=dhcp_cfg1, headers=headers,
                           verify=False) as res:
                    write_log(f'{no}_{schName} : {res.url} / {res.text}')

                with s.put(MAIN_URL + DHCP_API2, json=dhcp_cfg2, headers=headers, verify=False) as res:
                    write_log(f'{no}_{schName} : {res.url} / {res.text}')

                with s.put(MAIN_URL + DHCP_API3, headers=headers, verify=False) as res:
                    write_log(f'{no}_{schName} : {res.url} / {res.text}')

                with s.get(MAIN_URL + HA_API + DHCP_API, headers=headers, verify=False) as res:
                    write_log(f'{no}_{schName} : {res.url} / {res.text}')
                    var = res.json()

                    idx = var['result'][0]['area_index']
                    dhcp_cfg1, dhcp_cfg2 = make_dhcp_config(dhcp_start, dhcp_end, dhcp_net, dhcp_mask, dhcp_gateway, idx)

                with s.put(MAIN_URL + HA_API + DHCP_API + '/' + str(idx), json=dhcp_cfg1, headers=headers,
                           verify=False) as res:
                    write_log(f'{no}_{schName} : {res.url} / {res.text}')

                with s.put(MAIN_URL + HA_API + DHCP_API2, json=dhcp_cfg2, headers=headers,
                           verify=False) as res:
                    write_log(f'{no}_{schName} : {res.url} / {res.text}')

                with s.put(MAIN_URL + HA_API + DHCP_API3, headers=headers, verify=False) as res:
                    write_log(f'{no}_{schName} : {res.url} / {res.text}')

            # Backup File Download
            with s.post(MAIN_URL + BACKUP_API, json=BACKUP_PARAM, headers=headers, verify=False) as res:
                write_log(f'{no}_{schName} : {res.url} / {res.headers}')
                if not (os.path.isdir(CUR_PATH + f'\\{execution}')):
                    os.makedirs(CUR_PATH + f'\\{execution}')

                if not (os.path.isdir(CUR_PATH + f'\\{execution}\\{exec_date}')):
                    os.makedirs(CUR_PATH + f'\\{execution}\\{exec_date}')
                f = open(f"{CUR_PATH}\\{execution}\\{exec_date}\\{no}_{orgName}_{schName}.tar", 'wb',)
                f.write(res.content)
                f.close()
            s.close()

