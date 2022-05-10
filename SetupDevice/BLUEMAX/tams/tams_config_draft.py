import requests
import base64
import openpyxl
import json
import os, sys
import random, string
import socket
import hashlib

from requests_toolbelt import MultipartEncoder
from Cryptodome.Cipher import AES
from Cryptodome import Random

from datetime import date
import time
import urllib3
import random

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

CUR_PATH = os.getcwd()
START_DATE = date.today()

proxy = {'https': 'http://127.0.0.1:8080'}

LOGIN_API = '/api/au/login'
login_json = {"username": "secuiadmin",
              "password": "YmFhNjBhNTE1ZjYwZmI4Nzg2Njk2ODMyYjFmYzQzNzNjNjY1ZDA5ZjA5Mjk3Y2I3MzNlNmYzZjMwOGIzMzg0MQ==",
              "force": "True",
              "ip_force": "True",
              "csrf_token": "0aafba5acfd72972f27044c2eaa72dfeeeb0c26ae55a3a5410b750f1a2a0bb71"}

DEVICE_LIST_API = '/api/sm/devices-tree?only_group_yn=0&only_config_yn=0&tams_yn=1&slave_yn=1&tree_type=0&only_policy_mng_yn=0&only_tams_yn=0'

CHECK_IN_API = '/api/cm/check-in'
check_in_json = {"admin_id": 3,
                 "menu_id": 110201,
                 "sub_id": 0}

CHECK_IN_API2 = '/api/cm/check-in?admin_id=3&menu_id=110201&sub_id=0'

DEVICE_ADD_API = '/api/sm/devices'
device_add_json1 = {"parent_device_id": 8,
                    "device_uname": "서울특별시교육청강서도서관_이용자망",
                    "device_type": 1,
                    "device_desc": "",
                    "device_ip": "10.138.100.86",
                    "device_api_port": 50005,
                    "device_mng_type": 1,
                    "device_auto_add_yn": 0,
                    "device_login_uname": "tamsadm01",
                    "device_login_pw": "cWhka3M5OUAh"}
device_add_json2 = {"parent_device_id": 8,
                    "device_uname": "서울특별시교육청강서도서관_이용자망",
                    "device_type": 3,
                    "device_desc": "",
                    "device_ip": "10.138.100.86",
                    "device_api_port": 50005,
                    "device_mng_type": 1}

DEVICE_APPLY_API = '/api/sm/devices/apply'


def write_log(string):
    dat = str(START_DATE).replace("-", "")
    f = open(CUR_PATH + f'\\log_{dat}.txt', "a+")
    now = time.localtime()
    cur_time = "%02d/%02d,%02d:%02d:%02d" % (now.tm_mon, now.tm_mday, now.tm_hour, now.tm_min, now.tm_sec)
    f.write(f'{cur_time};{string}\n')
    f.close()


def port_check(ip, port):
    try:
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(2)
        result = sock.connect_ex((ip, port))
        sock.close()
    except socket.error as e:
        result = 1
        print("Error : " + e)

    if result == 0:
        return True
    else:
        return False


def _pad(s):
    bs = 16
    return s + (bs - len(s) % bs) * chr(bs - len(s) % bs)


def print_progress(iteration, total, prefix='', suffix='', decimals=1, barLength=100):
    formatStr = "{0:." + str(decimals) + "f}"
    percent = formatStr.format(100 * (iteration / float(total)))
    filledLength = int(round(barLength * iteration / float(total)))
    bar = '#' * filledLength + '-' * (barLength - filledLength)
    sys.stdout.write('\r%s |%s| %s%s %s' % (prefix, bar, percent, '%', suffix)),
    if iteration == total:
        sys.stdout.write('\n')
    sys.stdout.flush()


if __name__ == "__main__":
    SRC_FILE = CUR_PATH + "\\tams_configuration.xlsx"
    wb = openpyxl.load_workbook(SRC_FILE)
    ws = wb.active

    mainUrl = ws[f'I1'].value
    userId = ws[f'J1'].value
    userPass = ws[f'K1'].value

    for row in range(2, ws.max_row + 1):
        schNo = ws[f'A{row}'].value
        schOrg = ws[f'B{row}'].value
        schHaCls = ws[f'C{row}'].value
        schName = ws[f'D{row}'].value
        schUtmIp1 = ws[f'E{row}'].value
        schUtmIp2 = ws[f'F{row}'].value
        schUtmAccessPort = ws[f'G{row}'].value

        if str(schName) == "None":
            continue

        haChk = False
        if schHaCls == "HA":
            haChk = True

        device_add_json1['device_uname'] = schName + "_FW1"
        device_add_json1['parent_device_id'] = parent_device_id
        device_add_json1['device_ip'] = schUtmIp1
        device_add_json1['device_api_port'] = int(schUtmAccessPort)

        if row % 20 == 2:
            with requests.Session() as s:
                # WEB GUI Initialize
                with s.get(mainUrl, verify=False) as res:
                    val = res.text
                    find_num = val.find("csrf_token")
                    csrf_token = val[find_num + 51:find_num + 115]
                    pw = userPass
                    iv = Random.new().read(AES.block_size)

                    key = bytes(csrf_token[:32], 'utf-8')
                    raw = bytes(_pad(pw), 'utf-8')
                    cipher = AES.new(key, AES.MODE_CBC, iv)
                    hex_data = (iv.hex() + cipher.encrypt(raw).hex()).encode('utf-8')
                    login_pw = base64.b64encode(hex_data).decode('utf-8')

                    login_json['password'] = login_pw
                    login_json['username'] = userId
                    login_json['csrf_token'] = csrf_token
                    print(csrf_token)

                # Login Phase
                with s.post(mainUrl + LOGIN_API, json=login_json, verify=False) as res:
                    res_dict = json.loads(res.text)
                    print(res_dict)
                    if res_dict['code'] == "ok":
                        cookies = res.cookies.get_dict()
                        auth_key = res_dict['result']['api_token']
                        secui_helper_key = res_dict['result']['secui_helper_key']
                        headers = {'Authorization': auth_key}
                        write_log(f"{LOGIN_API};OK;")
                    else:
                        write_log(f"{LOGIN_API};{res_dict['dev_t']};{res_dict['result']['msg']}")

                with s.get(mainUrl + DEVICE_LIST_API, headers=headers, verify=False) as res:
                    res_dict = json.loads(res.text)
                    print(res_dict)
                    if res_dict['code'] == "ok":
                        for values in res_dict['result']:
                            if values['device_uname'] == '이용기관':
                                parent_device_id = values['device_id']

                        write_log(f"{DEVICE_LIST_API};OK;")
                    else:
                        write_log(f"{DEVICE_LIST_API};{res_dict['dev_t']};{res_dict['result']['msg']}")

                with s.post(mainUrl + CHECK_IN_API, headers=headers, json=check_in_json, verify=False) as res:
                    res_dict = json.loads(res.text)
                    print(res_dict)
                    if res_dict['code'] == "ok":
                        write_log(f"{CHECK_IN_API};OK;")
                    else:
                        write_log(f"{CHECK_IN_API};{res_dict['dev_t']};{res_dict['result']['msg']}")

                with s.get(mainUrl + CHECK_IN_API2, headers=headers, verify=False) as res:
                    res_dict = json.loads(res.text)
                    print(res_dict)
                    if res_dict['code'] == "ok":
                        write_log(f"{CHECK_IN_API2};OK;")
                    else:
                        write_log(f"{CHECK_IN_API2};{res_dict['dev_t']};{res_dict['result']['msg']}")

            with s.post(mainUrl + DEVICE_ADD_API, headers=headers, json=device_add_json1, verify=False) as res:
                res_dict = json.loads(res.text)
                print(res_dict)
                if res_dict['code'] == "ok":
                    master_device_id = res_dict['result']['device_id']
                    write_log(f"{schNo}_{schName};{schUtmIp1};{DEVICE_ADD_API};OK;")
                else:
                    write_log(f"{schNo}_{schName};{schUtmIp1};{DEVICE_ADD_API};{res_dict['dev_t']};{res_dict['result']['msg']}")
                    continue

            if haChk:
                device_add_json2['device_uname'] = schName + "_FW2"
                device_add_json2['parent_device_id'] = master_device_id
                device_add_json2['device_ip'] = schUtmIp2
                device_add_json2['device_api_port'] = int(schUtmAccessPort)

                with s.post(mainUrl + DEVICE_ADD_API, headers=headers, json=device_add_json2, verify=False) as res:
                    res_dict = json.loads(res.text)
                    print(res_dict)
                    if res_dict['code'] == "ok":

                        write_log(f"{schNo}_{schName};{schUtmIp2};{DEVICE_ADD_API};OK;")
                    else:
                        write_log(f"{schNo}_{schName};{schUtmIp2};{DEVICE_ADD_API};{res_dict['dev_t']};{res_dict['result']['msg']}")
                        continue

        if row % 20 == 1:
            with s.get(mainUrl + DEVICE_APPLY_API, headers=headers, verify=False) as res:
                res_dict = json.loads(res.text)
                print(res_dict)
                if res_dict['code'] == "ok":

                    write_log(f"{schNo}_{schName};{schUtmIp1};{DEVICE_APPLY_API};OK;")
                else:
                    write_log(
                        f"{schNo}_{schName};{schUtmIp1};{DEVICE_APPLY_API};{res_dict['dev_t']};{res_dict['result']['msg']}")

            s.close()

    with s.get(mainUrl + DEVICE_APPLY_API, headers=headers, verify=False) as res:
        res_dict = json.loads(res.text)
        print(res_dict)
        if res_dict['code'] == "ok":

            write_log(f"{schNo}_{schName};{schUtmIp1};{DEVICE_APPLY_API};OK;")
        else:
            write_log(
                f"{schNo}_{schName};{schUtmIp1};{DEVICE_APPLY_API};{res_dict['dev_t']};{res_dict['result']['msg']}")