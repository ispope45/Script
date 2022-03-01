import openpyxl
import os
import socket
import paramiko


from datetime import date
import time

CUR_PATH = os.getcwd()
SRC_FILE = CUR_PATH + "\\fwList.xlsx"
START_DATE = date.today()

proxy = {'https': 'http://127.0.0.1:8080'}


def ssh_make_config(gateway):
    cmd = list()
    cmd.append(f'y\n')
    cmd.append(f'config\n')
    cmd.append(f'ip-route\n')
    cmd.append(f'add default metric 10 gw {gateway} dev eth12\n')
    cmd.append(f'delete default\n')
    cmd.append(f'\n')
    return cmd


def port_check(ip, port):
    try:
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(1)
        result = sock.connect_ex((ip, port))
        sock.close()
    except socket.error as e:
        result = 1
        print("Error : " + e)

    if result == 0:
        return True
    else:
        return False


def write_log(string):
    dat = str(START_DATE).replace("-", "")
    f = open(CUR_PATH + f'\\log_{dat}.txt', "a+")
    now = time.localtime()
    cur_time = "%02d/%02d,%02d:%02d:%02d" % (now.tm_mon, now.tm_mday, now.tm_hour, now.tm_min, now.tm_sec)
    f.write(f'{cur_time} ::: {string}\n')
    f.close()


def write_log_ssh(string):
    dat = str(START_DATE).replace("-", "")
    f = open(CUR_PATH + f'\\log_ssh_{dat}.txt', "a+")
    now = time.localtime()
    cur_time = "%02d/%02d,%02d:%02d:%02d" % (now.tm_mon, now.tm_mday, now.tm_hour, now.tm_min, now.tm_sec)
    f.write(f'{cur_time}:::{string}\n')
    f.close()


class SSHConnector():
    def __init__(self):
        self.preamble = ['y\n', 'config\n']

    def main(self, sw_ip, sw_user, sw_pass, sw_port, command):
        host = sw_ip
        username = sw_user
        password = sw_pass

        output = list()

        try:
            conn = paramiko.SSHClient()
            conn.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            conn.connect(host, username=username, password=password, port=sw_port, timeout=100)
            channel = conn.invoke_shell()
            time.sleep(2)

            for line in command:
                channel.send(line)

                if line != 'delete default\n':
                    out_data, err_data = self.wait_streams(channel)
                    output.append(out_data)
                    write_log_ssh(out_data)

            return output

        except Exception as e:
            if "port" in str(e):
                return "Port Error"
            if "WinError 10060" in str(e):
                return "Connection Error"
            print(e)
            return "Error"

        finally:
            if conn is not None:
                conn.close()

    @staticmethod
    def wait_streams(channel):
        time.sleep(1)
        out_data = ""
        err_data = ""

        while channel.recv_ready():
            out_data += channel.recv(1000).decode('ascii')
        while channel.recv_stderr_ready():
            err_data += channel.recv(1000).decode('ascii')

        # while True:
        #     time.sleep(2)
        #     if channel.recv_ready():
        #         out_data += channel.recv(1000).decode('ascii')
        #     if channel.recv_stderr_ready():
        #         err_data += channel.recv(1000).decode('ascii')
        #         if out_data.find("#") != -1 or out_data.find("[y|n]") != -1:
        #             break

        return out_data, err_data

    def ssh_connect(self, con_info, rt_conf):
        self.main(con_info, "admin", "secui00@!", 22, rt_conf)


if __name__ == "__main__":

    wb = openpyxl.load_workbook(SRC_FILE)
    ws = wb.active
    cnt = 0

    for row in range(2, ws.max_row + 1):
        schNo = ws[f'A{row}'].value
        schHaCls = ws[f'B{row}'].value
        schName = ws[f'D{row}'].value
        schUtmIp = ws[f'E{row}'].value
        schKTL3Ip = ws[f'F{row}'].value
        schKT_UTM_sIp = ws[f'G{row}'].value

        if schUtmIp and schKTL3Ip and schKT_UTM_sIp:
            skFwIp = schUtmIp
            ktFw2Ip = schKT_UTM_sIp
            ktL3Ip = schKTL3Ip
        else:
            write_log(f"{schNo}_{schName} / Ip address Empty")
            ws[f'I{row}'].value = f"Ip address Empty"
            continue

        if port_check(ktFw2Ip, 22) and port_check(ktFw2Ip, 443):
            if port_check(skFwIp, 22):
                print(f'{schNo}_{schName} Processing...')
            else:
                write_log(f"{schNo}_{schName} / SK 22 Port Closed")
                ws[f'I{row}'].value = f"SK 22 Port Closed"
                continue
        else:
            write_log(f"{schNo}_{schName} / KT Slave 22,443 Port Closed")
            ws[f'I{row}'].value = f"KT Slave 22,443 Port Closed"
            continue

        rt_conf = ssh_make_config(schKTL3Ip)

        ssh = SSHConnector()
        ssh.ssh_connect(schUtmIp, rt_conf)

        ws[f'I{row}'].value = "Success"

        cnt += 1
        if cnt % 5 == 0:
            wb.save(SRC_FILE)

    wb.save(SRC_FILE)

