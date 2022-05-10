import openpyxl
import os, sys


def print_progress(iteration, total, prefix='', suffix='', decimals=1, barLength=100):
    formatStr = "{0:." + str(decimals) + "f}"
    percent = formatStr.format(100 * (iteration / float(total)))
    filledLength = int(round(barLength * iteration / float(total)))
    bar = '#' * filledLength + '-' * (barLength - filledLength)
    sys.stdout.write('\r%s |%s| %s%s %s' % (prefix, bar, percent, '%', suffix)),
    if iteration == total:
        sys.stdout.write('\n')
    sys.stdout.flush()


if __name__ == "__main__":
    CUR_PATH = os.getcwd()
    base_file = CUR_PATH + "\\ip_list.xlsx"
    base_wb = openpyxl.load_workbook(base_file)
    base_ws = base_wb.active

    for row in range(2, base_ws.max_row + 1):
        schNo = base_ws[f'A{row}'].value
        schOrg = base_ws[f'B{row}'].value
        schHaCls = base_ws[f'C{row}'].value
        schName = base_ws[f'D{row}'].value

        utm_rip1 = base_ws[f'E{row}'].value
        utm_rip2 = base_ws[f'F{row}'].value

        sen_nms = base_ws[f'G{row}'].value
        nia_nms = base_ws[f'H{row}'].value

        proc_file = CUR_PATH + "\\blm.xlsx"
        proc_wb = openpyxl.load_workbook(proc_file)
        proc_ws = proc_wb.active

        if str(utm_rip1) != "None":
            proc_ws['Q4'].value = utm_rip1
            proc_ws['Q6'].value = utm_rip1

        if str(utm_rip2) != "None":
            proc_ws['Q5'].value = utm_rip2
            proc_ws['Q7'].value = utm_rip2
        else:
            proc_ws['A5'].value = ""

        proc_ws['K6'].value = sen_nms
        proc_ws['K7'].value = nia_nms

        proc_wb.save(CUR_PATH + f"\\{schNo}_{schName}.xlsx")

        print_progress(row - 1, base_ws.max_row - 1, 'Progress:', 'Complete ', 1, 50)









