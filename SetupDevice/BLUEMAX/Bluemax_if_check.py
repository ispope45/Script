import requests
import base64
import openpyxl
import json
import os, sys

from Cryptodome.Cipher import AES
from Cryptodome import Random

from datetime import date
import time
import urllib3

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

LOGIN_API = '/api/au/login'
HA_API = '/ha:443:ha_grp_2'
INTERFACE_API = '/api/sm/interfaces'
PING_API = '/api/co/tools/ping'

CUR_PATH = os.getcwd()
SRC_FILE = CUR_PATH + "\\FwTestForm.xlsx"
START_DATE = date.today()

proxy = {'https': 'http://127.0.0.1:8080'}


def make_login_data(login_pw, csrf_token):
    login_data = {
        "lang": "ko",
        "force": "true",
        "readonly": 0,
        "login_id": "admin",
        "login_pw": login_pw,
        "csrf_token": csrf_token,
        "type": "local"
    }
    return login_data


def make_ping_target(target_ip):
    ping_data = {
        "target": target_ip
    }
    return ping_data


def _pad(s):
    bs = 16
    return s + (bs - len(s) % bs) * chr(bs - len(s) % bs)


def printProgress(iteration, total, prefix='', suffix='', decimals=1, barLength=100):
    formatStr = "{0:." + str(decimals) + "f}"
    percent = formatStr.format(100 * (iteration / float(total)))
    filledLength = int(round(barLength * iteration / float(total)))
    bar = '#' * filledLength + '-' * (barLength - filledLength)
    sys.stdout.write('\r%s |%s| %s%s %s' % (prefix, bar, percent, '%', suffix)),
    if iteration == total:
        sys.stdout.write('\n')
    sys.stdout.flush()


def write_log(string):
    dat = str(START_DATE).replace("-", "")
    f = open(CUR_PATH + f'\\log_{dat}.txt', "a+")
    now = time.localtime()
    cur_time = "%02d/%02d,%02d:%02d:%02d" % (now.tm_mon, now.tm_mday, now.tm_hour, now.tm_min, now.tm_sec)
    f.write(f'{cur_time} ::: {string}\n')
    f.close()


if __name__ == "__main__":

    wb = openpyxl.load_workbook(SRC_FILE)
    ws = wb.active
    proc_wb = openpyxl.Workbook()
    proc_ws = proc_wb.active

    proc_row = 1
    cnt = 0
    for row in range(2, ws.max_row + 1):
        schNo = ws[f'A{row}'].value
        schHaCls = ws[f'B{row}'].value
        schName = ws[f'D{row}'].value
        schUtmIp = ws[f'E{row}'].value
        schKTL3Ip = ws[f'F{row}'].value
        schComCls = ws[f'G{row}'].value

        if schComCls != "O":
            continue

        if schUtmIp:
            fw_ip = schUtmIp
        else:
            write_log(f"{schNo}_{schName} / Ip address Empty")
            continue

        mainUrl = f"https://{fw_ip}"
        pingTarget = schKTL3Ip

        printProgress(row, ws.max_row, 'Progress:', 'Complete ', 1, 50)

        try:
            with requests.Session() as s:
                # WEB GUI Initialize
                with s.get(mainUrl, verify=False) as res:
                    val = res.text
                    find_num = val.find("csrf_token")
                    csrf_token = val[find_num + 51:find_num + 115]
                    pw = "secui00@!"
                    iv = Random.new().read(AES.block_size)

                    key = bytes(csrf_token[:32], 'utf-8')
                    raw = bytes(_pad(pw), 'utf-8')
                    cipher = AES.new(key, AES.MODE_CBC, iv)
                    hex_data = (iv.hex() + cipher.encrypt(raw).hex()).encode('utf-8')
                    login_pw = base64.b64encode(hex_data).decode('utf-8')

                    login_data = make_login_data(login_pw, csrf_token)

                # Login Phase
                with s.post(mainUrl + LOGIN_API, json=login_data, verify=False) as res:
                    cookies = res.cookies.get_dict()
                    res_dict = json.loads(res.text)
                    auth_key = res_dict['result']['api_token']
                    headers = {'Authorization': auth_key}
        except Exception as e:
            write_log(f"{schNo}_{schName} / Connection Error : {e}")
            continue

        # Interface Information gathering
        with s.get(mainUrl + INTERFACE_API, headers=headers, verify=False) as res:
            res_dict = json.loads(res.text)
            M_ifInfo = {}
            for item in res_dict['result']:
                M_ifInfo[item['name'] + '_ip_addr'] = item['ipv4address']
                M_ifInfo[item['name'] + '_link_state'] = item['link_up']

        if schHaCls == "HA":
            try:
                with s.get(mainUrl + HA_API + INTERFACE_API, headers=headers, verify=False) as res:
                    res_dict = json.loads(res.text)
                    S_ifInfo = {}
                    for item in res_dict['result']:
                        S_ifInfo[item['name'] + '_ip_addr'] = item['ipv4address']
                        S_ifInfo[item['name'] + '_link_state'] = item['link_up']
            except Exception as e:
                write_log(f"{schNo}_{schName} / HA Device Connection Error : {e}")
                continue

        if_arr = ['eth1', 'eth2', 'eth3', 'eth4', 'eth8', 'eth9']
        for if_val in if_arr:
            if M_ifInfo[f'{if_val}_ip_addr']:
                for ip_val in M_ifInfo[f'{if_val}_ip_addr']:
                    proc_row += 1

                    proc_ws[f'A{proc_row}'].value = schNo
                    proc_ws[f'B{proc_row}'].value = schName
                    proc_ws[f'C{proc_row}'].value = "MASTER"
                    proc_ws[f'G{proc_row}'].value = if_val
                    proc_ws[f'H{proc_row}'].value = ip_val

            if S_ifInfo[f'{if_val}_ip_addr']:
                for ip_val in S_ifInfo[f'{if_val}_ip_addr']:
                    proc_row += 1
                    proc_ws[f'A{proc_row}'].value = schNo
                    proc_ws[f'B{proc_row}'].value = schName
                    proc_ws[f'C{proc_row}'].value = "SLAVE"
                    proc_ws[f'G{proc_row}'].value = if_val
                    proc_ws[f'H{proc_row}'].value = ip_val
        cnt += 1
        if cnt % 10 == 0:
            proc_wb.save(CUR_PATH + "\\ifCheck_result.xlsx")

    proc_wb.save(CUR_PATH + "\\ifCheck_result.xlsx")

