import requests
import base64
import openpyxl
import json
import os
import sys


from Cryptodome.Cipher import AES
from Cryptodome import Random

from datetime import date
import time
import urllib3

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

LOGIN_API = '/api/au/login'
LOGIN_DATA = {
        "lang": "ko",
        "force": "true",
        "readonly": 0,
        "login_id": "admin",
        "login_pw": "",
        "csrf_token": "",
        "type": "local"
    }

BACKUP_API = '/api/sm/backup/manual'
BACKUP_PARAM = {"ha_backup": 1, "target": "POVSL"}

proxy = {'https': 'http://127.0.0.1:8080'}

CUR_PATH = os.getcwd()
SRC_FILE = CUR_PATH + "\\backup.xlsx"
START_DATE = date.today()

LOG_NAME = "Bluemax_BackupLog"


def _pad(s):
    bs = 16
    return s + (bs - len(s) % bs) * chr(bs - len(s) % bs)


def write_log(string):
    dat = str(START_DATE).replace("-", "")
    f = open(CUR_PATH + f'\\{LOG_NAME}_{dat}.txt', "a+")
    now = time.localtime()
    cur_time = "%02d/%02d,%02d:%02d:%02d" % (now.tm_mon, now.tm_mday, now.tm_hour, now.tm_min, now.tm_sec)
    f.write(f'{cur_time};{string}\n')
    f.close()


def print_progress(iteration, total, prefix='', suffix='', decimals=1, barLength=100):
    formatStr = "{0:." + str(decimals) + "f}"
    percent = formatStr.format(100 * (iteration / float(total)))
    filledLength = int(round(barLength * iteration / float(total)))
    bar = '#' * filledLength + '-' * (barLength - filledLength)
    sys.stdout.write('\r%s |%s| %s%s %s' % (prefix, bar, percent, '%', suffix)),
    if iteration == total:
        sys.stdout.write('\n')
    sys.stdout.flush()


if __name__ == "__main__":
    wb = openpyxl.load_workbook(SRC_FILE)
    ws = wb.active

    for row in range(2, ws.max_row + 1):
        schNo = ws[f'A{row}'].value
        schOrg = ws[f'B{row}'].value
        schHaCls = ws[f'C{row}'].value
        schName = ws[f'D{row}'].value
        schUtmIp = ws[f'E{row}'].value
        schUtmAccessPort = ws[f'F{row}'].value

        schUtmId = ws[f'G{row}'].value
        schUtmPw = ws[f'H{row}'].value

        MAIN_URL = f"https://{schUtmIp}:{schUtmAccessPort}"

        if schHaCls == "HA":
            BACKUP_PARAM['ha_backup'] = 1
        else:
            BACKUP_PARAM['ha_backup'] = 0

        with requests.Session() as s:
            # WEB GUI Initialize
            with s.get(MAIN_URL, verify=False) as res:
                val = res.text
                find_num = val.find("csrf_token")
                csrf_token = val[find_num + 51:find_num + 115]
                iv = Random.new().read(AES.block_size)

                key = bytes(csrf_token[:32], 'utf-8')
                raw = bytes(_pad(schUtmPw), 'utf-8')
                cipher = AES.new(key, AES.MODE_CBC, iv)
                hex_data = (iv.hex() + cipher.encrypt(raw).hex()).encode('utf-8')
                login_pw = base64.b64encode(hex_data).decode('utf-8')

                LOGIN_DATA['login_id'] = schUtmId
                LOGIN_DATA['login_pw'] = login_pw
                LOGIN_DATA['csrf_token'] = csrf_token

            # Login Phase
            with s.post(MAIN_URL + LOGIN_API, json=LOGIN_DATA, verify=False) as res:
                cookies = res.cookies.get_dict()
                res_dict = json.loads(res.text)
                auth_key = res_dict['result']['api_token']
                headers = {'Authorization': auth_key}

            # Backup File Download
            with s.post(MAIN_URL + BACKUP_API, json=BACKUP_PARAM, headers=headers, verify=False) as res:
                write_log(f"{schNo}_{schName};{schUtmIp};{BACKUP_API};OK;")
                if not (os.path.isdir(CUR_PATH + f'\\{schOrg}')):
                    os.makedirs(CUR_PATH + f'\\{schOrg}')

                f = open(f"{CUR_PATH}\\{schOrg}\\{schNo}_{schOrg}_{schName}.tar", 'wb',)
                f.write(res.content)
                f.close()
            s.close()

