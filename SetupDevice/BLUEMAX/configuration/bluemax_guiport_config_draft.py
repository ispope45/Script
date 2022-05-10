# from SSHConnector import SSHConnector

import openpyxl
import os
import paramiko
import socket

from datetime import date
import time

CUR_PATH = os.getcwd()
SRC_FILE = CUR_PATH + "\\src.xlsx"
START_DATE = date.today()


class SSHConnector():
    def __init__(self):
        self.preamble = ['y\n', 'config\n']
        self.hostname = ['hostname\n', 'set hostname\n', 'exit\n']
        self.interface = ['interface eth1\n', 'no ip add\n', 'ip add 100.100.100.100/24\n', 'exit\n', 'y\n']

    def main(self, sw_ip, sw_user, sw_pass, sw_port, command_set):
        host = sw_ip
        username = sw_user
        password = sw_pass

        output = list()

        try:
            conn = paramiko.SSHClient()
            conn.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            conn.connect(host, username=username, password=password, port=sw_port, timeout=100)
            channel = conn.invoke_shell()
            time.sleep(1)
            for command in command_set:
                for line in command:
                    channel.send(line)

                    out_data, err_data = self.wait_streams(channel)
                    output.append(out_data)
                    print(out_data)
            return output

        except Exception as e:
            if "port" in str(e):
                return "Port Error"
            if "WinError 10060" in str(e):
                return "Connection Error"
            print(e)
            return "Error"

        finally:
            if conn is not None:
                conn.close()

    @staticmethod
    def wait_streams(channel):
        time.sleep(1)
        out_data = ""
        err_data = ""
        while True:
            time.sleep(2)
            if channel.recv_ready():
                out_data += channel.recv(1000).decode('ascii')
                if channel.recv_stderr_ready():
                    err_data += channel.recv(1000).decode('ascii')
                if out_data.find("#") != -1 or out_data.find("[y|n]") != -1:
                    break

        return out_data, err_data

    def ssh_connect(self, con_info, ip_cfg):
        self.main(con_info[0], con_info[1], con_info[2], con_info[3], ip_cfg)


def write_log(string):
    dat = str(START_DATE).replace("-", "")
    f = open(CUR_PATH + f'\\log_{dat}.txt', "a+")
    now = time.localtime()
    cur_time = "%02d/%02d,%02d:%02d:%02d" % (now.tm_mon, now.tm_mday, now.tm_hour, now.tm_min, now.tm_sec)
    f.write(f'{cur_time};{string}\n')
    f.close()


def port_check(ip, port):
    try:
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(2)
        result = sock.connect_ex((ip, port))
        sock.close()
    except socket.error as e:
        result = 1
        print("Error : " + e)

    if result == 0:
        return True
    else:
        return False


if __name__ == "__main__":
    # Excel Data Formatting
    SRC_FILE = CUR_PATH + "\\bluemax_configuration.xlsx"
    wb = openpyxl.load_workbook(SRC_FILE)
    ws = wb.active

    for row in range(2, ws.max_row + 1):
        schNo = ws[f'A{row}'].value
        schOrg = ws[f'B{row}'].value
        schHaCls = ws[f'C{row}'].value
        schName = ws[f'D{row}'].value
        schUtmIp = ws[f'E{row}'].value
        schUtmAccessPort = ws[f'F{row}'].value

        set_gui_port = ws[f'J{row}'].value

        ssh_access_id = ws[f'N{row}'].value
        ssh_access_pw = ws[f'O{row}'].value
        ssh_access_port = ws[f'P{row}'].value

        if port_check(schUtmIp, ssh_access_port):
            write_log(f"{schNo}_{schName};{schUtmIp}:{ssh_access_port} Port Check;OK;")
            # mainUrl = f"{baseUrl}:{schUtmAccessPort}"
        else:
            write_log(f"{schNo}_{schName};{schUtmIp}:{ssh_access_port} Port Check;False;")
            # Excel Comment
            # ws[f'N{row}'].value = f"{schNo}_{schName};{schUtmIp}:{ssh_access_port} Port Check;False;"
            # wb.save(SRC_FILE)
            continue

        con_info = [schUtmIp, ssh_access_id, ssh_access_pw, ssh_access_port]
        print(con_info)

        ifCfg_M = [['y\n',
                    'config\n',
                    'admin generic\n',
                    f'set gui_port {set_gui_port}\n',
                    'exit\n',
                    'y\n',
                    'y\n',
                    '\n']]

        ssh = SSHConnector()
        ssh.ssh_connect(con_info, ifCfg_M)

    # time.sleep(5)
    #
    # # INTERFACE & HOSTNAME Setting
    # ssh = SSHConnector()
    # th1 = Process(target=ssh.ssh_connect, args=(con_info[0], ifCfg_M))
    #
    # ssh2 = SSHConnector()
    # th2 = Process(target=ssh2.ssh_connect, args=(con_info[1], ifCfg_S))
    #
    # th1.start()
    # time.sleep(30)
    # th2.start()
    #
    # th1.join()
    # th2.join()
