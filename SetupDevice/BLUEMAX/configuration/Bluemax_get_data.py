import requests
import base64
import openpyxl
import json
import os, sys
import random, string
import socket

from requests_toolbelt import MultipartEncoder
from Cryptodome.Cipher import AES
from Cryptodome import Random

from datetime import date
import time
import urllib3
import random

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

LOGIN_API = '/api/au/login'

CUR_PATH = os.getcwd()
START_DATE = date.today()

proxy = {'https': 'http://127.0.0.1:8080'}

# ### INTERFACE GET API
INTERFACE_API = '/api/sm/interfaces'

# ### VIP GET API
VIPS_API = '/api/sm/ha/virtual-ips'

# ### DHCP AREA GET API
DHCP_AREA_API = '/api/sm/dhcp/server/dynamic-areas'


def make_login_data(login_pw, csrf_token):
    login_data = {
        "lang": "ko",
        "force": "true",
        "readonly": 0,
        "login_id": "admin",
        "login_pw": login_pw,
        "csrf_token": csrf_token,
        "type": "local"
    }
    return login_data


def _pad(s):
    bs = 16
    return s + (bs - len(s) % bs) * chr(bs - len(s) % bs)


def print_progress(iteration, total, prefix='', suffix='', decimals=1, barLength=100):
    formatStr = "{0:." + str(decimals) + "f}"
    percent = formatStr.format(100 * (iteration / float(total)))
    filledLength = int(round(barLength * iteration / float(total)))
    bar = '#' * filledLength + '-' * (barLength - filledLength)
    sys.stdout.write('\r%s |%s| %s%s %s' % (prefix, bar, percent, '%', suffix)),
    if iteration == total:
        sys.stdout.write('\n')
    sys.stdout.flush()


def write_log(string):
    dat = str(START_DATE).replace("-", "")
    f = open(CUR_PATH + f'\\log_{dat}.txt', "a+")
    now = time.localtime()
    cur_time = "%02d/%02d,%02d:%02d:%02d" % (now.tm_mon, now.tm_mday, now.tm_hour, now.tm_min, now.tm_sec)
    f.write(f'{cur_time};{string}\n')
    f.close()


def port_check(ip, port):
    try:
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(2)
        result = sock.connect_ex((ip, port))
        sock.close()
    except socket.error as e:
        result = 1
        print("Error : " + e)

    if result == 0:
        return True
    else:
        return False


if __name__ == "__main__":

    SRC_FILE = CUR_PATH + "\\bluemax_get_data.xlsx"
    wb = openpyxl.load_workbook(SRC_FILE)
    ws = wb.active

    interface_check_wb = openpyxl.Workbook()
    interface_check_ws = interface_check_wb.active
    i_c_row = 1

    vips_check_wb = openpyxl.Workbook()
    vips_check_ws = vips_check_wb.active
    v_c_row = 1

    dhcp_check_wb = openpyxl.Workbook()
    dhcp_check_ws = dhcp_check_wb.active
    d_c_row = 1

    for row in range(2, ws.max_row + 1):
        schNo = ws[f'A{row}'].value
        schOrg = ws[f'B{row}'].value
        schHaCls = ws[f'C{row}'].value
        schName = ws[f'D{row}'].value
        schUtmIp = ws[f'E{row}'].value
        schUtmAccessPort = ws[f'F{row}'].value

        interface_check_val = ws[f'G{row}'].value
        vips_check_val = ws[f'H{row}'].value
        dhcp_check_val = ws[f'I{row}'].value

        baseUrl = f"https://{schUtmIp}"
        ha_api = f'/ha:{schUtmAccessPort}:ha_grp_2'

        haChk = False
        if schHaCls == "HA":
            haChk = True

        if port_check(schUtmIp, schUtmAccessPort):
            write_log(f"{schNo}_{schName};{schUtmIp}:{schUtmAccessPort} Port Check;OK;")
            mainUrl = f"{baseUrl}:{schUtmAccessPort}"
        else:
            write_log(f"{schNo}_{schName};{schUtmIp}:{schUtmAccessPort} Port Check;False;")
            continue

        with requests.Session() as s:
            # WEB GUI Initialize
            with s.get(mainUrl, verify=False) as res:
                val = res.text
                find_num = val.find("csrf_token")
                csrf_token = val[find_num + 51:find_num + 115]
                pw = "secui00@!"
                iv = Random.new().read(AES.block_size)

                key = bytes(csrf_token[:32], 'utf-8')
                raw = bytes(_pad(pw), 'utf-8')
                cipher = AES.new(key, AES.MODE_CBC, iv)
                hex_data = (iv.hex() + cipher.encrypt(raw).hex()).encode('utf-8')
                login_pw = base64.b64encode(hex_data).decode('utf-8')

                login_data = make_login_data(login_pw, csrf_token)

            # Login Phase
            with s.post(mainUrl + LOGIN_API, json=login_data, verify=False) as res:
                cookies = res.cookies.get_dict()
                res_dict = json.loads(res.text)
                auth_key = res_dict['result']['api_token']
                headers = {'Authorization': auth_key}

            # INTERFACE GET
            if str(interface_check_val) != "None":
                with s.get(mainUrl + INTERFACE_API, headers=headers, verify=False) as res:
                    res_dict = json.loads(res.text)
                    if res_dict['code'] == "ok":
                        write_log(f"{schNo}_{schName};{schUtmIp};{INTERFACE_API};OK;")
                        if_list = res_dict['result']
                        for if_val in if_list:
                            if if_val['ipv4address']:
                                for ip in if_val['ipv4address']:
                                    interface_check_ws[f'A{i_c_row}'].value = schNo
                                    interface_check_ws[f'B{i_c_row}'].value = schName
                                    interface_check_ws[f'C{i_c_row}'].value = if_val['name']
                                    interface_check_ws[f'D{i_c_row}'].value = ip
                                    i_c_row += 1
                    else:
                        write_log(
                            f"{schNo}_{schName};{schUtmIp};{INTERFACE_API};{res_dict['dev_t']};{res_dict['message']}")
                if haChk:
                    with s.get(mainUrl + ha_api + INTERFACE_API, headers=headers, verify=False) as res:
                        res_dict = json.loads(res.text)
                        if res_dict['code'] == "ok":
                            write_log(f"{schNo}_{schName};{schUtmIp};{ha_api}{INTERFACE_API};OK;")
                            if_list = res_dict['result']
                            for if_val in if_list:
                                if if_val['ipv4address']:
                                    for ip in if_val['ipv4address']:
                                        interface_check_ws[f'A{i_c_row}'].value = schNo
                                        interface_check_ws[f'B{i_c_row}'].value = schName
                                        interface_check_ws[f'C{i_c_row}'].value = if_val['name']
                                        interface_check_ws[f'D{i_c_row}'].value = ip
                                        i_c_row += 1
                        else:
                            write_log(
                                f"{schNo}_{schName};{schUtmIp};{ha_api}{INTERFACE_API};{res_dict['dev_t']};{res_dict['message']}")

            if str(vips_check_val) != "None":
                with s.get(mainUrl + VIPS_API, headers=headers, verify=False) as res:
                    res_dict = json.loads(res.text)
                    if res_dict['code'] == "ok":
                        write_log(f"{schNo}_{schName};{schUtmIp};{VIPS_API};OK;")
                        vip_list = res_dict['result']
                        for vip_val in vip_list:
                            for val in vip_val['children']:
                                vips_check_ws[f'A{v_c_row}'].value = schNo
                                vips_check_ws[f'B{v_c_row}'].value = schName
                                vips_check_ws[f'C{v_c_row}'].value = val['ifc_name']
                                vips_check_ws[f'D{v_c_row}'].value = val['mmbr_uuid']
                                vips_check_ws[f'E{v_c_row}'].value = val['ip']
                                vips_check_ws[f'F{v_c_row}'].value = val['netmask']
                                v_c_row += 1
                    else:
                        write_log(
                            f"{schNo}_{schName};{schUtmIp};{VIPS_API};{res_dict['dev_t']};{res_dict['message']}")

            if str(dhcp_check_val) != "None":
                with s.get(mainUrl + DHCP_AREA_API, headers=headers, verify=False) as res:
                    res_dict = json.loads(res.text)
                    if res_dict['code'] == "ok":
                        write_log(f"{schNo}_{schName};{schUtmIp};{DHCP_AREA_API};OK;")
                        dhcp_list = res_dict['result']
                        # print(dhcp_list)
                        for dhcp_val in dhcp_list:
                            dhcp_check_ws[f'A{d_c_row}'].value = schNo
                            dhcp_check_ws[f'B{d_c_row}'].value = schName
                            dhcp_check_ws[f'C{d_c_row}'].value = dhcp_val['net_addr']
                            dhcp_check_ws[f'D{d_c_row}'].value = dhcp_val['subnet_mask']
                            d_c_row += 1

                    else:
                        write_log(
                            f"{schNo}_{schName};{schUtmIp};{DHCP_AREA_API};{res_dict['dev_t']};{res_dict['message']}")
                if haChk:
                    with s.get(mainUrl + ha_api + DHCP_AREA_API, headers=headers, verify=False) as res:
                        res_dict = json.loads(res.text)
                        if res_dict['code'] == "ok":
                            write_log(f"{schNo}_{schName};{schUtmIp};{ha_api}{DHCP_AREA_API};OK;")
                            dhcp_list = res_dict['result']
                            # print(dhcp_list)
                            for dhcp_val in dhcp_list:
                                dhcp_check_ws[f'A{d_c_row}'].value = schNo
                                dhcp_check_ws[f'B{d_c_row}'].value = schName
                                dhcp_check_ws[f'C{d_c_row}'].value = dhcp_val['net_addr']
                                dhcp_check_ws[f'D{d_c_row}'].value = dhcp_val['subnet_mask']
                                d_c_row += 1
                        else:
                            write_log(
                                f"{schNo}_{schName};{schUtmIp};{ha_api}{DHCP_AREA_API};{res_dict['dev_t']};{res_dict['message']}")

        if str(interface_check_val) != "None":
            interface_check_wb.save(CUR_PATH + "\\interface_check_result.xlsx")
        if str(vips_check_val) != "None":
            vips_check_wb.save(CUR_PATH + "\\vips_check_result.xlsx")
        if str(dhcp_check_val) != "None":
            dhcp_check_wb.save(CUR_PATH + "\\dhcp_check_result.xlsx")

        print_progress(row - 1, ws.max_row - 1, 'Progress:', 'Complete ', 1, 50)


