import requests
import base64
import openpyxl
import json
import os, sys
import socket
import hashlib

from Cryptodome.Cipher import AES
from Cryptodome import Random

from datetime import date
import time
import urllib3

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

CUR_PATH = os.getcwd()
START_DATE = date.today()

proxy = {'https': 'http://127.0.0.1:8080'}

# ### LOGIN API
LOGIN_API = '/api/au/login'
LOGIN_json = {
        "lang": "ko",
        "force": "true",
        "readonly": 0,
        "login_id": "",
        "login_pw": "",
        "csrf_token": "",
        "type": "local"
    }

# ### LOGOUT API
LOGOUT_API = '/api/au/logout'

# ### BACKUP API
BACKUP_API = '/api/sm/backup/manual'
BACKUP_json = {"ha_backup": 1, "target": "POVSL"}


def _pad(s):
    bs = 16
    return s + (bs - len(s) % bs) * chr(bs - len(s) % bs)


def make_key():
    now = int(time.time())
    print(now)

    key = str(now).encode('utf-8')
    print(key)
    return hashlib.sha256(key).hexdigest()


def write_log(string):
    dat = str(START_DATE).replace("-", "")
    f = open(CUR_PATH + f'\\log_{dat}.txt', "a+")
    now = time.localtime()
    cur_time = "%02d/%02d,%02d:%02d:%02d" % (now.tm_mon, now.tm_mday, now.tm_hour, now.tm_min, now.tm_sec)
    f.write(f'{cur_time};{string}\n')
    f.close()


def port_check(ip, port):
    try:
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(2)
        result = sock.connect_ex((ip, port))
        sock.close()
    except socket.error as e:
        result = 1
        print("Error : " + e)

    if result == 0:
        return True
    else:
        return False


def print_progress(iteration, total, prefix='', suffix='', decimals=1, barLength=100):
    formatStr = "{0:." + str(decimals) + "f}"
    percent = formatStr.format(100 * (iteration / float(total)))
    filledLength = int(round(barLength * iteration / float(total)))
    bar = '#' * filledLength + '-' * (barLength - filledLength)
    sys.stdout.write('\r%s |%s| %s%s %s' % (prefix, bar, percent, '%', suffix)),
    if iteration == total:
        sys.stdout.write('\n')
    sys.stdout.flush()


if __name__ == "__main__":
    SRC_FILE = CUR_PATH + "\\bluemax_backup.xlsx"
    wb = openpyxl.load_workbook(SRC_FILE)
    ws = wb.active

    for row in range(2, ws.max_row + 1):
        schNo = ws[f'A{row}'].value
        schOrg = ws[f'B{row}'].value
        schHaCls = ws[f'C{row}'].value
        schName = ws[f'D{row}'].value
        schUtmIp = ws[f'E{row}'].value
        schUtmAccessPort = ws[f'F{row}'].value

        loginId = ws[f'G{row}'].value
        loginPass = ws[f'H{row}'].value

        baseUrl = f"https://{schUtmIp}"
        ha_api = f'/ha:{schUtmAccessPort}:ha_grp_2'

        haChk = False
        if schHaCls == "HA":
            haChk = True

        if haChk:
            BACKUP_json['ha_backup'] = 1
        else:
            BACKUP_json['ha_backup'] = 0

        if str(schUtmIp) == "None" or str(schUtmAccessPort) == "None":
            write_log(f"{schNo}_{schName};No IP Addr or No Access Port;False;")
            continue

        if port_check(schUtmIp, schUtmAccessPort):
            write_log(f"{schNo}_{schName};{schUtmIp}:{schUtmAccessPort} Port Check;OK;")
            mainUrl = f"{baseUrl}:{schUtmAccessPort}"
        else:
            write_log(f"{schNo}_{schName};{schUtmIp}:{schUtmAccessPort} Port Check;False;")
            continue

        with requests.Session() as s:
            # WEB GUI Initialize
            with s.get(mainUrl, verify=False) as res:
                val = res.text
                find_num = val.find("csrf_token")
                csrf_token = val[find_num + 51:find_num + 115]

                iv = Random.new().read(AES.block_size)
                key = bytes(csrf_token[:32], 'utf-8')
                raw = bytes(_pad(loginPass), 'utf-8')
                cipher = AES.new(key, AES.MODE_CBC, iv)
                hex_data = (iv.hex() + cipher.encrypt(raw).hex()).encode('utf-8')
                enc_pw = base64.b64encode(hex_data).decode('utf-8')

                LOGIN_json['login_id'] = loginId
                LOGIN_json['login_pw'] = enc_pw
                LOGIN_json['csrf_token'] = csrf_token

            # Login Phase
            with s.post(mainUrl + LOGIN_API, json=LOGIN_json, verify=False) as res:
                cookies = res.cookies.get_dict()
                res_dict = json.loads(res.text)
                if res_dict['code'] == "ok":
                    write_log(f"{schNo}_{schName};{schUtmIp};{LOGIN_API};OK;")
                    auth_key = res_dict['result']['api_token']
                    secui_helper_key = res_dict['result']['secui_helper_key']
                    headers = {'Authorization': auth_key}
                else:
                    write_log(f"{schNo}_{schName};{schUtmIp};{LOGIN_API};{res_dict['dev_t']};{res_dict['message']}")
                    continue

            # Backup File Download
            with s.post(mainUrl + BACKUP_API, json=BACKUP_json, headers=headers, verify=False) as res:
                write_log(f'{schNo}_{schOrg}_{schName};{res.url};{res.headers}')
                if not (os.path.isdir(CUR_PATH + f'\\{schOrg}')):
                    os.makedirs(CUR_PATH + f'\\{schOrg}')

                f = open(f"{CUR_PATH}\\{schOrg}\\{schNo}_{schOrg}_{schName}.tar", 'wb', )
                f.write(res.content)
                f.close()

        print_progress(row - 1, ws.max_row - 1, 'Progress:', 'Complete ', 1, 50)