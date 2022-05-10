import requests
import base64
import openpyxl
import json
import os, sys
import random, string
import socket
import hashlib

from requests_toolbelt import MultipartEncoder
from Cryptodome.Cipher import AES
from Cryptodome import Random

from datetime import date
import time
import urllib3
import random

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

CUR_PATH = os.getcwd()
START_DATE = date.today()

proxy = {'https': 'http://127.0.0.1:8080'}

# ### LOGIN API
LOGIN_API = '/api/au/login'

# ### LOGOUT API
LOGOUT_API = '/api/au/logout'

# ### INTERFACE API
INTERFACE_API = '/api/sm/interfaces'

# ### POLICY BATCH API
POLICY_API = '/api/po/fw/4/rules?key='
POLICY_IMPORT_API = '/api/co/file/import'

POLICY_BATCH_APPLY_API = '/api/po/fw/4/rules/batch'
pol_import_batch = {"file": "2ffbe3f5-1db5-48a9-a1b8-bf8ebf1b8644.xlsx",
                    "excel_pos": 1,
                    "allow_dup": 0,
                    "pre_rule_id": -1,
                    "top_rule_id_for_group": {"default": 5}}

POLICY_APPLY_API = '/api/po/command/fw-4-policies/apply'
pol_apply_json = {'mod_rule': 1}

# ### SNMP API CONFIG
SNMP_SETTING_API = '/api/sm/server/snmp/users'
snmp_setting_json = {"svr_name": "NMS",
                     "usr_name": "schoolnet4",
                     "security": 2,
                     "auth_type": 0,
                     "msg_enc_type": 0,
                     "auth_pw": "U2VuMTY3MDEzOTZA",
                     "msg_enc_pw": "U2VuMTY3MDEzOTZA"}

SNMP_V3_API = '/api/sm/server/snmp/v3'
snmp_v3_json = {"use": 1,
                "snmp_usr_id": 1}

SNMP_ENABLE_API = '/api/sm/server/snmp'
snmp_enable_json = {"use": 1,
                    "community": ""}

SNMP_APPLY_API = '/api/sm/command/server-snmp/apply'

# ### SYSLOG API
SYSLOG_SETTING_API = '/api/sm/server/syslog'
syslog_setting_json = {"svr1_use": 1,
                       "svr1_addr": "192.168.10.254",
                       "svr1_port": 7514,
                       "svr1_func": 0,
                       "svr1_form": 4,
                       "svr1_desc": "",
                       "svr2_use": 0,
                       "svr2_addr": "",
                       "svr2_port": 514,
                       "svr2_func": 0,
                       "svr2_form": 0,
                       "svr2_desc": "",
                       "svr3_use": 0,
                       "svr3_addr": "",
                       "svr3_port": 514,
                       "svr3_func": 0,
                       "svr3_form": 0,
                       "svr3_desc": "",
                       "svr4_use": 0,
                       "svr4_addr": "",
                       "svr4_port": 514,
                       "svr4_func": 0,
                       "svr4_form": 0,
                       "svr4_desc": "",
                       "svr5_use": 0,
                       "svr5_addr": "",
                       "svr5_port": 514,
                       "svr5_func": 0,
                       "svr5_form": 0,
                       "svr5_desc": ""}

SYSLOG_APPLY_API = '/api/sm/command/server-syslog/apply'

LOG_SETTING_API = '/api/lr/monitor/setup/logs'
log_setting_json = {"items": [
    {"log_id": 130001, "mod_id": 1, "name": "traffic_session", "local_use": 1, "syslog1_use": 1, "syslog2_use": 0, "syslog3_use": 0, "syslog4_use": 0, "syslog5_use": 0, "snmptrap1_use": 0, "snmptrap2_use": 0, "snmptrap3_use": 0, "auto_backup_use": 0, "smtp_use": 0},
    {"log_id": 100, "mod_id": 3, "name": "fw_traffic", "local_use": 1, "syslog1_use": 1, "syslog2_use": 0, "syslog3_use": 0, "syslog4_use": 0, "syslog5_use": 0, "snmptrap1_use": 0, "snmptrap2_use": 0, "snmptrap3_use": 0, "auto_backup_use": 0, "smtp_use": 0},
    {"log_id": 101, "mod_id": 3, "name": "fw_rule_traffic", "local_use": 1, "syslog1_use": 1, "syslog2_use": 0, "syslog3_use": 0, "syslog4_use": 0, "syslog5_use": 0, "snmptrap1_use": 0, "snmptrap2_use": 0, "snmptrap3_use": 0, "auto_backup_use": 0, "smtp_use": 0},
    {"log_id": 5, "mod_id": 3, "name": "nat_session", "local_use": 1, "syslog1_use": 1, "syslog2_use": 0, "syslog3_use": 0, "syslog4_use": 0, "syslog5_use": 0, "snmptrap1_use": 0, "snmptrap2_use": 0, "snmptrap3_use": 0, "auto_backup_use": 0, "smtp_use": 0},
    {"log_id": 104, "mod_id": 3, "name": "nat_traffic", "local_use": 1, "syslog1_use": 1, "syslog2_use": 0, "syslog3_use": 0, "syslog4_use": 0, "syslog5_use": 0, "snmptrap1_use": 0, "snmptrap2_use": 0, "snmptrap3_use": 0, "auto_backup_use": 0, "smtp_use": 0},
    {"log_id": 8, "mod_id": 3, "name": "ad_group_mapping", "local_use": 1, "syslog1_use": 1, "syslog2_use": 0, "syslog3_use": 0, "syslog4_use": 0, "syslog5_use": 0, "snmptrap1_use": 0, "snmptrap2_use": 0, "snmptrap3_use": 0, "auto_backup_use": 0, "smtp_use": 0},
    {"log_id": 6, "mod_id": 3, "name": "userauth", "local_use": 1, "syslog1_use": 1, "syslog2_use": 0, "syslog3_use": 0, "syslog4_use": 0, "syslog5_use": 0, "snmptrap1_use": 0, "snmptrap2_use": 0, "snmptrap3_use": 0, "auto_backup_use": 0, "smtp_use": 0},
    {"log_id": 7, "mod_id": 3, "name": "fqdn_object_management", "local_use": 1, "syslog1_use": 1, "syslog2_use": 0, "syslog3_use": 0, "syslog4_use": 0, "syslog5_use": 0, "snmptrap1_use": 0, "snmptrap2_use": 0, "snmptrap3_use": 0, "auto_backup_use": 0, "smtp_use": 0},
    {"log_id": 1001, "mod_id": 9, "name": "ha_event", "local_use": 1, "syslog1_use": 1, "syslog2_use": 0, "syslog3_use": 0, "syslog4_use": 0, "syslog5_use": 0, "snmptrap1_use": 0, "snmptrap2_use": 0, "snmptrap3_use": 0, "auto_backup_use": 0, "smtp_use": 0},
    {"log_id": 1100, "mod_id": 9, "name": "ha_traffic", "local_use": 1, "syslog1_use": 1, "syslog2_use": 0, "syslog3_use": 0, "syslog4_use": 0, "syslog5_use": 0, "snmptrap1_use": 0, "snmptrap2_use": 0, "snmptrap3_use": 0, "auto_backup_use": 0, "smtp_use": 0},
    {"log_id": 1101, "mod_id": 9, "name": "ha_status", "local_use": 1, "syslog1_use": 1, "syslog2_use": 0, "syslog3_use": 0, "syslog4_use": 0, "syslog5_use": 0, "snmptrap1_use": 0, "snmptrap2_use": 0, "snmptrap3_use": 0, "auto_backup_use": 0, "smtp_use": 0},
    {"log_id": 5003, "mod_id": 9, "name": "iap_interworking", "local_use": 1, "syslog1_use": 1, "syslog2_use": 0, "syslog3_use": 0, "syslog4_use": 0, "syslog5_use": 0, "snmptrap1_use": 0, "snmptrap2_use": 0, "snmptrap3_use": 0, "auto_backup_use": 0, "smtp_use": 0},
    {"log_id": 5007, "mod_id": 9, "name": "zombie_pc_solution", "local_use": 1, "syslog1_use": 1, "syslog2_use": 0, "syslog3_use": 0, "syslog4_use": 0, "syslog5_use": 0, "snmptrap1_use": 0, "snmptrap2_use": 0, "snmptrap3_use": 0, "auto_backup_use": 0, "smtp_use": 0},
    {"log_id": 5004, "mod_id": 9, "name": "blacklist_regist", "local_use": 1, "syslog1_use": 1, "syslog2_use": 0, "syslog3_use": 0, "syslog4_use": 0, "syslog5_use": 0, "snmptrap1_use": 0, "snmptrap2_use": 0, "snmptrap3_use": 0, "auto_backup_use": 0, "smtp_use": 0},
    {"log_id": 5006, "mod_id": 9, "name": "line_management", "local_use": 1, "syslog1_use": 1, "syslog2_use": 0, "syslog3_use": 0, "syslog4_use": 0, "syslog5_use": 0, "snmptrap1_use": 0, "snmptrap2_use": 0, "snmptrap3_use": 0, "auto_backup_use": 0, "smtp_use": 0},
    {"log_id": 5100, "mod_id": 9, "name": "system_resource", "local_use": 1, "syslog1_use": 1, "syslog2_use": 0, "syslog3_use": 0, "syslog4_use": 0, "syslog5_use": 0, "snmptrap1_use": 0, "snmptrap2_use": 0, "snmptrap3_use": 0, "auto_backup_use": 0, "smtp_use": 0},
    {"log_id": 5101, "mod_id": 9, "name": "daemon", "local_use": 1, "syslog1_use": 1, "syslog2_use": 0, "syslog3_use": 0, "syslog4_use": 0, "syslog5_use": 0, "snmptrap1_use": 0, "snmptrap2_use": 0, "snmptrap3_use": 0, "auto_backup_use": 0, "smtp_use": 0},
    {"log_id": 5102, "mod_id": 9, "name": "interface_traffic", "local_use": 1, "syslog1_use": 1, "syslog2_use": 0, "syslog3_use": 0, "syslog4_use": 0, "syslog5_use": 0, "snmptrap1_use": 0, "snmptrap2_use": 0, "snmptrap3_use": 0, "auto_backup_use": 0, "smtp_use": 0},
    {"log_id": 3103, "mod_id": 9, "name": "oversubscription", "local_use": 1, "syslog1_use": 1, "syslog2_use": 0, "syslog3_use": 0, "syslog4_use": 0, "syslog5_use": 0, "snmptrap1_use": 0, "snmptrap2_use": 0, "snmptrap3_use": 0, "auto_backup_use": 0, "smtp_use": 0},
    {"log_id": 106, "mod_id": 9, "name": "qos", "local_use": 1, "syslog1_use": 1, "syslog2_use": 0, "syslog3_use": 0, "syslog4_use": 0, "syslog5_use": 0, "snmptrap1_use": 0, "snmptrap2_use": 0, "snmptrap3_use": 0, "auto_backup_use": 0, "smtp_use": 0},
    {"log_id": 5107, "mod_id": 9, "name": "vlan_traffic", "local_use": 1, "syslog1_use": 1, "syslog2_use": 0, "syslog3_use": 0, "syslog4_use": 0, "syslog5_use": 0, "snmptrap1_use": 0, "snmptrap2_use": 0, "snmptrap3_use": 0, "auto_backup_use": 0, "smtp_use": 0},
    {"log_id": 10001, "mod_id": 10, "name": "alert", "local_use": 1, "syslog1_use": 1, "syslog2_use": 0, "syslog3_use": 0, "syslog4_use": 0, "syslog5_use": 0, "snmptrap1_use": 0, "snmptrap2_use": 0, "snmptrap3_use": 0, "auto_backup_use": 0, "smtp_use": 0},
    {"log_id": 5001, "mod_id": 11, "name": "audit", "local_use": 1, "syslog1_use": 1, "syslog2_use": 0, "syslog3_use": 0, "syslog4_use": 0, "syslog5_use": 0, "snmptrap1_use": 0, "snmptrap2_use": 0, "snmptrap3_use": 0, "auto_backup_use": 0, "smtp_use": 0}]}

LOG_SETTING_APPLY_API = '/api/lr/command/setup-logs/apply'

# ### ADMIN IP ACL
ADMIN_IP_ACL_API = '/api/sm/manager/sys-ips'
admin_ip_acl_json = {"ip_ver": 1,
                     "ip": "1.1.1.3",
                     "prefix": 32,
                     "desc": ""}

ADMIN_IP_ACL_APPLY_API = '/api/sm/manager/sys-ips/apply'

# ### Manager Access port
MANAGER_ACCESS_PORT_API = '/api/sm/manager/config'
manager_access_port_json = {"gui_idle_exit": 0,
                            "gui_idle_use": 0,
                            "gui_idle_tm": 30,
                            "login_fail_cnt": 3,
                            "cli_port": 22,
                            "gui_port": 443,
                            "ext_auth_use": 0,
                            "auth_prof": None}

MANAGER_ACCESS_PORT_APPLY_API = '/api/sm/manager/config/apply'

# ### DHCP CONFIG Lease time
DHCP_CONFIG_API = '/api/sm/dhcp/server/config'
dhcp_config_json = {"normal_serv_use_enable": 1,
                    "normal_serv_rent_tm": 480,
                    "normal_serv_auth": 0,
                    "dns_serv1": "168.126.63.1",
                    "dns_serv2": "8.8.8.8",
                    "wins_serv1": "",
                    "wins_serv2": ""}

DHCP_CONFIG_APPLY_API = '/api/sm/dhcp/server/apply'

# ### MANAGER ACCOUNT CONFIG
MANAGER_ACCOUNT_CONFIG_API = '/api/sm/manager/users'
manager_account_json1 = {"type": 0,
                         "login_id": "senadm01",
                         "login_pw": "ZDMxZmNmYWQwNGQ3MTA5NTU5YTgzY2NmYTBmM2Y5NzY5YjQyMTgyYjI1NzJmODRjNDYxOGJiM2FhNzIwNWUwMA==",
                         "name": "SEN_ADMIN",
                         "pw_expire_use": 0,
                         "pw_expire_prd": 60,
                         "level": 3,
                         "depart": "",
                         "phone": "",
                         "email": "",
                         "email_alarm_use": 0,
                         "id_expire_use": 0,
                         "id_expire_prd": 30,
                         "ext_clnt_use": 0,
                         "otp_use": 0,
                         "otp_key": "",
                         "adm_prf": None,
                         "adm_ip_use": 0,
                         "adm_ip": []}
manager_account_json2 = {"type": 0,
                         "login_id": "tamsadm01",
                         "login_pw": "Yjk0MmEzM2VmOWJiZDBlYzdjYmMzOGRlNjYyYjg2MGE0YzMwZjZiYjhjMmI1MzYzYTBmNWRjMDVjOWJjMjQ0OQ==",
                         "name": "TAMS_ADMIN",
                         "pw_expire_use": 0,
                         "pw_expire_prd": 60,
                         "level": 3,
                         "depart": "",
                         "phone": "",
                         "email": "",
                         "email_alarm_use": 0,
                         "id_expire_use": 0,
                         "id_expire_prd": 30,
                         "ext_clnt_use": 0,
                         "otp_use": 0,
                         "otp_key": "",
                         "adm_prf": None,
                         "adm_ip_use": 0,
                         "adm_ip": []}

MANAGER_ACCOUNT_CONFIG_APPLY_API = '/api/sm/manager/users/apply'


def make_login_data(login_pw, csrf_token):
    login_data = {
        "lang": "ko",
        "force": "true",
        "readonly": 0,
        "login_id": "admin",
        "login_pw": login_pw,
        "csrf_token": csrf_token,
        "type": "local"
    }
    return login_data


def _pad(s):
    bs = 16
    return s + (bs - len(s) % bs) * chr(bs - len(s) % bs)


def make_key():
    now = int(time.time())
    print(now)

    key = str(now).encode('utf-8')
    print(key)
    return hashlib.sha256(key).hexdigest()


def write_log(string):
    dat = str(START_DATE).replace("-", "")
    f = open(CUR_PATH + f'\\log_{dat}.txt', "a+")
    now = time.localtime()
    cur_time = "%02d/%02d,%02d:%02d:%02d" % (now.tm_mon, now.tm_mday, now.tm_hour, now.tm_min, now.tm_sec)
    f.write(f'{cur_time};{string}\n')
    f.close()


def port_check(ip, port):
    try:
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(2)
        result = sock.connect_ex((ip, port))
        sock.close()
    except socket.error as e:
        result = 1
        print("Error : " + e)

    if result == 0:
        return True
    else:
        return False


def print_progress(iteration, total, prefix='', suffix='', decimals=1, barLength=100):
    formatStr = "{0:." + str(decimals) + "f}"
    percent = formatStr.format(100 * (iteration / float(total)))
    filledLength = int(round(barLength * iteration / float(total)))
    bar = '#' * filledLength + '-' * (barLength - filledLength)
    sys.stdout.write('\r%s |%s| %s%s %s' % (prefix, bar, percent, '%', suffix)),
    if iteration == total:
        sys.stdout.write('\n')
    sys.stdout.flush()


if __name__ == "__main__":
    SRC_FILE = CUR_PATH + "\\bluemax_configuration.xlsx"
    wb = openpyxl.load_workbook(SRC_FILE)
    ws = wb.active

    for row in range(2, ws.max_row + 1):
        schNo = ws[f'A{row}'].value
        schOrg = ws[f'B{row}'].value
        schHaCls = ws[f'C{row}'].value
        schName = ws[f'D{row}'].value
        schUtmIp = ws[f'E{row}'].value
        schUtmAccessPort = ws[f'F{row}'].value

        admin_ip_acl_ip = ws[f'G{row}'].value
        syslog_setting_ip = ws[f'H{row}'].value
        syslog_setting_port = ws[f'I{row}'].value
        set_gui_port = ws[f'J{row}'].value
        set_cli_port = ws[f'K{row}'].value
        policy_import_filename = ws[f'L{row}'].value
        snmp_setting = ws[f'M{row}'].value
        dhcp_lease_time = ws[f'Q{row}'].value
        manager_account_setting = ws[f'R{row}'].value

        baseUrl = f"https://{schUtmIp}"
        ha_api = f'/ha:{schUtmAccessPort}:ha_grp_2'
        if str(policy_import_filename) != "None":
            upload_file = f"./{policy_import_filename}.xlsx"
            upload_data = open(upload_file, 'rb')
            files = {"file_name": upload_data}

        haChk = False
        if schHaCls == "HA":
            haChk = True

        if str(schUtmIp) == "None" or str(schUtmAccessPort) == "None":
            write_log(f"{schNo}_{schName};No IP Addr or No Access Port;False;")
            continue

        if port_check(schUtmIp, schUtmAccessPort):
            write_log(f"{schNo}_{schName};{schUtmIp}:{schUtmAccessPort} Port Check;OK;")
            mainUrl = f"{baseUrl}:{schUtmAccessPort}"
        else:
            write_log(f"{schNo}_{schName};{schUtmIp}:{schUtmAccessPort} Port Check;False;")
            continue

        with requests.Session() as s:
            # WEB GUI Initialize
            with s.get(mainUrl, verify=False) as res:
                val = res.text
                find_num = val.find("csrf_token")
                csrf_token = val[find_num + 51:find_num + 115]
                pw = "secui00@!"
                iv = Random.new().read(AES.block_size)

                key = bytes(csrf_token[:32], 'utf-8')
                raw = bytes(_pad(pw), 'utf-8')
                cipher = AES.new(key, AES.MODE_CBC, iv)
                hex_data = (iv.hex() + cipher.encrypt(raw).hex()).encode('utf-8')
                login_pw = base64.b64encode(hex_data).decode('utf-8')

                login_data = make_login_data(login_pw, csrf_token)

            # Login Phase
            with s.post(mainUrl + LOGIN_API, json=login_data, verify=False) as res:
                cookies = res.cookies.get_dict()
                res_dict = json.loads(res.text)
                auth_key = res_dict['result']['api_token']
                secui_helper_key = res_dict['result']['secui_helper_key']
                headers = {'Authorization': auth_key}

            # POLICY IMPORT
            if str(policy_import_filename) != "None":
                with s.get(mainUrl + POLICY_API + "fwKey_4_admin_" + str(random.randint(32000, 39999)), headers=headers, verify=False) as res:
                    res_dict = json.loads(res.text)
                    rule_cnt = res_dict['result'][0]['rule_cnt']
                    top_rule_id = res_dict['result'][1]['fw_rule_id']
                    if res_dict['code'] == "ok":
                        write_log(f"{schNo}_{schName};{schUtmIp};{POLICY_API};OK;")
                    else:
                        write_log(f"{schNo}_{schName};{schUtmIp};{POLICY_API};{res_dict['dev_t']};{res_dict['message']}")

                fields = {
                    'file': (upload_file, upload_data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                }
                boundary = '----WebKitFormBoundary' \
                           + ''.join(random.sample(string.ascii_letters + string.digits, 16))
                m = MultipartEncoder(fields=fields, boundary=boundary)
                pol_import_headers = {'Authorization': auth_key,
                                      'Content-Type': m.content_type}

                with s.post(mainUrl + POLICY_IMPORT_API, data=m, verify=False, headers=pol_import_headers) as res:
                    res_dict = json.loads(res.text)
                    file_name = res_dict['result']['file']
                    pol_import_batch['file'] = file_name
                    pol_import_batch['top_rule_id_for_group']['default'] = top_rule_id
                    if res_dict['code'] == "ok":
                        write_log(f"{schNo}_{schName};{schUtmIp};{POLICY_IMPORT_API};OK;")
                    else:
                        write_log(f"{schNo}_{schName};{schUtmIp};{POLICY_IMPORT_API};{res_dict['dev_t']};{res_dict['message']}")

                with s.post(mainUrl + POLICY_BATCH_APPLY_API, json=pol_import_batch, verify=False, headers=headers) as res:
                    res_dict = json.loads(res.text)
                    if res_dict['code'] == "async":
                        write_log(f"{schNo}_{schName};{schUtmIp};{POLICY_BATCH_APPLY_API};OK;")
                    else:
                        write_log(f"{schNo}_{schName};{schUtmIp};{POLICY_BATCH_APPLY_API};{res_dict['dev_t']};{res_dict['message']}")

                with s.put(mainUrl + POLICY_APPLY_API, json=pol_apply_json, verify=False, headers=headers) as res:
                    res_dict = json.loads(res.text)
                    if res_dict['code'] == "ok":
                        write_log(f"{schNo}_{schName};{schUtmIp};{POLICY_APPLY_API};OK;")
                    else:
                        write_log(f"{schNo}_{schName};{schUtmIp};{POLICY_APPLY_API};{res_dict['dev_t']};{res_dict['message']}")

            # SYSLOG Setting
            if str(syslog_setting_ip) != 'None':
                syslog_setting_json['svr1_addr'] = syslog_setting_ip
                syslog_setting_json['svr1_port'] = int(syslog_setting_port)
                with s.put(mainUrl + SYSLOG_SETTING_API, json=syslog_setting_json, headers=headers, verify=False) as res:
                    res_dict = json.loads(res.text)
                    if res_dict['code'] == "ok":
                        write_log(f"{schNo}_{schName};{schUtmIp};{SYSLOG_SETTING_API};OK;")
                    else:
                        write_log(
                            f"{schNo}_{schName};{schUtmIp};{SYSLOG_SETTING_API};{res_dict['dev_t']};{res_dict['message']}")

                with s.put(mainUrl + SYSLOG_APPLY_API, headers=headers, verify=False) as res:
                    res_dict = json.loads(res.text)
                    if res_dict['code'] == "ok":
                        write_log(f"{schNo}_{schName};{schUtmIp};{SYSLOG_APPLY_API};OK;")
                    else:
                        write_log(
                            f"{schNo}_{schName};{schUtmIp};{SYSLOG_APPLY_API};{res_dict['dev_t']};{res_dict['message']}")

                with s.put(mainUrl + LOG_SETTING_API, json=log_setting_json, headers=headers, verify=False) as res:
                    res_dict = json.loads(res.text)
                    if res_dict['code'] == "ok":
                        write_log(f"{schNo}_{schName};{schUtmIp};{LOG_SETTING_API};OK;")
                    else:
                        write_log(
                            f"{schNo}_{schName};{schUtmIp};{LOG_SETTING_API};{res_dict['dev_t']};{res_dict['message']}")

                with s.put(mainUrl + LOG_SETTING_APPLY_API, headers=headers, verify=False) as res:
                    res_dict = json.loads(res.text)
                    if res_dict['code'] == "ok":
                        write_log(f"{schNo}_{schName};{schUtmIp};{LOG_SETTING_APPLY_API};OK;")
                    else:
                        write_log(
                            f"{schNo}_{schName};{schUtmIp};{LOG_SETTING_APPLY_API};{res_dict['dev_t']};{res_dict['message']}")

            # SNMP Setting
            if str(snmp_setting) != "None":
                with s.post(mainUrl + SNMP_SETTING_API, json=snmp_setting_json, headers=headers, verify=False) as res:
                    res_dict = json.loads(res.text)
                    if res_dict['code'] == "ok":
                        write_log(f"{schNo}_{schName};{schUtmIp};{SNMP_SETTING_API};OK;")
                    else:
                        write_log(f"{schNo}_{schName};{schUtmIp};{SNMP_SETTING_API};{res_dict['dev_t']};{res_dict['message']}")

                with s.post(mainUrl + SNMP_V3_API, json=snmp_v3_json, headers=headers, verify=False) as res:
                    res_dict = json.loads(res.text)
                    if res_dict['code'] == "ok":
                        write_log(f"{schNo}_{schName};{schUtmIp};{SNMP_V3_API};OK;")
                    else:
                        write_log(f"{schNo}_{schName};{schUtmIp};{SNMP_V3_API};{res_dict['dev_t']};{res_dict['message']}")

                with s.put(mainUrl + SNMP_ENABLE_API, json=snmp_enable_json, headers=headers, verify=False) as res:
                    res_dict = json.loads(res.text)
                    if res_dict['code'] == "ok":
                        write_log(f"{schNo}_{schName};{schUtmIp};{SNMP_ENABLE_API};OK;")
                    else:
                        write_log(f"{schNo}_{schName};{schUtmIp};{SNMP_ENABLE_API};{res_dict['dev_t']};{res_dict['message']}")

                with s.put(mainUrl + SNMP_APPLY_API, headers=headers, verify=False) as res:
                    res_dict = json.loads(res.text)
                    if res_dict['code'] == "ok":
                        write_log(f"{schNo}_{schName};{schUtmIp};{SNMP_APPLY_API};OK;")
                    else:
                        write_log(f"{schNo}_{schName};{schUtmIp};{SNMP_APPLY_API};{res_dict['dev_t']};{res_dict['message']}")

            # ADMIN ACL Setting
            if str(admin_ip_acl_ip) != "None":
                admin_ip_acl_json['ip'] = admin_ip_acl_ip
                with s.post(mainUrl + ADMIN_IP_ACL_API, json=admin_ip_acl_json, headers=headers, verify=False) as res:
                    res_dict = json.loads(res.text)
                    if res_dict['code'] == "ok":
                        write_log(f"{schNo}_{schName};{schUtmIp};{ADMIN_IP_ACL_API};OK;")
                    else:
                        write_log(
                            f"{schNo}_{schName};{schUtmIp};{ADMIN_IP_ACL_API};{res_dict['dev_t']};{res_dict['message']}")

                with s.put(mainUrl + ADMIN_IP_ACL_APPLY_API, headers=headers, verify=False) as res:
                    res_dict = json.loads(res.text)
                    if res_dict['code'] == "ok":
                        write_log(f"{schNo}_{schName};{schUtmIp};{ADMIN_IP_ACL_APPLY_API};OK;")
                    else:
                        write_log(f"{schNo}_{schName};{schUtmIp};{ADMIN_IP_ACL_APPLY_API};{res_dict['dev_t']};{res_dict['message']}")

            # MANAGER ACCOUNT
            if str(dhcp_lease_time) != "None":
                pw = "qhdks00@!"
                key = bytes(secui_helper_key[:32], 'utf-8')
                iv = Random.new().read(AES.block_size)
                raw = bytes(_pad(pw), 'utf-8')
                cipher = AES.new(key, AES.MODE_CBC, iv)
                hex_data = (iv.hex() + cipher.encrypt(raw).hex()).encode('utf-8')
                user_pw = base64.b64encode(hex_data).decode('utf-8')
                manager_account_json1['login_pw'] = user_pw

                pw = "qhdks99@!"
                key = bytes(secui_helper_key[:32], 'utf-8')
                iv = Random.new().read(AES.block_size)
                raw = bytes(_pad(pw), 'utf-8')
                cipher = AES.new(key, AES.MODE_CBC, iv)
                hex_data = (iv.hex() + cipher.encrypt(raw).hex()).encode('utf-8')
                user_pw = base64.b64encode(hex_data).decode('utf-8')
                manager_account_json2['login_pw'] = user_pw

                with s.post(mainUrl + MANAGER_ACCOUNT_CONFIG_API, json=manager_account_json1, headers=headers, verify=False) as res:
                    res_dict = json.loads(res.text)
                    if res_dict['code'] == "ok":
                        write_log(f"{schNo}_{schName};{schUtmIp};{MANAGER_ACCOUNT_CONFIG_API}_JSON1;OK;")
                    else:
                        write_log(f"{schNo}_{schName};{schUtmIp};{MANAGER_ACCOUNT_CONFIG_API}_JSON1;{res_dict['dev_t']};{res_dict['message']}")

                with s.post(mainUrl + MANAGER_ACCOUNT_CONFIG_API, json=manager_account_json2, headers=headers, verify=False) as res:
                    res_dict = json.loads(res.text)
                    if res_dict['code'] == "ok":
                        write_log(f"{schNo}_{schName};{schUtmIp};{MANAGER_ACCOUNT_CONFIG_API}_JSON2;OK;")
                    else:
                        write_log(f"{schNo}_{schName};{schUtmIp};{MANAGER_ACCOUNT_CONFIG_API}_JSON2;{res_dict['dev_t']};{res_dict['message']}")

                with s.put(mainUrl + MANAGER_ACCOUNT_CONFIG_APPLY_API, headers=headers, verify=False) as res:
                    res_dict = json.loads(res.text)
                    if res_dict['code'] == "ok":
                        write_log(f"{schNo}_{schName};{schUtmIp};{MANAGER_ACCOUNT_CONFIG_APPLY_API};OK;")
                    else:
                        write_log(f"{schNo}_{schName};{schUtmIp};{MANAGER_ACCOUNT_CONFIG_APPLY_API};{res_dict['dev_t']};{res_dict['message']}")

            # DHCP SETTING
            if str(dhcp_lease_time) != "None":
                dhcp_config_json['normal_serv_rent_tm'] = int(dhcp_lease_time)
                with s.put(mainUrl + DHCP_CONFIG_API, json=dhcp_config_json, headers=headers, verify=False) as res:
                    res_dict = json.loads(res.text)
                    if res_dict['code'] == "ok":
                        write_log(f"{schNo}_{schName};{schUtmIp};{DHCP_CONFIG_API};OK;")
                    else:
                        write_log(f"{schNo}_{schName};{schUtmIp};{DHCP_CONFIG_API};{res_dict['dev_t']};{res_dict['message']}")

                with s.put(mainUrl + DHCP_CONFIG_APPLY_API, headers=headers, verify=False) as res:
                    res_dict = json.loads(res.text)
                    if res_dict['code'] == "ok":
                        write_log(f"{schNo}_{schName};{schUtmIp};{DHCP_CONFIG_APPLY_API};OK;")
                    else:
                        write_log(f"{schNo}_{schName};{schUtmIp};{DHCP_CONFIG_APPLY_API};{res_dict['dev_t']};{res_dict['message']}")

                with s.put(mainUrl + ha_api + DHCP_CONFIG_API, json=dhcp_config_json, headers=headers, verify=False) as res:
                    res_dict = json.loads(res.text)
                    if res_dict['code'] == "ok":
                        write_log(f"{schNo}_{schName};{schUtmIp};{ha_api}{DHCP_CONFIG_API};OK;")
                    else:
                        write_log(f"{schNo}_{schName};{schUtmIp};{ha_api}{DHCP_CONFIG_API};{res_dict['dev_t']};{res_dict['message']}")

                with s.put(mainUrl + ha_api + DHCP_CONFIG_APPLY_API, headers=headers, verify=False) as res:
                    res_dict = json.loads(res.text)
                    if res_dict['code'] == "ok":
                        write_log(f"{schNo}_{schName};{schUtmIp};{ha_api}{DHCP_CONFIG_APPLY_API};OK;")
                    else:
                        write_log(f"{schNo}_{schName};{schUtmIp};{ha_api}{DHCP_CONFIG_APPLY_API};{res_dict['dev_t']};{res_dict['message']}")

            # ADMIN GENERIC Setting
            if str(set_gui_port) != "None":
                manager_access_port_json['cli_port'] = int(set_cli_port)
                manager_access_port_json['gui_port'] = int(set_gui_port)
                with s.put(mainUrl + MANAGER_ACCESS_PORT_API, json=manager_access_port_json, headers=headers, verify=False) as res:
                    res_dict = json.loads(res.text)
                    if res_dict['code'] == "ok":
                        write_log(f"{schNo}_{schName};{schUtmIp};{MANAGER_ACCESS_PORT_API};OK;")
                    else:
                        write_log(f"{schNo}_{schName};{schUtmIp};{MANAGER_ACCESS_PORT_API};{res_dict['dev_t']};{res_dict['message']}")

                with s.put(mainUrl + MANAGER_ACCESS_PORT_APPLY_API, headers=headers, verify=False) as res:
                    res_dict = json.loads(res.text)
                    if res_dict['code'] == "ok":
                        write_log(f"{schNo}_{schName};{schUtmIp};{MANAGER_ACCESS_PORT_APPLY_API};OK;")
                    else:
                        write_log(f"{schNo}_{schName};{schUtmIp};{MANAGER_ACCESS_PORT_APPLY_API};{res_dict['dev_t']};{res_dict['message']}")

            # LOGOUT
            # with s.delete(mainUrl + LOGOUT_API, headers=headers, verify=False) as res:
            #     res_dict = json.loads(res.text)

        print_progress(row - 1, ws.max_row - 1, 'Progress:', 'Complete ', 1, 50)