import openpyxl
import os, sys
import socket

from datetime import date
import time

START_DATE = date.today()

def port_check(ip, port):
    try:
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(1)
        result = sock.connect_ex((ip, port))
        sock.close()
    except socket.error as e:
        result = 1
        print("Error : " + e)

    if result == 0:
        return True
    else:
        return False


def write_log(string):
    dat = str(START_DATE).replace("-", "")
    f = open(CUR_PATH + f'\\log_{dat}.txt', "a+")
    now = time.localtime()
    cur_time = "%02d/%02d,%02d:%02d:%02d" % (now.tm_mon, now.tm_mday, now.tm_hour, now.tm_min, now.tm_sec)
    f.write(f'{cur_time};{string}\n')
    f.close()


def printProgress(iteration, total, prefix='', suffix='', decimals=1, barLength=100):
    formatStr = "{0:." + str(decimals) + "f}"
    percent = formatStr.format(100 * (iteration / float(total)))
    filledLength = int(round(barLength * iteration / float(total)))
    bar = '#' * filledLength + '-' * (barLength - filledLength)
    sys.stdout.write('\r%s |%s| %s%s %s' % (prefix, bar, percent, '%', suffix)),
    if iteration == total:
        sys.stdout.write('\n')
    sys.stdout.flush()


if __name__ == "__main__":
    CUR_PATH = os.getcwd()
    SRC_FILE = CUR_PATH + "\\bluemax_portcheck.xlsx"

    wb = openpyxl.load_workbook(SRC_FILE)
    ws = wb.active
    cnt = 0

    for row in range(2, ws.max_row + 1):
        schNo = ws[f'A{row}'].value
        schHaCls = ws[f'B{row}'].value
        schName = ws[f'D{row}'].value
        utm_rip1 = ws[f'E{row}'].value
        utm_rip2 = ws[f'F{row}'].value
        utm_access_port = ws[f'G{row}'].value

        port_check_val1 = ws[f'H{row}'].value
        port_check_val2 = ws[f'I{row}'].value

        if str(utm_rip1) != "None" or str(port_check_val1) != "None":
            if port_check(utm_rip1, port_check_val1):
                ws[f'J{row}'].value = "O"
            else:
                ws[f'J{row}'].value = "X"

        if str(utm_rip1) != "None" or str(port_check_val2) != "None":
            if port_check(utm_rip1, port_check_val2):
                ws[f'K{row}'].value = "O"
            else:
                ws[f'K{row}'].value = "X"

        if str(utm_rip2) != "None" or str(port_check_val1) != "None":
            if port_check(utm_rip2, port_check_val1):
                ws[f'L{row}'].value = "O"
            else:
                ws[f'L{row}'].value = "X"

        if str(utm_rip2) != "None" or str(port_check_val2) != "None":
            if port_check(utm_rip2, port_check_val2):
                ws[f'M{row}'].value = "O"
            else:
                ws[f'M{row}'].value = "X"

        printProgress(row, ws.max_row, 'Progress:', 'Complete ', 1, 50)

        cnt += 1
        if cnt % 10 == 0:
            wb.save(SRC_FILE)

    wb.save(SRC_FILE)

