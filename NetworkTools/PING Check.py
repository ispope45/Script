import socket
import os
import openpyxl

# GLOBAL
OS_HOME_DRIVE = os.environ['HOMEDRIVE']
OS_HOME_PATH = os.environ['HOMEPATH']
HOME_PATH = OS_HOME_DRIVE + OS_HOME_PATH

SRC_PATH = HOME_PATH + '\\Desktop\\src\\'
DST_PATH = HOME_PATH + '\\Desktop\\dst\\'

# SRC_FILE = SRC_PATH + 'Dev_List_20210706_PoE.xlsx'
# SRC_FILE = SRC_PATH + 'Dev_List_20210706_AP.xlsx'
SRC_FILE = SRC_PATH + 'Davo_list.xlsx'

# test_ip = ['192.168.0.254', '192.168.0.253', '192.168.0.252', '192.168.0.251', '192.168.0.250']
# test_port = [22, 80, 50005]
form_desc = '■ PING Test\n' \
            'Read IP and PING test\n' \
            'A:No(r) / B:Name(r) / C:IP(r) / D:Port(r) / E:Desc(w)\n'

START_LINE = 0
END_LINE = 0


# FUNCTION
def ping_check(ip):
    try:
        response = os.system("ping -n 1 " + ip)
        if response == 0:
            return True
        else:
            return False

    except Exception as e:
        print("Error : " + e)


if __name__ == "__main__":
    print(form_desc)

    wb = openpyxl.load_workbook(SRC_FILE)
    ws = wb.active

    cnt = 0

    if START_LINE == 0:
        startRow = 2
    else:
        startRow = START_LINE

    if END_LINE == 0:
        totalRows = ws.max_row + 1
    else:
        totalRows = END_LINE

    for i in range(startRow, totalRows):
        row = str(i)

        val_devName = ws['B' + row].value
        val_ip = ws['C' + row].value

        print(f"########### {i - 1} / {totalRows - 2} ###########")
        res = ping_check(val_ip)
        if res:
            print(f"{val_devName} / {str(val_ip)} PING Ok\n")
            ws['E' + row].value = "PINK OK"
        else:
            print(f"{val_devName} / {str(val_ip)} Not Connected\n")
            ws['E' + row].value = "Not Connected"

        cnt = cnt + 1
        if cnt == 50:
            wb.save(filename=SRC_FILE)
            cnt = 0

    wb.save(filename=SRC_FILE)
