import socket
import os
import openpyxl
import time
from icmplib import ping

# GLOBAL
OS_HOME_DRIVE = os.environ['HOMEDRIVE']
OS_HOME_PATH = os.environ['HOMEPATH']
HOME_PATH = OS_HOME_DRIVE + OS_HOME_PATH

SRC_PATH = HOME_PATH + '\\Desktop\\src\\'
DST_PATH = HOME_PATH + '\\Desktop\\dst\\'

# SRC_FILE = SRC_PATH + 'Dev_List_20210706_PoE.xlsx'
SRC_FILE = SRC_PATH + 'Dev_List_PoE.xlsx'

# test_ip = ['192.168.0.254', '192.168.0.253', '192.168.0.252', '192.168.0.251', '192.168.0.250']
# test_port = [22, 80, 50005]
form_desc = '■ TCP Port Test\n' \
            'Read IP, Port number and TCP Port test\n' \
            'A:No(r) / B:Name(r) / C:IP(r) / D:Port(r) / E:Desc(w)\n'

START_LINE = 0
END_LINE = 0


# FUNCTION
def port_check(ip, port):
    try:
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(1)
        result = sock.connect_ex((ip, port))
        sock.close()
    except socket.error as e:
        result = 1
        print("Error : " + e)

    if result == 0:
        return True
    else:
        return False


def ping_check(ip):
    try:
        response = ping(ip, count=1, timeout=1)
        if response.is_alive:
            return True
        else:
            return False

    except Exception as e:
        print("Error : " + e)


if __name__ == "__main__":
    print(form_desc)
    startTime = time.time()

    wb = openpyxl.load_workbook(SRC_FILE)
    ws = wb.active

    cnt = 0

    if START_LINE == 0:
        startRow = 2
    else:
        startRow = START_LINE

    if END_LINE == 0:
        totalRows = ws.max_row + 1
    else:
        totalRows = END_LINE

    for i in range(startRow, totalRows):
        row = str(i)

        val_devName = ws['B' + row].value
        val_ip = ws['C' + row].value
        val_port = ws['D' + row].value
        nowTime = time.time()

        print(f"########### {i - 1} / {totalRows - 2} ########### Remain Time : {nowTime - startTime}")

        res1 = ping_check(val_ip)
        if res1:
            print(f"{val_devName} / {str(val_ip)}:ICMP Opened\n")
            ws['E' + row].value = "Opened"
        else:
            print(f"{val_devName} / {str(val_ip)}:ICMP Not Connected\n")
            ws['E' + row].value = "Not Connected"

        res2 = port_check(val_ip, val_port)
        if res2:
            print(f"{val_devName} / {str(val_ip)}:{str(val_port)} Opened\n")
            ws['F' + row].value = "Opened"
        else:
            print(f"{val_devName} / {str(val_ip)}:{str(val_port)} Not Connected\n")
            ws['F' + row].value = "Not Connected"

        cnt = cnt + 1
        if cnt == 1000:
            wb.save(filename=SRC_FILE)
            cnt = 0

    wb.save(filename=SRC_FILE)
