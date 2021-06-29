from openpyxl import load_workbook, Workbook
import glob
'''
여러 엑셀파일의 시트안 특정값 추출하여 1개의 엑셀파일로 정리

'''
f_list = glob.glob("C:\\Users\\Jungly\\Desktop\\아비\\*")
print(f_list)
aa = f_list[0].split('\\')
#var 선언
fileList = []
fileNum = []
fileName = []
i = 0
row = 0

for v in f_list:
    fileName = v.split('\\')
    if fileName[5].find("~$") == -1:
        fileList.append(fileName[5])

for v1 in fileList:
    #print(v1)
    tmp = v1.split('_')
    fileNum.append(tmp[0])

write_wb = Workbook()
write_ws = write_wb.active

for vv in fileList:
    print(vv)
    load_wb = load_workbook("C:\\Users\\Jungly\\Desktop\\아비\\" + str(vv), data_only=True)
    load_ws = load_wb.active

    idx = vv.split('_')

    align_num = 1

    collist = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']
    collist2 = ['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']

    for a in range(99, 153):
        row = row + 1
        write_ws['A' + str(row)] = int(align_num)
        write_ws['B' + str(row)] = int(idx[0])
        write_ws['C' + str(row)] = idx[1]
        write_ws['D' + str(row)] = idx[2]
        for colnum in range(len(collist)):
            write_ws[collist2[colnum] + str(row)] = load_ws[collist[colnum] + str(a)].value

        align_num = align_num + 1

    i = i + 1

write_wb.save("C:\\Users\\Jungly\\Desktop\\test.xlsx")
