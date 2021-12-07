import requests
import time
import base64
import binascii
import hashlib
import openpyxl
import os

MAIN_URL = "https://192.168.0.254:50005/"
LOGIN_API = "login/loginAction"
BACKUP_API = "firewall/firewall/ipv4Policy/list?json=true&exportAction=true&queryString=&"

proxy = {'https': 'http://127.0.0.1:8080'}

login_data = {'disconnectType': '',
              'webDeviceForceLogin': '',
              'force': 'false',
              'userId': 'manager',
              'password': 'qwe123!@#',
              'language': 0}
CUR_PATH = os.getcwd()
SRC_DIR = CUR_PATH + "\\"
SRC_FILE = "form.xlsx"


if __name__ == "__main__":
    wb = openpyxl.load_workbook(SRC_DIR + SRC_FILE)
    ws = wb.active

    for row in range(2, ws.max_row + 1):
        no = str(ws[f'A{row}'].value)
        org = ws[f'B{row}'].value
        name = ws[f'C{row}'].value
        utmIp = ws[f'D{row}'].value

        MAIN_URL = f"https://{utmIp}:50005/"

        with requests.Session() as s:
            with s.post(MAIN_URL + LOGIN_API, proxies=proxy, data=login_data, verify=False) as res:
                print(res.headers)
                print(res.text)

            with s.get(MAIN_URL + BACKUP_API, proxies=proxy, verify=False) as res:
                print(res.headers)
                print(res.text)
                result = res.text.replace("\n\n", "\n")
                f = open(SRC_DIR + f"{no}_{org}_{name}.csv", 'w', newline='')
                f.write(result)
                f.close()
        s.close()
    wb.close()
    # # Initialize
    # with s.get(MAIN_URL, proxies=proxy, verify=False) as res:
    #     val = res.text
    #     find_num = val.find("csrf_token")
    #     csrf_token = val[find_num+51:find_num+115]
    #     pw = "secui00@!"
    #     iv = Random.new().read(AES.block_size)
    #
    #     key = bytes(csrf_token[:32], 'utf-8')
    #     raw = bytes(_pad(pw), 'utf-8')
    #     cipher = AES.new(key, AES.MODE_CBC, iv)
    #     hex_data = (iv.hex() + cipher.encrypt(raw).hex()).encode('utf-8')
    #     login_pw = base64.b64encode(hex_data).decode('utf-8')
    #
    #     login_data = make_login_data(login_pw, csrf_token)
    #
    # # Login Phase
    # with s.post(MAIN_URL + LOGIN_API, json=login_data, proxies=proxy, verify=False) as res:
    #     print(res.cookies.get_dict())
    #     cookies = res.cookies.get_dict()
    #     print(res.headers)
    #     # print(type(cookies))
    #     # "_static = 88d4bcd01551a342f397603aace6b3dd74fcab1d77800188e0b0e8bf058bbdf3"
    #     print(cookies)
    #     val = res.json()
    #     # e_id = val['token']['id']
    #     print(res.text)
    #     res_dict = json.loads(res.text)
    #     print(res_dict['result']['api_token'])
    #     auth_key = res_dict['result']['api_token']
    #     headers = {'Authorization': auth_key}