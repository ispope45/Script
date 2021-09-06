import openpyxl
import xlrd
import os
from xml.etree.ElementTree import parse

# 전체 0:0
START_LINE = 0
END_LINE = 0


# Excel File extender is xls to xlsx
def convert_xls_to_xlsx(xls_file_path):
    xlsBook = xlrd.open_workbook(xls_file_path)
    workbook = openpyxl.Workbook()

    for i in range(0, xlsBook.nsheets):
        xlsSheet = xlsBook.sheet_by_index(i)
        sheet = workbook.active if i == 0 else workbook.create_sheet()
        sheet.title = xlsSheet.name

        for row in range(0, xlsSheet.nrows):
            for col in range(0, xlsSheet.ncols):
                colvalue = xlsSheet.cell_value(row, col)
                if isinstance(colvalue, str):
                    colvalue = colvalue.replace('L', ' ', 3)

                sheet.cell(row=row + 1, column=col + 1).value = colvalue
    return workbook


def parse_xml(src_file, xmlroot, arg):
    tree = parse(src_file)
    root = tree.getroot()
    xml_root = root.findall(xmlroot)

    xml_val = list()
    for i in range(0, len(arg)):
        xml_val.append([x.findtext(arg[i]) for x in xml_root])

    return xml_val


if __name__ == "__main__":

    filename = "C:\\Users\\Jungly\\Downloads\\물품보유현황_20210906.xlsx"
    classfication_filename = "C:\\Users\\Jungly\\Downloads\\classification.xml"

    print(filename.find(".xlsx"))

    if filename.find(".xlsx") == -1:
        # load_wb = convert_xls_to_xlsx(FILENAME)
        b_filename = filename
        a_filename = filename.replace(".xls", ".xlsx")
        os.rename(b_filename, a_filename)
        filename = a_filename

    load_wb = openpyxl.load_workbook(filename, data_only=True)
    load_ws = load_wb.active

    if START_LINE == 0:
        startRow = 2
    else:
        startRow = START_LINE

    if END_LINE == 0:
        totalRows = load_ws.max_row + 1
    else:
        totalRows = END_LINE

    xml_root = "classification"
    xml_src = ["classname", "midclass", "highclass"]
    parseData = parse_xml(classfication_filename, xml_root, xml_src)

    classname = parseData[0]
    midclass = parseData[1]
    highclass = parseData[2]

    load_ws.insert_cols(2)
    load_ws.insert_cols(2)
    load_ws['B1'].value = "대분류"
    load_ws['C1'].value = "중분류"

    load_ws.Range(f'A2:Q{str(totalRows)}').Sort(Key1=load_ws.Range('B1:D1'), Order1=1, Orientation=1)

    for i in range(startRow, totalRows):
        row = str(i)
        idx = classname.index(load_ws['D' + row].value)

        load_ws['B' + row].value = midclass[idx]
        load_ws['C' + row].value = midclass[idx]

    # print(classname)
    # print(len(classname))
    # print(midclass)
    # print(len(midclass))
    # print(highclass)
    # print(len(highclass))
    #
    # idx = classname.index('강연대')
    # print(classname.index('강연대'))
    # print(classname[idx])
    # print(midclass[idx])
    # print(highclass[idx])





    load_wb.save(filename)

    # sheets = load_wb.sheetnames
    #
    # print(sheets)


