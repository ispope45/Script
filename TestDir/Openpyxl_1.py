from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.merge_cells('B2:F4')
top_left_cell = ws['B2']
top_left_cell.value = "My Cell"

thin = Side(border_style="thin", color="000000")
double = Side(border_style="double", color="ff0000")

top_left_cell.border = Border(top=double, left=thin, right=thin, bottom=double)
top_left_cell.fill = PatternFill("solid", fgColor="DDDDDD")
top_left_cell.fill = fill = GradientFill(stop=("000000", "FFFFFF"))
top_left_cell.font = Font(b=True, color="FF0000")
top_left_cell.alignment = Alignment(horizontal="center", vertical="center")


def merge_border_set(ws, s_row, s_col, e_row, e_col):
    wb.save("styled.xlsx")