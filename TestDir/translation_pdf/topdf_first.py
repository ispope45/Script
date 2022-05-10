import sys
import os
import win32com.client
import openpyxl

from datetime import date
import time

CUR_PATH = os.getcwd()
START_DATE = date.today()


def print_progress(iteration, total, prefix='', suffix='', decimals=1, barLength=100):
    formatStr = "{0:." + str(decimals) + "f}"
    percent = formatStr.format(100 * (iteration / float(total)))
    filledLength = int(round(barLength * iteration / float(total)))
    bar = '#' * filledLength + '-' * (barLength - filledLength)
    sys.stdout.write('\r%s |%s| %s%s %s' % (prefix, bar, percent, '%', suffix)),
    if iteration == total:
        sys.stdout.write('\n')
    sys.stdout.flush()


def write_log(string):
    dat = str(START_DATE).replace("-", "")
    f = open(CUR_PATH + f'\\log_{dat}.txt', "a+")
    now = time.localtime()
    cur_time = "%02d/%02d,%02d:%02d:%02d" % (now.tm_mon, now.tm_mday, now.tm_hour, now.tm_min, now.tm_sec)
    f.write(f'{cur_time};{string}\n')
    f.close()


if __name__ == "__main__":

    wdFormatPDF = 17

    SRC_FILE = CUR_PATH + "\\src_20220331.xlsx"
    wb = openpyxl.load_workbook(SRC_FILE)
    ws = wb.active
    fileClass = ["_개통확인서", "_설치확인서"]

    word = win32com.client.Dispatch('Word.Application')
    time.sleep(3)

    for row in range(2, ws.max_row + 1):
        result_filename = ws[f'P{row}'].value

        print_progress(row - 1, ws.max_row - 1, 'Progress:', 'Complete ', 1, 50)

        for cls in fileClass:
            try:
                in_file = os.path.abspath(result_filename + cls + ".doc")
                out_file = os.path.abspath(result_filename + cls + ".pdf")

                doc = word.Documents.Open(in_file)
                doc.SaveAs(out_file, FileFormat=wdFormatPDF)

                doc.Close()

            except Exception as e:
                write_log(f"{result_filename + cls};{e}")

        print_progress(row - 1, ws.max_row - 1, 'Progress:', 'Complete ', 1, 50)

    word.Quit()

