import requests
import openpyxl
import os
from datetime import date

CUR_PATH = os.getcwd()
START_DATE = date.today()

URL = 'http://emis.lkinv.com'

LOGIN_API = '/cms/login.do'
login_param = {
    "kind": "login",
    "f_user_id": "admin",
    "f_user_pw": "admin1234"
}

LIST_API = "/cms/sc_list.do"
list_param = {
    "currPage": 1,
    "company_id": "",
    "s_nm_company": "",
    "s_nm_schoolname": "위례솔중학교",
    "s_nm_area": "",
    "s_nm_collection": "",
    "s_nm_sig": "",
    "s_nm_address": "",
    "s_nm_manager": "",
    "s_nm_manager_phone": "",
    "s_nm_installer": "",
    "s_nm_end": "",
    "s_st_created": "",
    "s_ed_created": "",
    "s_st_updatedate": "",
    "s_ed_updatedate": ""
}

proxy = {'https': 'http://127.0.0.1:8080'}

if __name__ == "__main__":
    SRC_FILE = CUR_PATH + "\\완료학교.xlsx"
    wb = openpyxl.load_workbook(SRC_FILE)
    ws = wb.active

    with requests.Session() as s:
        with s.post(URL + LOGIN_API, data=login_param, verify=False, proxies=proxy) as res:
            print("Login OK")

        for row in range(2, ws.max_row + 1):
            schName = ws[f'B{row}'].value
            schNo = ws[f'D{row}'].value
            schClear = ws[f'C{row}'].value

            if str(schName) == "None":
                continue
            elif str(schNo) == "None":
                continue
            elif schClear != "O":
                continue

            DOWNLOAD_API = "/cms/report_summary.do?school_no=1561&proctype=download&step_cls=2"

            with s.get(URL + DOWNLOAD_API, verify=False, proxies=proxy) as res:

                f1 = open(f'./text1.pdf', 'wb')
                f1.write(res.content)