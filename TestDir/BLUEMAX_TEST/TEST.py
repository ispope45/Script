from SSHConnector import SSHConnector
import requests
import time
import base64
import binascii
import hashlib
import openpyxl
from threading import Thread
from multiprocessing import Process

from Crypto.Cipher import AES
from Crypto import Random
from bs4 import BeautifulSoup as bs
import json

import random


MAIN_URL = 'https://192.168.10.10'
LOGIN_API = '/api/au/login'
HA_API = '/ha:443:ha_grp_2'
IF_API_BASE = '/api/sm/interfaces'
BACKUP_API = '/api/sm/backup/manual'
HOSTNAME_API = '/api/sm/info/equipment'
ROUTING_API = '/api/sm/static-routings/1'
VIP_API = '/api/sm/ha/virtual-ips'
DHCP_API = '/api/sm/dhcp/server/dynamic-areas'
DHCP_API2 = '/api/sm/dhcp/server/config'
DHCP_API3 = '/api/sm/dhcp/server/apply'

BAKCUP_PARAM = {"ha_backup": 1, "target": "POVS"}
proxy = {'https': 'http://127.0.0.1:8080'}
cookies = {'_next_login': 'index', '_static': '88d4bcd01551a342f397603aace6b3dd74fcab1d77800188e0b0e8bf058bbdf3'}

SRC_FILE = "C:\\Users\\work.Jungly\\Desktop\\src.xlsx"


# ### gw1: KT, gw2: SK
def make_routing_config(gw1, gw2, gw3):
    cfg1 = {"type": 0, "dest_addr": "0.0.0.0", "dest_mask": 0, "metric": 0,
            "gw": [
                {"id": f"{str(random.randint(2000, 9999))}_{gw1}", "addr": gw1, "ifc_id": 12, "ifc_name": "eth12",
                 "weight": 2, "itemIndex": 0},
                {"id": f"{str(random.randint(2000, 9999))}_{gw2}", "addr": gw2, "ifc_id": 13, "ifc_name": "eth11",
                 "weight": 1, "ifauto": 0, "invalid": 1, "itemIndex": 1}
                ],
            "gwTostring": "", "route_type": 0, "status_check": 0, "desc": ""}
    cfg2 = {"type": 0, "dest_addr": "0.0.0.0", "dest_mask": 0, "metric": 0,
            "gw": [
                {"id": f"{str(random.randint(2000, 9999))}_{gw3}", "addr": gw3, "ifc_id": 13, "ifc_name": "eth12",
                 "weight": 1, "itemIndex": 0}
            ],
            "gwTostring": "", "route_type": 0, "status_check": 0, "desc": ""}
    return cfg1, cfg2


def make_vip_url():
    print("A")


def make_vip_config(zone, zone_id, mmbr_id, mmbr_uuid, mmbr_hostname, dev_name, vrid, vip, vip_mask):
    cfg1 = {"zone": zone, "zone_id": zone_id, "ha_mmbr_id": mmbr_id, "mmbr_uuid": mmbr_uuid,
            "mmbr_hostname": mmbr_hostname,"ifc_name": dev_name, "vrid": vrid, "ip_ver": 4,
            "ip": vip, "netmask": vip_mask, "rip": None}
    return cfg1


def make_dhcp_config(start, end, net, subnet, gw, idx):
    cfg1 = {"area_index": idx, "start_area": start, "end_area": end, "net_addr": net,
            "subnet_mask": subnet, "default_gw": gw}
    cfg2 = {"normal_serv_use_enable": 1, "normal_serv_rent_tm": 1440, "normal_serv_auth": 0,
            "dns_serv1": "168.126.63.1", "dns_serv2": "8.8.8.8", "wins_serv1": "", "wins_serv2": ""}
    return cfg1, cfg2


def make_login_data(login_pw, csrf_token):
    login_data = {
        "lang": "ko",
        "force": "true",
        "readonly": 0,
        "login_id": "admin",
        "login_pw": login_pw,
        "csrf_token": csrf_token,
        "type": "local"
    }
    return login_data


def _pad(s):
    bs = 16
    return s + (bs - len(s) % bs) * chr(bs - len(s) % bs)


if __name__ == "__main__":
    # Excel Data Formatting
    wb = openpyxl.load_workbook(SRC_FILE)
    ws = wb.active

    for i in range(2, ws.max_row + 1):
        row = str(i)
        s_row = str(i + 1)
        if i % 2 == 1:
            continue

        # COMMON
        no = ws[f'A{row}'].value
        name = ws[f'B{row}'].value

        # MASTER
        m_ha = ws[f'C{row}'].value
        m_COL = [['E', 'F'], ['G', 'H'], ['I', 'J'], ['M', 'N'], ['K', 'L'], ['O', 'P'], ['Q', 'R']]
        m_con_info = '192.168.10.10'
        m_hostname = ws[f'D{row}'].value
        m_ipSet = []
        for item in m_COL:
            m_item = []
            m_item.append(ws[item[0] + row].value)
            m_item.append(ws[item[1] + row].value)
            m_ipSet.append(m_item)

        m_vip1 = ws[f'S{row}'].value
        m_vip2 = ws[f'T{row}'].value
        m_vip3 = ws[f'U{row}'].value
        m_vip9 = ws[f'V{row}'].value
        m2_vip9 = ws[f'V{s_row}'].value
        m_vip9_2 = ws[f'W{row}'].value
        m2_vip9_2 = ws[f'W{s_row}'].value
        m_vip12 = ws[f'X{row}'].value

        m_dhcp_net = ws[f'Y{row}'].value
        m_dhcp_mask = ws[f'Z{row}'].value
        m_dhcp_start = ws[f'AA{row}'].value
        m_dhcp_end = ws[f'AB{row}'].value

        m_gw1 = ws[f'AC{row}'].value
        m_gw2 = ws[f'AD{row}'].value

        # SLAVE

        s_ha = ws[f'C{s_row}'].value
        s_COL = [['E', 'F'], ['G', 'H'], ['I', 'J'], ['K', 'L'], ['M', 'N'], ['Q', 'R']]
        s_con_info = '192.168.10.11'
        s_hostname = ws[f'D{s_row}'].value
        s_ipSet = []
        for item in s_COL:
            s_item = []
            s_item.append(ws[item[0] + s_row].value)
            s_item.append(ws[item[1] + s_row].value)
            s_ipSet.append(s_item)

        s_dhcp_net = ws[f'Y{row}'].value
        s_dhcp_mask = ws[f'Z{row}'].value
        s_dhcp_start = ws[f'AA{row}'].value
        s_dhcp_end = ws[f'AB{row}'].value

        s_gw2 = ws[f'AD{row}'].value

        with requests.Session() as s:
            # WEB GUI Initialize
            with s.get(MAIN_URL, proxies=proxy, verify=False) as res:
                val = res.text
                find_num = val.find("csrf_token")
                csrf_token = val[find_num + 51:find_num + 115]
                pw = "secui00@!"
                iv = Random.new().read(AES.block_size)

                key = bytes(csrf_token[:32], 'utf-8')
                raw = bytes(_pad(pw), 'utf-8')
                cipher = AES.new(key, AES.MODE_CBC, iv)
                hex_data = (iv.hex() + cipher.encrypt(raw).hex()).encode('utf-8')
                login_pw = base64.b64encode(hex_data).decode('utf-8')

                login_data = make_login_data(login_pw, csrf_token)

            # Login Phase
            with s.post(MAIN_URL + LOGIN_API, json=login_data, proxies=proxy, verify=False) as res:
                print(res.cookies.get_dict())
                cookies = res.cookies.get_dict()
                print(res.headers)
                # print(type(cookies))
                # "_static = 88d4bcd01551a342f397603aace6b3dd74fcab1d77800188e0b0e8bf058bbdf3"
                print(cookies)
                val = res.json()
                # e_id = val['token']['id']
                print(res.text)
                res_dict = json.loads(res.text)
                print(res_dict['result']['api_token'])
                auth_key = res_dict['result']['api_token']
                headers = {'Authorization': auth_key}

            # VIP Setting
            dev_name = ['eth1', 'eth1', 'eth2', 'eth2', 'eth3', 'eth3', 'eth9', 'eth9', 'eth9', 'eth9', 'eth12',
                        'eth12']
            zone = [1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 2, 2]
            zone_id = [0, 0, 1, 1, 2, 2, 3, 3, 3, 3, 4, 4]
            vrid = [1, 1, 2, 2, 3, 3, 4, 5, 6, 7, 8, 8]
            vip_id = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12]
            vip = [m_vip1, m_vip1, m_vip2, m_vip2, m_vip3, m_vip3, m_vip9, m2_vip9, m_vip9_2, m2_vip9_2, m_vip12,
                   m_vip12]
            vip_mask = [m_ipSet[0][1], m_ipSet[0][1], m_ipSet[1][1], m_ipSet[1][1], m_ipSet[2][1], m_ipSet[2][1],
                        m_ipSet[4][1], m_ipSet[4][1], m_ipSet[3][1], m_ipSet[3][1], m_ipSet[6][1], m_ipSet[6][1]]
            vip_cfg = []

            # VIP Delete
            for j in range(0, 12):
                if j % 2 == 0:
                    mmbr_id = 1
                    mmbr_uuid = "ha_grp_1"
                    mmbr_hostname = m_hostname
                else:
                    mmbr_id = 2
                    mmbr_uuid = "ha_grp_2"
                    mmbr_hostname = s_hostname

                vip_item = make_vip_config(zone[j], zone_id[j], mmbr_id, mmbr_uuid, mmbr_hostname, dev_name[j],
                                           vrid[j], vip[j], vip_mask[j])
                vip_cfg.append(vip_item)
                with s.delete(MAIN_URL + VIP_API + '/' + str(j + 1), proxies=proxy, headers=headers,
                              verify=False) as res:
                    print(res.text)

            with s.put(MAIN_URL + VIP_API + '/apply', proxies=proxy, headers=headers, verify=False) as res:
                print(res.text)

            s.close()

        # INTERFACE & HOSTNAME Setting
        ssh = SSHConnector()
        # ssh.ssh_connect(s_con_info, s_hostname, s_ipSet, s_ha)
        th1 = Process(target=ssh.ssh_connect, args=(s_con_info, s_hostname, s_ipSet, s_ha))
        ssh2 = SSHConnector()
        # ssh2.ssh_connect(m_con_info, m_hostname, m_ipSet, m_ha)
        th2 = Process(target=ssh2.ssh_connect, args=(m_con_info, m_hostname, m_ipSet, m_ha))

        th1.start()
        th2.start()
        th1.join()
        th2.join()

        print(vip_cfg)
        with requests.Session() as s:
            # WEB GUI Initialize
            with s.get(MAIN_URL, proxies=proxy, verify=False) as res:
                val = res.text
                find_num = val.find("csrf_token")
                csrf_token = val[find_num + 51:find_num + 115]
                pw = "secui00@!"
                iv = Random.new().read(AES.block_size)

                key = bytes(csrf_token[:32], 'utf-8')
                raw = bytes(_pad(pw), 'utf-8')
                cipher = AES.new(key, AES.MODE_CBC, iv)
                hex_data = (iv.hex() + cipher.encrypt(raw).hex()).encode('utf-8')
                login_pw = base64.b64encode(hex_data).decode('utf-8')

                login_data = make_login_data(login_pw, csrf_token)

            # Login Phase
            with s.post(MAIN_URL + LOGIN_API, json=login_data, proxies=proxy, verify=False) as res:
                print(res.cookies.get_dict())
                cookies = res.cookies.get_dict()
                print(res.headers)
                # print(type(cookies))
                # "_static = 88d4bcd01551a342f397603aace6b3dd74fcab1d77800188e0b0e8bf058bbdf3"
                print(cookies)
                val = res.json()
                # e_id = val['token']['id']
                print(res.text)
                res_dict = json.loads(res.text)
                print(res_dict['result']['api_token'])
                auth_key = res_dict['result']['api_token']
                headers = {'Authorization': auth_key}
            # VIP Setting
            for cfg in vip_cfg:
                with s.post(MAIN_URL + VIP_API, json=cfg, proxies=proxy, headers=headers, verify=False) as res:
                    print(res.text)

            with s.put(MAIN_URL + VIP_API + '/apply', proxies=proxy, headers=headers, verify=False) as res:
                print(res.text)

            # Routing Setting
            rt_cfg1, rt_cfg2 = make_routing_config(m_gw1, m_gw2, s_gw2)
            with s.put(MAIN_URL + ROUTING_API, json=rt_cfg1, proxies=proxy, headers=headers, verify=False) as res:
                print(res.text)

            with s.put(MAIN_URL + HA_API + ROUTING_API, json=rt_cfg2, proxies=proxy, headers=headers, verify=False) as res:
                print(res.text)

            # DHCP Setting
            with s.get(MAIN_URL + DHCP_API, proxies=proxy, headers=headers, verify=False) as res:
                var = res.json()
                # var = json.loads(res.text)
                print(var)

                idx = var['result'][0]['area_index']
                dhcp_cfg1, dhcp_cfg2 = make_dhcp_config(m_dhcp_start, m_dhcp_end, m_dhcp_net, m_dhcp_mask, m2_vip9, idx)

            with s.put(MAIN_URL + DHCP_API + '/' + str(idx), json=dhcp_cfg1, proxies=proxy, headers=headers,
                       verify=False) as res:
                print(res.text)

            with s.put(MAIN_URL + DHCP_API2, json=dhcp_cfg2, proxies=proxy, headers=headers, verify=False) as res:
                print(res.text)

            with s.put(MAIN_URL + DHCP_API3, proxies=proxy, headers=headers, verify=False) as res:
                print(res.text)

            with s.get(MAIN_URL + HA_API + DHCP_API, proxies=proxy, headers=headers, verify=False) as res:
                var = res.json()
                # var = json.loads(res.text)
                print(var)

                idx = var['result'][0]['area_index']
                dhcp_cfg1, dhcp_cfg2 = make_dhcp_config(m_dhcp_start, m_dhcp_end, m_dhcp_net, m_dhcp_mask, m2_vip9, idx)

            with s.put(MAIN_URL + HA_API + DHCP_API + '/' + str(idx), json=dhcp_cfg1, proxies=proxy, headers=headers,
                       verify=False) as res:
                print(res.text)

            with s.put(MAIN_URL + HA_API + DHCP_API2, json=dhcp_cfg2, proxies=proxy, headers=headers,
                       verify=False) as res:
                print(res.text)

            with s.put(MAIN_URL + HA_API + DHCP_API3, proxies=proxy, headers=headers, verify=False) as res:
                print(res.text)

            # Backup File Download
            with s.post(MAIN_URL + "/api/sm/backup/manual", json=BAKCUP_PARAM, proxies=proxy, headers=headers, verify=False) as res:
                # print(res.text)
                print(res.headers)
                f = open(f"C:\\Users\\work.Jungly\\Desktop\\{no}_{name}.tar", 'wb',)
                f.write(res.content)
                f.close()
            s.close()

