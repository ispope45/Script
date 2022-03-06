# from SSHConnector import SSHConnector
import requests
import base64
import openpyxl
import json
import os
import paramiko

from multiprocessing import Process

from Cryptodome.Cipher import AES
from Cryptodome import Random

import random
from datetime import date
import time
import urllib3

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

MAIN_URL = 'https://192.168.10.10'
LOGIN_API = '/api/au/login'
HA_API = '/ha:443:ha_grp_2'
BACKUP_API = '/api/sm/backup/manual'
ROUTING_API = '/api/sm/static-routings/1'
VIP_API = '/api/sm/ha/virtual-ips'
DHCP_API = '/api/sm/dhcp/server/dynamic-areas'
DHCP_API2 = '/api/sm/dhcp/server/config'
DHCP_API3 = '/api/sm/dhcp/server/apply'
EQUIP_API = "/api/sm/info/equipment"

BACKUP_PARAM = {"ha_backup": 1, "target": "POVSL"}

proxy = {'https': 'http://127.0.0.1:8080'}

CUR_PATH = os.getcwd()
SRC_FILE = CUR_PATH + "\\src.xlsx"
START_DATE = date.today()


def ssh_make_config(interface, del_cnt, ip_set):
    cmd = list()
    cmd.append(f'interface {interface}\n')
    for delCnt in range(1, del_cnt + 1):
        cmd.append('no ip add\n')

    for ip in ip_set:
        cmd.append(f'ip add {ip}\n')

    cmd.append('exit\n')
    cmd.append('y\n')

    return cmd


class SSHConnector():
    def __init__(self):
        self.preamble = ['y\n', 'config\n']
        self.hostname = ['hostname\n', 'set hostname\n', 'exit\n']
        self.interface = ['interface eth1\n', 'no ip add\n', 'ip add 100.100.100.100/24\n', 'exit\n', 'y\n']

    def main(self, sw_ip, sw_user, sw_pass, sw_port, command_set):
        host = sw_ip
        username = sw_user
        password = sw_pass

        output = list()

        try:
            conn = paramiko.SSHClient()
            conn.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            conn.connect(host, username=username, password=password, port=sw_port, timeout=100)
            channel = conn.invoke_shell()
            time.sleep(3)
            for command in command_set:
                for line in command:
                    channel.send(line)

                    out_data, err_data = self.wait_streams(channel)
                    output.append(out_data)
            return output

        except Exception as e:
            if "port" in str(e):
                return "Port Error"
            if "WinError 10060" in str(e):
                return "Connection Error"
            print(e)
            return "Error"

        finally:
            if conn is not None:
                conn.close()

    @staticmethod
    def wait_streams(channel):
        time.sleep(1)
        out_data = ""
        err_data = ""
        while True:
            time.sleep(2)
            if channel.recv_ready():
                out_data += channel.recv(1000).decode('ascii')
                if channel.recv_stderr_ready():
                    err_data += channel.recv(1000).decode('ascii')
                if out_data.find("#") != -1 or out_data.find("[y|n]") != -1:
                    break

        return out_data, err_data

    def ssh_connect(self, con_info, ip_cfg):
        self.main(con_info[0], con_info[1], con_info[2], con_info[3], ip_cfg)


if __name__ == "__main__":
    # Excel Data Formatting
    SRC_FILE = CUR_PATH + "\\bluemax_configuration.xlsx"
    wb = openpyxl.load_workbook(SRC_FILE)
    ws = wb.active

    for row in range(2, ws.max_row + 1):
        schNo = ws[f'A{row}'].value
        schOrg = ws[f'B{row}'].value
        schHaCls = ws[f'C{row}'].value
        schName = ws[f'D{row}'].value
        schUtmIp = ws[f'E{row}'].value
        schUtmAccessPort = ws[f'F{row}'].value

        set_gui_port = ws[f'J{row}'].value

        ssh_access_id = ws[f'N{row}'].value
        ssh_access_pw = ws[f'O{row}'].value
        ssh_access_port = ws[f'P{row}'].value

        con_info = [schUtmIp, ssh_access_id, ssh_access_pw, ssh_access_port]

        ifCfg_M = [['y\n',
                    'config\n',
                    'admin generic\n',
                    f'set gui_port {set_gui_port}\n',
                    'y\n',
                    'y\n',
                    '\n']]

        ssh = SSHConnector()
        ssh.ssh_connect(con_info, ifCfg_M)

    # time.sleep(5)
    #
    # # INTERFACE & HOSTNAME Setting
    # ssh = SSHConnector()
    # th1 = Process(target=ssh.ssh_connect, args=(con_info[0], ifCfg_M))
    #
    # ssh2 = SSHConnector()
    # th2 = Process(target=ssh2.ssh_connect, args=(con_info[1], ifCfg_S))
    #
    # th1.start()
    # time.sleep(30)
    # th2.start()
    #
    # th1.join()
    # th2.join()
