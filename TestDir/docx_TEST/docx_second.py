import openpyxl
import os, sys

CUR_PATH = os.getcwd()


if __name__ == "__main__":

    SRC_FILE = CUR_PATH + "\\backup.xlsx"
    wb = openpyxl.load_workbook(SRC_FILE)
    ws = wb.active

    target_word1 = '>교사망<'
    new_word1 = '>업무망<'

    target_word2 = '>학생망<'
    new_word2 = '>이용자망<'

    for row in range(2, ws.max_row + 1):

        filename = ws[f'P{row}'].value + "_개통확인서.doc"

        # text_file_path = "2.doc"
        new_text_content = ''

        with open(filename, 'r', encoding='utf8') as f:
            lines = f.readlines()

            for i, l in enumerate(lines):
                # new_string = l.strip().replace(target_word1, new_word1)
                # new_text_content += new_string
                #
                # new_string = l.strip().replace(target_word2, new_word2)
                # new_text_content += new_string
                if i in range(391, 409):
                    continue
                else:
                    new_string = l.strip().replace(target_word1, new_word1).replace(target_word2, new_word2)
                    if new_string:
                        new_text_content += new_string

                # print(l)
                # new_string = l.strip().replace(target_word, new_word)

        with open(filename, 'wb') as f:
            res_text = bytes(new_text_content, 'utf-8')
            f.write(res_text)
