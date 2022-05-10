import requests
import openpyxl
import os
import sys
import time

from datetime import date

CUR_PATH = os.getcwd()
START_DATE = date.today()

URL = 'http://emis.lkinv.com'

LOGIN_API = '/cms/login.do'
login_param = {
    "kind": "login",
    "f_user_id": "admin",
    "f_user_pw": "admin1234"
}

LIST_API = "/cms/sc_list.do"
list_param = {
    "currPage": 1,
    "company_id": "",
    "s_nm_company": "",
    "s_nm_schoolname": "위례솔중학교",
    "s_nm_area": "",
    "s_nm_collection": "",
    "s_nm_sig": "",
    "s_nm_address": "",
    "s_nm_manager": "",
    "s_nm_manager_phone": "",
    "s_nm_installer": "",
    "s_nm_end": "",
    "s_st_created": "",
    "s_ed_created": "",
    "s_st_updatedate": "",
    "s_ed_updatedate": ""
}

proxy = {'https': 'http://127.0.0.1:8080'}


def write_log(string):
    dat = str(START_DATE).replace("-", "")
    f = open(CUR_PATH + f'\\log_{dat}.txt', "a+")
    now = time.localtime()
    cur_time = "%02d/%02d,%02d:%02d:%02d" % (now.tm_mon, now.tm_mday, now.tm_hour, now.tm_min, now.tm_sec)
    f.write(f'{cur_time};{string}\n')
    f.close()


def print_progress(iteration, total, prefix='', suffix='', decimals=1, barLength=100):
    formatStr = "{0:." + str(decimals) + "f}"
    percent = formatStr.format(100 * (iteration / float(total)))
    filledLength = int(round(barLength * iteration / float(total)))
    bar = '#' * filledLength + '-' * (barLength - filledLength)
    sys.stdout.write('\r%s |%s| %s%s %s' % (prefix, bar, percent, '%', suffix)),
    if iteration == total:
        sys.stdout.write('\n')
    sys.stdout.flush()


if __name__ == "__main__":
    SRC_FILE = CUR_PATH + "\\완료학교.xlsx"
    wb = openpyxl.load_workbook(SRC_FILE)
    ws = wb.active

    with requests.Session() as s:
        with s.post(URL + LOGIN_API, data=login_param, verify=False, proxies=proxy) as res:
            print("Login OK")

        for row in range(2, ws.max_row + 1):
            schName = ws[f'B{row}'].value
            schNo = ws[f'D{row}'].value
            schClear = ws[f'C{row}'].value

            print_progress(row - 1, ws.max_row - 1, 'Progress:', 'Complete ', 1, 50)

            if str(schName) == "None":
                continue
            elif str(schNo) == "None":
                continue
            elif schClear != "O":
                continue

            DOWNLOAD_API = f"/cms/report_summary.do?school_no={str(schNo)}&proctype=download&step_cls=2"
            try:
                with s.get(URL + DOWNLOAD_API, verify=False, proxies=proxy) as res:
                    f1 = open(f'./{schName}.pdf', 'wb')
                    f1.write(res.content)
                    write_log(f"{schName};{DOWNLOAD_API};OK")
            except Exception as e:
                write_log(f"{schName};{e}")

