from win32com.client import Dispatch
from copy import copy
from openpyxl import Workbook, load_workbook, worksheet
import glob

PATH = "C:\\Users\\Jungly\\Desktop\\"

'''
엑셀 특정 시트를 찾아("2.설치 사진첩(공통)") 각각 별도파일로 저장
'''
f_list = glob.glob(PATH + "src1\\*")

worksheet.copier.copy_worksheet()
'''
a = 0
for f in f_list:
    print(f)
    load_wb = load_workbook(f)
    for s in range(len(load_wb.sheetnames)):
        if load_wb.sheetnames[s] == '설치 사진첩':
            a = s
            break

    load_wb.active = a
    print(load_wb.active)
    load_wb.save(f)

'''