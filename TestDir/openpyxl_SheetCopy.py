from openpyxl import Workbook, load_workbook

load_wb = load_workbook("C:\\Users\\Jungle\\Desktop\\TEST3\\1_서부_서울고은초등학교_SKB준공서류.xlsx", data_only=True)
sheets = load_wb.sheetnames

for s in sheets:
    if s != '2.설치 사진첩(공통)':
        sheet_name = load_wb.get_sheet_by_name(s)
        load_wb.remove_sheet(sheet_name)

#write_ws = write_wb.copy_worksheet(load_ws)

load_wb.save("C:\\Users\\Jungle\\Desktop\\test3.xlsx")