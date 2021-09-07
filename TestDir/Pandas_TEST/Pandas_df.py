import pandas as pd
import openpyxl

from openpyxl.styles import Font
from xml.etree.ElementTree import parse

START_LINE = 0
END_LINE = 0


def parse_xml(src_file, xmlroot, arg):
    tree = parse(src_file)
    root = tree.getroot()
    xml_root = root.findall(xmlroot)

    xml_val = list()
    for i in range(0, len(arg)):
        xml_val.append([x.findtext(arg[i]) for x in xml_root])

    return xml_val


if __name__ == "__main__":

    filename = "C:\\Users\\Jungly\\Downloads\\물품보유현황_20210906_1.xlsx"
    classfication_filename = "C:\\Users\\Jungly\\Downloads\\classification.xml"

    load_wb = openpyxl.load_workbook(filename, data_only=True)
    load_ws = load_wb.active

    if START_LINE == 0:
        startRow = 2
    else:
        startRow = START_LINE

    if END_LINE == 0:
        totalRows = load_ws.max_row + 1
    else:
        totalRows = END_LINE

    xml_root = "classification"
    xml_src = ["classname", "midclass", "highclass"]
    parseData = parse_xml(classfication_filename, xml_root, xml_src)

    classname = parseData[0]
    midclass = parseData[1]
    highclass = parseData[2]

    load_ws.insert_cols(2)
    load_ws.insert_cols(2)
    load_ws['B1'].value = "대분류"
    load_ws['C1'].value = "중분류"

    for i in range(startRow, totalRows):
        row = str(i)
        try:
            idx = classname.index(load_ws['D' + row].value)

            load_ws['B' + row].value = midclass[idx]
            load_ws['C' + row].value = midclass[idx]
        except Exception as e:
            print(e)
            load_ws['B' + row].value = "Not classified value"
            load_ws['C' + row].value = "Not classified value"

    load_wb.save("C:\\Users\\Jungly\\Downloads\\물품보유현황_20210906_2.xlsx")

    test = pd.read_excel("C:\\Users\\Jungly\\Downloads\\물품보유현황_20210906_2.xlsx")
    print(test)
    test = test.sort_values(by=['운용부서', '대분류', '중분류', '물품분류명'], ascending=[True, True, True, True])
    # test_list = test.columns.values.tolist() + test.values.tolist()
    condition = test.drop_duplicates(['운용부서'])['운용부서'].tolist()

    print(condition)

    # condition2 = condition['운용부서'].tolist()
    # print(condition2)
    # print(test_list)
    # print(test)
    # print(type(test))
    # print(test_list)
    src_wb = openpyxl.load_workbook("C:\\Users\\Jungly\\Downloads\\sampledata.xlsx")
    src_ws = src_wb.active

    for filter in condition:
        test_list = test[test['운용부서'] == filter].values.tolist()

        proc_ws = src_wb.copy_worksheet(src_ws)
        proc_ws.title = filter
        cnt = 0
        for data in test_list:
            cnt += 1
            data.insert(0, cnt)
            data.insert(10, '')
            data.insert(15, '')
            data.insert(16, '')
            proc_ws.append(data)
            proc_ws['B1'].value = filter

        for data in test_list:
            for row in range(2, len(data)):
                proc_ws.row_dimensions[row].height = 100
                for col in range(1, 21):
                    proc_ws.cell(row, col).font = Font(size=20)
    src_wb.save("C:\\Users\\Jungly\\Downloads\\sampledata3.xlsx")

    # test.to_excel("C:\\Users\\Jungly\\Downloads\\물품보유현황_20210906_3.xlsx")