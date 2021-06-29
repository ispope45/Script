from openpyxl import workbook, load_workbook

PATH = "C:\\Users\\Jungly\\Desktop\\"

load_wb = load_workbook(filename=PATH+"test1.xlsx")
load_ws = load_wb["시공사진_샘플"]

load_wb2 = load_workbook(filename=PATH+"test2.xlsx")
load_ws2 = load_wb2["1.설치 전 사진"]
load_ws3 = load_wb2["2.설치 후 사진(랙)"]
load_ws4 = load_wb2["3.설치 후 사진(교환기 등)"]

print(type(load_wb))
print(load_wb2.index(load_ws4))
print(type(load_ws))

write_wb = workbook.Workbook()
write_ws = write_wb.active

write_wb.copy_worksheet(load_ws)

write_wb.save(PATH + "test3.xlsx")

#wb = workbook.Workbook(filename=PATH+"test1.xlsx")
#ws = wb['']

#wb.move_sheet()

