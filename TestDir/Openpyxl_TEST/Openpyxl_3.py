import openpyxl
import os
from openpyxl.drawing.image import Image

# ### GLOBAL ########
OS_HOME_DRIVE = os.environ['HOMEDRIVE']
OS_HOME_PATH = os.environ['HOMEPATH']
HOME_PATH = OS_HOME_DRIVE + OS_HOME_PATH

SRC_PATH = HOME_PATH + '\\Desktop\\src\\'
DST_PATH = HOME_PATH + '\\Desktop\\dst\\'


srcWorkbook = openpyxl.load_workbook("C:\\Users\\Jungly\\Desktop\\33.xlsx")
srcWorkSheet = srcWorkbook.active  # 특정 시트 지정할경우 ws = wb['TEST1'] 의 형태로

START_LINE = 0
END_LINE = 0

if START_LINE == 0:
    startRow = 2
else:
    startRow = START_LINE

if END_LINE == 0:
    totalRows = srcWorkSheet.max_row + 1
else:
    totalRows = END_LINE

val = []
val2 = []

for i in range(startRow, totalRows):
    if srcWorkSheet['F' + str(i)].value == '':
        val.append('None')
        val2.append(['None', i])
    else:
        val.append(srcWorkSheet['F' + str(i)].value)
        val2.append([srcWorkSheet['F' + str(i)].value, i])

val2.sort()

if not (os.path.isdir(DST_PATH)):
    os.mkdir(SRC_PATH)
tmp = ''
for v1 in val2:
    fileName = str(v1[0]) + ".xlsx"
    rowNum = str(v1[1])
    if tmp == fileName:
        dstWorkbook = openpyxl.Workbook()
        dstWorkSheet = dstWorkbook.active

    colA = srcWorkSheet['A' + rowNum].value
    colB = srcWorkSheet['B' + rowNum].value
    colC = srcWorkSheet['C' + rowNum].value
    colD = srcWorkSheet['D' + rowNum].value
    colE = srcWorkSheet['E' + rowNum].value
    colF = srcWorkSheet['F' + rowNum].value
    colG = srcWorkSheet['G' + rowNum].value
    colH = srcWorkSheet['H' + rowNum].value
    colI = srcWorkSheet['I' + rowNum].value
    colJ = srcWorkSheet['J' + rowNum].value
    colK = srcWorkSheet['K' + rowNum].value
    colL = srcWorkSheet['L' + rowNum].value
    colM = srcWorkSheet['M' + rowNum].value
    colN = srcWorkSheet['N' + rowNum].value
    colO = srcWorkSheet['O' + rowNum].value


    # if not (os.path.isfile(DST_PATH + fileName)):




print(val)
print(len(val))

res = list(set(val))
print(res)
print(len(res))

print(val2)
print(len(val2))

val2.sort()
print(val2)


wb2.close()
