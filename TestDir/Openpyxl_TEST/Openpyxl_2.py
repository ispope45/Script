import openpyxl
from openpyxl.drawing.image import Image

# 선언
wb = openpyxl.Workbook()
wb2 = openpyxl.load_workbook("C:\\Users\\Jungly\\Desktop\\3.xlsx")

# 워크시트
wb.create_sheet('TEST1', 0)  # 0 : 왼쪽에 생성, 1 : 오른쪽에 생성
ws = wb.active  # 특정 시트 지정할경우 ws = wb['TEST1'] 의 형태로
ws2 = wb2['TEST1']

# 값입력
ws['A1'].value = "AAA"
ws.cell(row=1, column=2).value = "BBB"
ws.cell(row=2, column=1).value = "AAA2"
ws.cell(row=2, column=3).value = "CCC"
print(ws.cell(row=1, column=1).value)
print(ws['A1'].value)

# 행, 열 길이 반환
print(len(ws['1']))  # 1행의 열 길이 (A,B,C)
print(len(ws['A']))  # A열의 행 길이
print(len(ws['B']))  # B열의 행 길이
print(ws.max_row)  # 시트의 행 길이
print(ws.max_column)  # 시트의 열 길이
# 1 , 1 , 1
# 1 , 0 , 0
# 0 , 0 , 1
# 와 같은 행렬일때 열길이 3, 행길이 3 반환

for i in range(1, ws.max_row):
    if 'A' in ws.cell(row=i, column=1).value:
        print(ws.cell(row=i, column=1).offset(row=1).value)

print('')
for row in ws.iter_rows():
    for entry in row:
        try:
            if 'B' in entry.value:
                # print(type(entry.offset))
                # print(dir(entry))
                print(entry.row)
                print(entry.column)
                print(entry.value)
                # print(entry.offset)
                # print(entry.offset(row=1).value)
        except (AttributeError, TypeError):
            continue





wb.save(filename="C:\\Users\\Jungly\\Desktop\\2.xlsx")

wb.close()
wb2.close()
