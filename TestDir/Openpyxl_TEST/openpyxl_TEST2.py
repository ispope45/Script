from openpyxl import Workbook, load_workbook

load_wb = load_workbook("D:\\\\Desktop\\434.xlsx")
load_ws = load_wb['2.납품및설치완료 확인서']

write_wb = Workbook()
write_ws = write_wb.active

load_ws.insert_rows(9)

load_wb.save("D:\\\\Desktop\\434.xlsx")