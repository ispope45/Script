import openpyxl
import os

CUR_PATH = os.getcwd()

if __name__ == "__main__":
    SRC_FILE = CUR_PATH + "\\ip_src.xlsx"
    wb = openpyxl.load_workbook(SRC_FILE)
    ws = wb.active

    res_wb = openpyxl.Workbook()
    res_ws = res_wb.active

    res_cnt = 1

    for row in range(2, ws.max_row + 1):
        schNo = ws[f'A{row}'].value
        schName = ws[f'B{row}'].value
        ipa = ws[f'D{row}'].value
        ipb = ws[f'E{row}'].value
        ipc = ws[f'F{row}'].value
        ipd = ws[f'G{row}'].value
        subnet = ws[f'H{row}'].value
        startc = ws[f'I{row}'].value
        endc = ws[f'J{row}'].value

        for i in range(startc, endc+1):
            res_ws[f'A{res_cnt}'].value = schNo
            res_ws[f'B{res_cnt}'].value = schName
            res_ws[f'C{res_cnt}'].value = f"{ipa}.{ipb}.{i}.0"
            res_cnt += 1

        res_wb.save(CUR_PATH + "\\ip_result.xlsx")
