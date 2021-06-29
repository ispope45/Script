from openpyxl import load_workbook, Workbook
import glob
'''
여러 엑셀파일의 시트안 특정값 추출하여 1개의 엑셀파일로 정리

'''
f_list = glob.glob("C:\\Users\\Jungle\\Desktop\\TEST3\\*")
aa = f_list[0].split('\\')
#var 선언
fileList = []
fileNum = []
fileName = []
i = 0
row = 0

for v in f_list:
    fileName = v.split('\\')
    #print(fileName[5])
    if fileName[5].find("~$") == -1:
        fileList.append(fileName[5])

for v1 in fileList:
    #print(v1)
    tmp = v1.split('_')
    fileNum.append(tmp[0])

write_wb = Workbook()
write_ws = write_wb.active

for vv in fileList:
    #print(vv)
    load_wb = load_workbook("C:\\Users\\Jungle\\Desktop\\TEST3\\" + str(vv), data_only=True)
    load_ws = load_wb['3.SK 설치확인서(추가)']
    load_ws2 = load_wb['2.설치 사진첩(공통)']

    for a in range(13, 20):
        row = row + 1
        print(fileNum[i])
        write_ws['A' + str(row)] = fileNum[i]
        write_ws['B' + str(row)] = load_ws2['D6'].value
        write_ws['C' + str(row)] = load_ws['A' + str(a)].value
        write_ws['D' + str(row)] = load_ws['F' + str(a)].value
        write_ws['E' + str(row)] = load_ws2['D9'].value
        print(str(row) + " " + load_ws2['D6'].value)

    i = i + 1

write_wb.save("C:\\Users\\Jungle\\Desktop\\test.xlsx")
