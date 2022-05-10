import requests
from urllib.request import urlopen
import re
import wget
import math
import openpyxl
import os


proxy = {'http': 'http://127.0.0.1:8080'}
CUR_PATH = os.getcwd()


def string_escape(s, encoding='utf-8'):
    return (s.encode('latin1')         # To bytes, required by 'unicode-escape'
             .decode('unicode-escape') # Perform the actual octal-escaping decode
             .encode('latin1')         # 1:1 mapping back to bytes
             .decode(encoding))        # Decode original encoding


def bar_custom(current, total, width=80):
    width = 30
    avail_dots = width-2
    shaded_dots = int(math.floor(float(current) / total * avail_dots))
    percent_bar = '[' + '■'*shaded_dots + ' '*(avail_dots-shaded_dots) + ']'
    progress = "%d%% %s [%d / %d]" % (current / total * 100, percent_bar, current, total)
    return progress


def download(url, out_path, error_cnt):
    try:
        wget.download(url, out=out_path, bar=bar_custom)
        print(" Ok")
        return int(error_cnt)
    except Exception as e:
        print(url + " Error : 404 Not Exist")
        # f = open(DST_PATH + "log.txt", "a+")
        # f.write(out_path + f" Error : 404 Not Exist : {e}\n")
        # f.close()
        return int(error_cnt + 1)


if __name__ == "__main__":
    mainUrl = "http://sen1.kti.co.kr"
    REPORT_API = "/inc_report.php"
    # download_API = "/download.php?fnm="

    SRC_FILE = CUR_PATH + "\\장비시리얼.xlsx"
    wb = openpyxl.load_workbook(SRC_FILE)
    ws = wb.active
    for row in range(2, ws.max_row + 1):
        schNo = ws[f'A{row}'].value
        devNo = ws[f'B{row}'].value

        switch_param = f"?cateidx=9&idx={str(devNo)}&schidx={str(schNo)}"
        # fw_param = f"?cateidx=11&idx={str(devNo)}&schidx={str(schNo)}"

        with requests.Session() as s:
            with s.get(mainUrl + REPORT_API + switch_param, verify=False, proxies=proxy) as res:
                resultBody = res.text
                find_num = resultBody.find("&nbsp;")
                if int(devNo) > 3679:
                    serial_num = resultBody[find_num - 13:find_num]
                else:
                    serial_num = resultBody[find_num - 12:find_num]

                ws[f'D{row}'].value = serial_num

        if row % 100 == 0:
            wb.save(SRC_FILE)

    wb.save(SRC_FILE)