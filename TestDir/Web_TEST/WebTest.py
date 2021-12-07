import requests
import time
import base64
import binascii
import hashlib
import openpyxl

from Crypto.Cipher import AES
from Crypto import Random
from bs4 import BeautifulSoup as bs
import json

MAIN_URL = 'https://192.168.10.10'
LOGIN_API = '/api/au/login'
HA_API = '/ha:443:ha_grp_2'
IF_API_BASE = '/api/sm/interfaces'
BACKUP_API = '/api/sm/backup/manual'
HOSTNAME_API = '/api/sm/info/equipment'

BAKCUP_PARAM = {"ha_backup": 1, "target": "POVS"}
proxy = {'https': 'http://127.0.0.1:8080'}
cookies = {'_next_login': 'index','_static': '88d4bcd01551a342f397603aace6b3dd74fcab1d77800188e0b0e8bf058bbdf3'}

SRC_FILE = "C:\\Users\\work.Jungly\\Desktop\\src.xlsx"

# cookie1 = {'_static': '88d4bcd01551a342f397603aace6b3dd74fcab1d77800188e0b0e8bf058bbdf3'}
# if1_config = {'eth1_m':
#                   {'url': ['/api/sm/interfaces/2/config/ipv4s/2', '/api/sm/interfaces/2/references',
#                                  '/api/sm/interfaces/2/config', '/api/sm/interfaces/2/config/apply'],
#                    'cfg1': {"ipaddress": "10.10.10.15", "netmask": 24, "primary": 1},
#                    'cfg2': {"name": "eth1", "zone": 1, "mode":
#                        {"mode": 1, "ipv4":
#                            {"type": 1, "address": [{"id": 2, "ipaddress": "10.10.10.15", "netmask": 24, "primary": 1, "_itemKey": "163814914152635", "itemIndex": 0}], "dynamic": {"request_dns": 0, "request_defaultgw": 0}, "pppoe": []}, "ipv6": {"type": 0, "dynamic": {"request_dns": 0, "accept_ra_defrtr": 0}, "address": []}, "l2": {"type": 1}, "mirror": [], "nac": {"ip4": "", "type": 0, "ipv4ipval": "", "ipv4maskval": None}}
#                             }
#                    },
#               'eth1_s': ''}


# IF1_IP = {"ipaddress":"10.10.10.20","netmask":24,"primary":1}
# IF1_CONFIG = {"name":"eth1","zone":1,"mode":{"mode":1,"ipv4":{"type":1,"address":[{"id":2,"ipaddress":"10.10.10.20","netmask":24,"primary":1,"_itemKey":"163814914152635","itemIndex":0}],"dynamic":{"request_dns":0,"request_defaultgw":0},"pppoe":[]},"ipv6":{"type":0,"dynamic":{"request_dns":0,"accept_ra_defrtr":0},"address":[]},"l2":{"type":1},"mirror":[],"nac":{"ip4":"","type":0,"ipv4ipval":"","ipv4maskval":None}}}


def make_hostname_config(hostname):
    cfg1 = {"hostname": hostname, "location": "KR", "desc": ""}

    return cfg1


def make_if_url(interface, ha):
    ifnum = str(interface + 1)
    if ha == "M":
        api1 = f'{IF_API_BASE}/{ifnum}/config/ipv4s/{ifnum}'
        api2 = f'{IF_API_BASE}/{ifnum}/references'
        api3 = f'{IF_API_BASE}/{ifnum}/config'
        api4 = f'{IF_API_BASE}/{ifnum}/config/apply'
    elif ha == "S":
        api1 = f'{HA_API + IF_API_BASE}/{ifnum}/config/ipv4s/{ifnum}'
        api2 = f'{HA_API + IF_API_BASE}/{ifnum}/references'
        api3 = f'{HA_API + IF_API_BASE}/{ifnum}/config'
        api4 = f'{HA_API + IF_API_BASE}/{ifnum}/config/apply'

    return api1, api2, api3, api4


# ### ip(str), netmaks(int), primary(0,1), interface(str), zone(1,2,3), idx(int), itemkey(str), itemIdx(int)
def make_if_config(ha, ip, netmask, if_num, primary, zone, itemKey, itemIdx):
    ifnum = str(if_num + 1)
    if ha == "M":
        api1 = f'{IF_API_BASE}/{ifnum}/config/ipv4s/'
        api2 = f'{IF_API_BASE}/{ifnum}/references'
        api3 = f'{IF_API_BASE}/{ifnum}/config'
        api4 = f'{IF_API_BASE}/{ifnum}/config/apply'
    elif ha == "S":
        api1 = f'{HA_API + IF_API_BASE}/{ifnum}/config/ipv4s/'
        api2 = f'{HA_API + IF_API_BASE}/{ifnum}/references'
        api3 = f'{HA_API + IF_API_BASE}/{ifnum}/config'
        api4 = f'{HA_API + IF_API_BASE}/{ifnum}/config/apply'

    if_name = 'eth' + str(if_num)
    cfg1 = {"ipaddress": ip, "netmask": netmask, "primary": primary}
    cfg2 = {"name": if_name,
            "zone": zone,
            "mode":
                {"mode": 1,
                 "ipv4": {"type": 1, "address": [{"id": 0, "ipaddress": ip, "netmask": netmask, "primary": primary, "_itemKey": itemKey, "itemIndex": itemIdx}],
                          "dynamic": {"request_dns": 0, "request_defaultgw": 0}, "pppoe": []},
                 "ipv6": {"type": 0, "dynamic": {"request_dns": 0, "accept_ra_defrtr": 0}, "address": []},
                 "l2": {"type": 1}, "mirror": [], "nac": {"ip4": "", "type": 0, "ipv4ipval": "", "ipv4maskval": None}
                 }
            }
    return api1, api2, api3, api4, cfg1, cfg2


# ### gw1: KT, gw2: SK
def make_routing_config(gw1, gw2):
    cfg1 = {"type": 0, "dest_addr": "0.0.0.0", "dest_mask": 0, "metric": 0,
            "gw": [{"id": "5452_60.60.60.20", "addr": gw1, "ifc_id": 9, "ifc_name": "eth8", "weight": 1, "itemIndex": 0},
                   {"id": "2597_70.70.70.20", "addr": gw2, "ifc_id": 8, "ifc_name": "eth7", "weight": 2, "ifauto": 0, "invalid": 1, "itemIndex": 1}],
            "gwTostring": "", "route_type": 0, "status_check": 1, "desc": ""}
    cfg2 = {"type": 0, "dest_addr": "0.0.0.0", "dest_mask": 0, "metric": 0,
            "gw": [{"id": "5452_60.60.60.20", "addr": gw1, "ifc_id": 9, "ifc_name": "eth8", "weight": 1, "itemIndex": 0}],
            "gwTostring": "", "route_type": 0, "status_check": 1, "desc": ""}
    return cfg1, cfg2


def make_dhcp_config(start, end, net, subnet, gw):
    cfg1 = {"area_index": 1, "start_area": start, "end_area": end, "net_addr": net,
            "subnet_mask": subnet, "default_gw": gw}
    cfg2 = {"normal_serv_use_enable": 1, "normal_serv_rent_tm": 1440, "normal_serv_auth": 0,
            "dns_serv1": "168.126.63.1", "dns_serv2": "8.8.8.8", "wins_serv1": "", "wins_serv2": ""}
    return cfg1, cfg2


def make_login_data(login_pw, csrf_token):
    login_data = {
        "lang": "ko",
        "force": "true",
        "readonly": 0,
        "login_id": "admin",
        "login_pw": login_pw,
        "csrf_token": csrf_token,
        "type": "local"
    }
    return login_data


def _pad(s):
    bs = 16
    return s + (bs - len(s) % bs) * chr(bs - len(s) % bs)


with requests.Session() as s:
    # Initialize
    with s.get(MAIN_URL, proxies=proxy, verify=False) as res:
        val = res.text
        find_num = val.find("csrf_token")
        csrf_token = val[find_num+51:find_num+115]
        pw = "secui00@!"
        iv = Random.new().read(AES.block_size)

        key = bytes(csrf_token[:32], 'utf-8')
        raw = bytes(_pad(pw), 'utf-8')
        cipher = AES.new(key, AES.MODE_CBC, iv)
        hex_data = (iv.hex() + cipher.encrypt(raw).hex()).encode('utf-8')
        login_pw = base64.b64encode(hex_data).decode('utf-8')

        login_data = make_login_data(login_pw, csrf_token)

    # Login Phase
    with s.post(MAIN_URL + LOGIN_API, json=login_data, proxies=proxy, verify=False) as res:
        print(res.cookies.get_dict())
        cookies = res.cookies.get_dict()
        print(res.headers)
        # print(type(cookies))
        # "_static = 88d4bcd01551a342f397603aace6b3dd74fcab1d77800188e0b0e8bf058bbdf3"
        print(cookies)
        val = res.json()
        # e_id = val['token']['id']
        print(res.text)
        res_dict = json.loads(res.text)
        print(res_dict['result']['api_token'])
        auth_key = res_dict['result']['api_token']
        headers = {'Authorization': auth_key}

    # Excel Data Formatting
    wb = openpyxl.load_workbook(SRC_FILE)
    ws = wb.active

    # A:No, B:Name, C:HA, D:hostname, E~R:IP/NETMASK, S~X:VIP, AA~AB:DHCP, AC~AD:GATEWAY
    IP_COL = ['E', 'G', 'I', 'K', 'M', 'O', 'Q']
    NETMASK_COL = ['F', 'H', 'J', 'L', 'N', 'P', 'R']
    IF_NUM = [1, 2, 3, 9, 9, 11, 12]
    if_config = []
    for row in range(2, 3):  # ws.max_row + 1
        if_config.clear()
        if ws[f'C{str(row)}'].value == "S":
            continue

        # Hostname Setting
        host_cfg1 = make_hostname_config(ws[f'D{str(row)}'].value)
        with s.put(MAIN_URL + HOSTNAME_API, json=host_cfg1, proxies=proxy, headers=headers, verify=False) as res:
            print(res.text)

        host_cfg2 = make_hostname_config(ws[f'D{str(row+1)}'].value)
        with s.put(MAIN_URL + HA_API + HOSTNAME_API, json=host_cfg2, proxies=proxy, headers=headers, verify=False) as res:
            print(res.text)

        # Interface Setting
        # for col in range(0, len(IP_COL)):
        #     if IP_COL[col] == 'M':
        #         primary = 0
        #         idx = 1
        #     else:
        #         primary = 1
        #         idx = 0
        #
        #     if IP_COL[col] in ['O', 'Q']:
        #         zone = 2
        #     else:
        #         zone = 1
        #
        #     if_config.append([
        #         ws[f'C{str(row)}'].value,
        #         ws[IP_COL[col] + str(row)].value,
        #         ws[NETMASK_COL[col] + str(row)].value,
        #         IF_NUM[col],
        #         primary,
        #         zone,
        #         idx]
        #     )
        #
        #     if IP_COL[col] != 'O' :
        #         if_config.append([
        #             ws[f'C{str(row + 1)}'].value,
        #             ws[IP_COL[col] + str(row + 1)].value,
        #             ws[NETMASK_COL[col] + str(row + 1)].value,
        #             IF_NUM[col],
        #             primary,
        #             zone,
        #             idx]
        #         )
        #
        # print(if_config)
        # for val in if_config:
        #     API1, API2, API3, API4, if_cfg1, if_cfg2 = \
        #         make_if_config(val[0], val[1], val[2], val[3], val[4], val[5], "163814914152635", val[6])
        #     with s.get(MAIN_URL + API3, proxies=proxy, headers=headers, verify=False) as res:
        #         print(res.text)
        #         var = res.json()
        #         print(type(if_cfg2['mode']['ipv4']['address'][0]['id']))
        #         print(if_cfg2['mode']['ipv4']['address'][0]['id'])
        #         if len(var['result']['mode']['ipv4']['address']) < 2:
        #             if_cfg2['mode']['ipv4']['address'][0]['id'] = var['result']['mode']['ipv4']['address'][0]['id']
        #             API1 += str(var['result']['mode']['ipv4']['address'][0]['id'])
        #         else:
        #             if_cfg2['mode']['ipv4']['address'][0]['id'] = var['result']['mode']['ipv4']['address'][1]['id']
        #             API1 += str(var['result']['mode']['ipv4']['address'][1]['id'])
        #
        #         print(type(idx))
        #
        #     with s.put(MAIN_URL + API1, json=if_cfg1, proxies=proxy, headers=headers, verify=False) as res:
        #         print(res.text)
        #         # val = res.json()
        #         # print(type(if_cfg2['mode']['ipv4']['address'][0]['id']))
        #         # print(if_cfg2['mode']['ipv4']['address'][0]['id'])
        #         # if_cfg2['mode']['ipv4']['address'][0]['id'] = val['result']['id']
        #         # print(type(idx))
        #
        #     with s.get(MAIN_URL + API2, proxies=proxy, headers=headers, verify=False) as res:
        #         print(res.text)
        #
        #     with s.put(MAIN_URL + API3, json=if_cfg2, proxies=proxy, headers=headers, verify=False) as res:
        #         print(res.text)
        #
        #     with s.put(MAIN_URL + API4, proxies=proxy, headers=headers, verify=False) as res:
        #         print(res.text)

        # Virtual IP Setting

        # DHCP Setting

        # Routing Setting

        # # Backup File Download
        # with s.post(MAIN_URL + "/api/sm/backup/manual", json=BAKCUP_PARAM, proxies=proxy, headers=headers, verify=False) as res:
        #     # print(res.text)
        #     print(res.headers)
        #     f = open("C:\\Users\\work.Jungly\\Desktop\\test.tar", 'wb',)
        #     f.write(res.content)
        #     f.close()
    s.close()

