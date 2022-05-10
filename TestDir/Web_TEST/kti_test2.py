import requests
from urllib.request import urlopen
import re
import wget
import math
import openpyxl
import os


proxy = {'http': 'http://127.0.0.1:8080'}
CUR_PATH = os.getcwd()


def string_escape(s, encoding='utf-8'):
    return (s.encode('latin1')         # To bytes, required by 'unicode-escape'
             .decode('unicode-escape') # Perform the actual octal-escaping decode
             .encode('latin1')         # 1:1 mapping back to bytes
             .decode(encoding))        # Decode original encoding


def bar_custom(current, total, width=80):
    width = 30
    avail_dots = width-2
    shaded_dots = int(math.floor(float(current) / total * avail_dots))
    percent_bar = '[' + '■'*shaded_dots + ' '*(avail_dots-shaded_dots) + ']'
    progress = "%d%% %s [%d / %d]" % (current / total * 100, percent_bar, current, total)
    return progress


def download(url, out_path, error_cnt):
    try:
        wget.download(url, out=out_path, bar=bar_custom)
        print(" Ok")
        return int(error_cnt)
    except Exception as e:
        print(url + " Error : 404 Not Exist")
        # f = open(DST_PATH + "log.txt", "a+")
        # f.write(out_path + f" Error : 404 Not Exist : {e}\n")
        # f.close()
        return int(error_cnt + 1)


if __name__ == "__main__":
    mainUrl = "http://sen1.kti.co.kr"
    speed_API = "/inc_report.php"
    # download_API = "/download.php?fnm="

    SRC_FILE = CUR_PATH + "\\report_2.xlsx"
    wb = openpyxl.load_workbook(SRC_FILE)
    ws = wb.active
    for row in range(2, ws.max_row + 1):
        schNo = ws[f'A{row}'].value
        schName = ws[f'B{row}'].value
        devIdx = ws[f'D{row}'].value
        speed_param = f"?cateidx=11&idx={str(devIdx)}&schidx={str(schNo)}"

        with requests.Session() as s:
            with s.get(mainUrl + speed_API + speed_param, verify=False) as res:
                resultBody = res.text
                # print(res.text)
                # print(len(res.text))
                # print(res.text.findall("src"))
                key1 = [i.start() for i in re.finditer("src", res.text)]
                key2 = [i.start() for i in re.finditer("class=\"embed-responsive\">", res.text)]
                try:
                    fileUrl = "/" + resultBody[key1[1] + 5:key2[1] - 2]
                    ext = fileUrl.split(".")[1]
                    download(mainUrl + fileUrl, f"{schNo}_{schName}." + ext, 0)
                except Exception as e:
                    print(e)
                #
                # for i in range(0, len(key1)):
                #     # print(resultBody[key1[i]+5:key2[i]-2])
                #     fileUrl = "/" + resultBody[key1[i] + 5:key2[i] - 2]
                #     ext = fileUrl.split(".")[1]
                #     download(mainUrl + fileUrl, f"{schNo}_{schName}." + ext, 0)
