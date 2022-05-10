import requests
import os, sys
import openpyxl
import time
from datetime import date

proxy = {'http': 'http://127.0.0.1:8080'}
CUR_PATH = os.getcwd()
START_DATE = date.today()


def string_escape(s, encoding='utf-8'):
    return (s.encode('latin1')         # To bytes, required by 'unicode-escape'
             .decode('unicode-escape') # Perform the actual octal-escaping decode
             .encode('latin1')         # 1:1 mapping back to bytes
             .decode(encoding))        # Decode original encoding


def write_log(string):
    dat = str(START_DATE).replace("-", "")
    f = open(CUR_PATH + f'\\log_{dat}.txt', "a+")
    now = time.localtime()
    cur_time = "%02d/%02d,%02d:%02d:%02d" % (now.tm_mon, now.tm_mday, now.tm_hour, now.tm_min, now.tm_sec)
    f.write(f'{cur_time};{string}\n')
    f.close()


if __name__ == "__main__":
    mainUrl = "http://sen1.kti.co.kr"
    report_API = "/action_report_print.php"
    report2_API = "/action_report_print1.php"
    download_API = "/download.php?fnm="

    SRC_FILE = CUR_PATH + "\\src_20220331.xlsx"
    wb = openpyxl.load_workbook(SRC_FILE)
    ws = wb.active

    for row in range(2, ws.max_row + 1):
        KTI_no = ws[f'A{row}'].value
        upload_val = ws[f'K{row}'].value
        download_val = ws[f'J{row}'].value
        speed_val = ws[f'M{row}'].value
        line_val = ws[f'N{row}'].value

        result_filename = ws[f'P{row}'].value

        report_param = {'schidx': KTI_no, 'knd': 'btnPrintDoc'}

        with requests.Session() as s:
            with s.post(mainUrl + report_API, data=report_param, verify=False) as res:
                fileName = string_escape(res.text, 'utf-8')
            with s.get(mainUrl + download_API + fileName, verify=False) as res:
                filePath = result_filename + "_설치확인서.doc"
                f = open(filePath, 'wb')
                f.write(res.text.encode('utf-8'))
                f.close()

            with s.post(mainUrl + report2_API, data=report_param, verify=False) as res:
                fileName2 = string_escape(res.text, 'utf-8')

            with s.get(mainUrl + download_API + fileName2, verify=False) as res:
                filePath2 = result_filename + "_개통확인서.doc"
                f = open(filePath2, 'wb')
                f.write(res.text.encode('utf-8'))
                f.close()

            new_text_content = ''
            new_string = ''

            target_word1 = '></SPAN>'
            new_word1 = f'>{str(speed_val)}</SPAN>'

            target_word2 = '></SPAN>'
            new_word2 = f'>{str(line_val)}</SPAN>'

            target_word3 = '>M</SPAN>'
            new_word3 = f'>{str(upload_val)}M</SPAN>'

            target_word4 = '>M</SPAN>'
            new_word4 = f'>{str(download_val)}M</SPAN>'

            with open(filePath2, 'r', encoding='utf8') as f:
                lines = f.readlines()
                for i, l in enumerate(lines):
                    # print(i)
                    if i == 148:
                        new_string = l.strip().replace(target_word1, new_word1)
                        new_text_content += new_string
                        write_log(f"{filePath2};{new_string}")
                    elif i == 154:
                        new_string = l.strip().replace(target_word2, new_word2)
                        new_text_content += new_string
                        write_log(f"{filePath2};{new_string}")
                    elif i == 351:
                        new_string = l.strip().replace(target_word3, new_word3)
                        new_text_content += new_string
                        write_log(f"{filePath2};{new_string}")
                    elif i == 362:
                        new_string = l.strip().replace(target_word4, new_word4)
                        new_text_content += new_string
                        write_log(f"{filePath2};{new_string}")
                    else:
                        new_string = l
                        if new_string:
                            new_text_content += new_string
                    # print(l)
                    # new_string = l.strip().replace(target_word, new_word)

            with open(filePath2, 'wb') as f:
                res_text = bytes(new_text_content, 'utf-8')
                f.write(res_text)

            s.close()
